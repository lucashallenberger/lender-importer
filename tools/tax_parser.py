#!/usr/bin/env python3
"""
APN Tax Bill Tool — web version (Streamlit)
Upload an LA County secured property tax bill PDF, review the extracted
data, and download a formatted Excel matching the standard template.

Core parsing and Excel-building logic is unchanged from the CLI version.
"""

import re
import io
import base64
import tempfile
from pathlib import Path

import streamlit as st

from tools import hist_llm   # shared Claude plumbing (api_key/_client/MODEL/_json_response)

# All intermediate files (screenshot, xlsx) go to a per-session temp folder.
OUTPUT_DIR = Path(tempfile.gettempdir()) / "apn_tax_tool"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ──────────────────────────────────────────────────────────────────────────
#  OPTIONAL OCR  (for scanned bills with no text layer; degrades gracefully)
# ──────────────────────────────────────────────────────────────────────────

def ocr_available() -> bool:
    """True only if pytesseract AND the tesseract binary are both usable."""
    try:
        import pytesseract
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def ocr_pdf_text(pdf_path) -> str:
    """OCR every page of a scanned PDF. Returns '' if OCR isn't available so the
    caller can show a friendly 'looks scanned' message instead of crashing."""
    if not ocr_available():
        return ""
    try:
        import io as _io
        import fitz  # pymupdf (already a dependency)
        import pytesseract
        from PIL import Image
        out = []
        doc = fitz.open(str(pdf_path))
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(3, 3))  # ~216 dpi
            out.append(pytesseract.image_to_string(Image.open(_io.BytesIO(pix.tobytes("png")))))
        doc.close()
        return "\n".join(out)
    except Exception:
        return ""


# ──────────────────────────────────────────────────────────────────────────
#  PARSING + EXCEL LOGIC  (ported verbatim from apn_tax_agent_28.py)
# ──────────────────────────────────────────────────────────────────────────

def parse_pdf(pdf_path: Path) -> dict:
    """
    Extract: Agency+Rate, Direct Assessments+Amount, Taxable Value from the PDF.
    Returns structured dict matching the Example.xlsx layout.
    """
    import pdfplumber

    data = {
        "apn": None,
        "mill_rates": [],          # list of (agency, rate)
        "direct_assessments": [],  # list of (name, amount)
        "taxable_value": {
            "land": None,
            "improvements": None,
            "pers_property": None,
        },
        "property_tax_hardcoded": None,
        "tax_year": None,
    }

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # Scanned bill (no text layer)? Try optional OCR.
    scanned_no_text = False
    if not full_text.strip():
        full_text = ocr_pdf_text(pdf_path)
        scanned_no_text = not full_text.strip()  # still nothing -> OCR unavailable/failed

    lines = [l.strip() for l in full_text.splitlines() if l.strip()]

    # ── Extract APN ──────────────────────────────────────────────────────
    for line in lines:
        # Match formats: 2558-022-010 or 2558 022 010
        m = re.search(r'(\d{4}[-\s]\d{3}[-\s]\d{3})', line)
        if m:
            data["apn"] = m.group(1)
            break

    # ── Extract Tax (Fiscal) Year ────────────────────────────────────────
    # e.g. "FISCAL YEAR JULY 1, 2024 ...", "2024-25", "2025-2026", "2024 ANNUAL"
    m = (re.search(r'FISCAL\s+YEAR\s+JULY\s+1,?\s+(20[123]\d)', full_text, re.IGNORECASE)
         or re.search(r'\b(20[123]\d)\s*[-–/]\s*(?:20)?\d{2}\b', full_text)
         or re.search(r'\b(20[123]\d)\s+ANNUAL', full_text, re.IGNORECASE))
    if m:
        data["tax_year"] = int(m.group(1))

    # ── Sections: find bounds ─────────────────────────────────────────────
    section_keywords = {
        "GENERAL TAX LEVY": "mill_rate_start",
        "MILL RATE": "mill_rate_start",
        "DIRECT ASSESSMENT": "direct_start",
        "TAXABLE VALUE": "taxable_start",
        "SUMMARY": "summary_start",
    }

    idx_map = {}
    for i, line in enumerate(lines):
        for kw, tag in section_keywords.items():
            if kw in line.upper() and tag not in idx_map:
                idx_map[tag] = i

    # ── Mill Rates (Agency + Rate) ────────────────────────────────────────
    # Known agency name fragments to match against (case-insensitive)
    KNOWN_AGENCIES = [
        "ALL AGENCIES", "CITY-LOS ANGELES", "CITY LOS ANGELES", "LA STORMWATER", "LA STORM WATER",
        "METRO WATER DIST", "METROPOLITAN WATER", "COMMNTY COLLEGE",
        "COMMUNITY COLLEGE", "UNIFIED SCHOOLS", "UNIFIED SCHOOL",
        "SCHOOL DISTRICT", "COUNTY", "FIRE", "LIBRARY", "WATER DIST",
        "SANITATION", "FLOOD CONTROL", "HARBOR", "AIRPORT",
    ]

    start = idx_map.get("mill_rate_start", 0)
    end   = idx_map.get("direct_start",    len(lines))
    for line in lines[start+1:end]:
        # Skip pure header/label lines
        upper = line.upper()
        if any(kw in upper for kw in ['AGENCY PHONE', 'AGENCY  PHONE', 'RATE', 'GENERAL TAX LEVY', 'VOTED INDEBTEDNESS']):
            continue

        # The rate is always a small decimal like 1.000000 or .012232
        # Match it at the end of the line (with optional $ amount after)
        rate_match = re.search(r'(?<![\d])(\d{0,2}\.\d{3,8})(?![\d])', line)
        if not rate_match:
            continue
        try:
            rate = float(rate_match.group(1))
        except ValueError:
            continue
        if rate <= 0 or rate > 10:  # sanity check - mill rates are never > 10
            continue

        # Extract agency name: everything before the rate
        agency_raw = line[:rate_match.start()].strip()

        # Remove phone numbers
        agency_raw = re.sub(r'\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}', '', agency_raw).strip()

        # Remove leading numeric junk (IDs, codes like "PDFEC 001 1234--")
        agency_raw = re.sub(r'^[\d\s\-–]+', '', agency_raw).strip()
        agency_raw = re.sub(r'\w{3,6}\s+\d+\s+\d+[-–]+\s*\d*\s*', '', agency_raw).strip()
        agency_raw = re.sub(r'^\W+', '', agency_raw).strip()

        # If multiple words remain, try to find a known agency name within
        matched_agency = None
        for known in KNOWN_AGENCIES:
            if known in agency_raw.upper():
                # Find the position and take from there
                idx = agency_raw.upper().find(known)
                matched_agency = agency_raw[idx:idx+len(known)].title()
                break

        if matched_agency:
            agency = matched_agency
        else:
            # Fall back: take the last run of alpha words (strip leading junk)
            words = re.findall(r'[A-Za-z][A-Za-z\s\-/&\.]{1,30}', agency_raw)
            agency = words[-1].strip() if words else agency_raw

        agency = agency.strip()
        if agency:
            data["mill_rates"].append((agency, rate))

    # ── Direct Assessments ────────────────────────────────────────────────
    start = idx_map.get("direct_start", 0)
    end   = idx_map.get("taxable_start", len(lines))
    for line in lines[start+1:end]:
        # Amount is always at end of line: e.g. "132.27", ".93", or "4,884.14"
        amount_match = re.search(r'\$?\s*([\d,]*\.\d{2})\s*$', line)
        if not amount_match:
            continue
        try:
            amount = float(amount_match.group(1).replace(',', ''))
        except ValueError:
            continue

        # Everything before the amount
        before_amount = line[:amount_match.start()].strip()

        # Strip phone number e.g. "(213) 485-4094"
        before_amount = re.sub(r'\(?\d{3}\)?\s*\d{3}-\d{4}', '', before_amount).strip()

        # Strip leading "$" marker
        before_amount = re.sub(r'\$\s*$', '', before_amount).strip()

        # Strip leading junk text that got merged by PDF (e.g. "SaveMoney–SaveTime–PayOnline")
        # Real agency names are ALL CAPS — find the last run of uppercase words
        parts = re.findall(r'[A-Z][A-Z\s/&\.\-]{2,}', before_amount)
        if parts:
            name = parts[-1].strip()
        else:
            name = re.sub(r'^[^A-Za-z]+', '', before_amount).strip()

        name = name.strip()
        if name and amount > 0:
            data["direct_assessments"].append((name, amount))

    # ── Taxable Value ─────────────────────────────────────────────────────
    # Search up to 25 lines after the TAXABLE VALUE header; if any field is
    # still missing after that window, fall back to a full-text scan so that
    # spaced-out or late-appearing lines (e.g. IMPROVEMENTS) are never missed.
    tv_start = idx_map.get("taxable_start", 0)

    def _parse_tv_line(line, data):
        # Collapse spaced-out letters pdfplumber emits for bold/wide PDF fonts
        # e.g. "L A N D" -> "LAND", "I M P R O V E M E N T S" -> "IMPROVEMENTS"
        normalized = re.sub(r'L\s+A\s+N\s+D', 'LAND', line, flags=re.IGNORECASE)
        normalized = re.sub(r'I\s+M\s+P\s+R\s+O\s+V\s+E\s+M\s+E\s+N\s+T\s+S',
                            'IMPROVEMENTS', normalized, flags=re.IGNORECASE)
        label = normalized.lower()
        is_land = bool(re.search(r'\bland\b', label))
        is_impr = bool(re.search(r'\bimprov', label))
        is_pers = bool(re.search(r'\bpers\b', label))
        if not (is_land or is_impr or is_pers):
            return
        # The PDF has two value columns (Current Assessed / Taxable) which are
        # identical. The line may also contain address/lot numbers (e.g. "LOT 16").
        # Property values are always comma-formatted (>=1,000); filter on that first,
        # then deduplicate and take the leftmost unique value.
        raw_numbers = re.findall(r'[\d,]+', normalized)
        seen, unique = set(), []
        for num in raw_numbers:
            if ',' not in num:          # skip bare integers like lot/address numbers
                continue
            try:
                v = int(num.replace(",", ""))
                if v > 0 and v not in seen:
                    seen.add(v)
                    unique.append(v)
            except ValueError:
                pass
        value = unique[0] if unique else None
        if is_land and data["taxable_value"]["land"] is None:
            data["taxable_value"]["land"] = value
        elif is_impr and data["taxable_value"]["improvements"] is None:
            data["taxable_value"]["improvements"] = value
        elif is_pers and data["taxable_value"]["pers_property"] is None:
            data["taxable_value"]["pers_property"] = value

    # First pass: 25-line window starting at the section header
    for line in lines[tv_start:tv_start + 25]:
        _parse_tv_line(line, data)

    # Second pass (fallback): scan the entire document for any field still None
    if any(v is None for v in data["taxable_value"].values()):
        for line in lines:
            _parse_tv_line(line, data)
            if all(v is not None for v in data["taxable_value"].values()):
                break

    # ── Property Tax total ────────────────────────────────────────────────
    for line in lines:
        # Try existing pattern first
        m = re.search(r'(?:Total Tax|Property Tax)[^\d]*([\d,]+\.\d{2})', line, re.IGNORECASE)
        if m:
            try:
                data["property_tax_hardcoded"] = float(m.group(1).replace(",", ""))
            except ValueError:
                pass
        # Grab the annual total: line with 3 dollar amounts e.g. "$3,337.84 $3,337.83 $6,675.67"
        # The third (largest) figure is the combined annual total
        if not data["property_tax_hardcoded"]:
            amounts = re.findall(r'\$([\d,]+\.\d{2})', line)
            if len(amounts) == 3:
                try:
                    data["property_tax_hardcoded"] = float(amounts[2].replace(",", ""))
                except ValueError:
                    pass

    data["scanned_no_text"] = scanned_no_text
    return data


def pdf_to_screenshot(pdf_path: Path) -> Path:
    """Render first page of PDF to PNG."""
    png_path = OUTPUT_DIR / (pdf_path.stem + "_screenshot.png")

    # Try pymupdf (fitz) - most reliable cross-platform option
    try:
        import fitz  # pymupdf
        doc = fitz.open(str(pdf_path))
        page = doc[0]
        mat = fitz.Matrix(2.0, 2.0)  # 2x zoom = ~150dpi
        pix = page.get_pixmap(matrix=mat)
        pix.save(str(png_path))
        doc.close()
        print(f"[4a] Screenshot rendered via pymupdf")
        return png_path
    except ImportError:
        pass
    except Exception as e:
        print(f"[WARN] pymupdf failed: {e}")

    # Try pdf2image
    try:
        from pdf2image import convert_from_path
        pages = convert_from_path(str(pdf_path), dpi=150, first_page=1, last_page=1)
        pages[0].save(str(png_path))
        print(f"[4a] Screenshot rendered via pdf2image")
        return png_path
    except Exception as e:
        print(f"[WARN] pdf2image failed: {e}")

    print(f"[WARN] Could not render PDF screenshot - no image will be embedded")
    return None


# ──────────────────────────────────────────────────────────────────────────
#  OPTIONAL CLAUDE EXTRACTION  (reliability layer — mirrors the Historicals/
#  Rent Roll tools; activates only when ANTHROPIC_API_KEY is present)
# ──────────────────────────────────────────────────────────────────────────

_TAX_PROMPT = """This is an LA County (California) SECURED PROPERTY TAX BILL.
Transcribe faithfully — do not compute or invent anything.

Extract:
- apn: the Assessor's Parcel Number (e.g. "5208-020-009"), or null
- tax_year: the fiscal year START as a 4-digit year (for "FISCAL YEAR JULY 1,
  2024" or "2024-25" return 2024), or null
- mill_rates: every taxing agency in the general tax levy / voted-indebtedness
  rate table, each with its RATE exactly as printed (a small number such as
  1.000000 or 0.012232). Include an "All Agencies" line if one is shown.
- direct_assessments: every direct-assessment line with its dollar amount
- taxable_value: the assessed land, improvements, and personal-property values
  as integers (null for any not shown)
- property_tax_total: the total ANNUAL tax for the year (the combined total of
  both installments), or null

Use the exact printed labels. Do not merge or skip lines."""

_TAX_SCHEMA = {
    "type": "object",
    "properties": {
        "apn": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "tax_year": {"anyOf": [{"type": "integer"}, {"type": "null"}]},
        "mill_rates": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {"agency": {"type": "string"}, "rate": {"type": "number"}},
                "required": ["agency", "rate"], "additionalProperties": False,
            },
        },
        "direct_assessments": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {"name": {"type": "string"}, "amount": {"type": "number"}},
                "required": ["name", "amount"], "additionalProperties": False,
            },
        },
        "taxable_value": {
            "type": "object",
            "properties": {
                "land": {"anyOf": [{"type": "integer"}, {"type": "null"}]},
                "improvements": {"anyOf": [{"type": "integer"}, {"type": "null"}]},
                "pers_property": {"anyOf": [{"type": "integer"}, {"type": "null"}]},
            },
            "required": ["land", "improvements", "pers_property"], "additionalProperties": False,
        },
        "property_tax_total": {"anyOf": [{"type": "number"}, {"type": "null"}]},
    },
    "required": ["apn", "tax_year", "mill_rates", "direct_assessments",
                 "taxable_value", "property_tax_total"],
    "additionalProperties": False,
}


def ai_available() -> bool:
    return hist_llm.available()


def extract_tax_bill(pdf_bytes: bytes) -> dict:
    """Claude reads the bill PDF and returns the same dict shape as parse_pdf."""
    resp = hist_llm._client().messages.create(
        model=hist_llm.MODEL,
        max_tokens=16000,
        thinking={"type": "adaptive"},
        messages=[{
            "role": "user",
            "content": [
                {"type": "document",
                 "source": {"type": "base64", "media_type": "application/pdf",
                            "data": base64.standard_b64encode(pdf_bytes).decode()}},
                {"type": "text", "text": _TAX_PROMPT},
            ],
        }],
        output_config={"format": {"type": "json_schema", "schema": _TAX_SCHEMA}},
    )
    j = hist_llm._json_response(resp)
    tv = j.get("taxable_value") or {}
    return {
        "apn": j.get("apn"),
        "tax_year": j.get("tax_year"),
        "mill_rates": [(m["agency"], float(m["rate"])) for m in (j.get("mill_rates") or [])],
        "direct_assessments": [(d["name"], float(d["amount"])) for d in (j.get("direct_assessments") or [])],
        "taxable_value": {"land": tv.get("land"), "improvements": tv.get("improvements"),
                          "pers_property": tv.get("pers_property")},
        "property_tax_hardcoded": j.get("property_tax_total"),
    }


def _weak_tax(data: dict) -> bool:
    """True when the regex parse looks thin — trigger the Claude fallback."""
    tv = data.get("taxable_value") or {}
    return (len(data.get("mill_rates") or []) < 2
            or not data.get("direct_assessments")
            or tv.get("land") is None
            or data.get("property_tax_hardcoded") is None)


def _merge_tax(base: dict, ai: dict) -> dict:
    """Fill gaps in the regex parse with Claude's read; prefer the longer list."""
    out = dict(base)
    for k in ("apn", "tax_year", "property_tax_hardcoded"):
        if not out.get(k) and ai.get(k) is not None:
            out[k] = ai[k]
    if len(ai.get("mill_rates") or []) > len(out.get("mill_rates") or []):
        out["mill_rates"] = ai["mill_rates"]
    if len(ai.get("direct_assessments") or []) > len(out.get("direct_assessments") or []):
        out["direct_assessments"] = ai["direct_assessments"]
    tv = dict(out.get("taxable_value") or {})
    for k, v in (ai.get("taxable_value") or {}).items():
        if tv.get(k) is None and v is not None:
            tv[k] = v
    out["taxable_value"] = tv
    return out


def build_excel(data: dict, screenshot_path: Path | None, apn: str) -> Path:
    """
    Creates the Excel file matching the Example.xlsx layout.
    Layout (column A = labels, column C = values):
      Row 1:  MILL RATE header
      Row 2:  APN
      Row 3+: each agency + rate
      Row N:  Total (SUM formula)
      ...     DIRECT ASSESSMENTS
      ...     each name + amount
      ...     Total (SUM)
      ...     TAXABLE VALUE
      ...     Land / Improvements / Total
      Row X:  Property Tax formulas
      Screenshot embedded at right (col E)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "W - RE Taxes"

    bold        = Font(bold=True, name="Arial")
    bold_white  = Font(bold=True, name="Arial", color="FFFFFF")
    normal      = Font(name="Arial")
    thin        = Side(style="thin")
    border      = Border(top=thin, bottom=thin)
    dark_grey   = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

    header_rows = set()  # track which rows get dark grey treatment

    def cell(row, col, value, is_bold=False, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = bold if is_bold else normal
        if fmt:
            c.number_format = fmt
        return c

    def header_cell(row, col, value):
        """Dark grey background, white bold text."""
        header_rows.add(row)
        c = ws.cell(row=row, column=col, value=value)
        c.font = bold_white
        c.fill = dark_grey
        return c

    def apply_header_fill(row):
        """Fill columns A-E of a header row with dark grey."""
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.fill = dark_grey

    row = 1

    # ── MILL RATE section ────────────────────────────────────────────────
    # Format APN as XXXX-XXX-XXX (first 10 digits with dashes)
    _apn_raw = re.sub(r"[^0-9]", "", apn)[:10]
    apn_digits = f"{_apn_raw[:4]}-{_apn_raw[4:7]}-{_apn_raw[7:10]}" if len(_apn_raw) >= 10 else _apn_raw
    header_cell(row, 1, "MILL RATE"); apply_header_fill(row)
    apn_row = row
    c = ws.cell(row=row, column=2, value=apn_digits); c.font = bold_white; c.fill = dark_grey
    row += 1

    rate_start = row
    if data["mill_rates"]:
        for agency, rate in data["mill_rates"]:
            cell(row, 1, agency)
            cell(row, 2, rate / 100, fmt='0.00000000%')
            row += 1
    else:
        # Placeholder rows matching example
        for placeholder in [("All Agencies", 0.01), ("City-Los Angeles", 0.00012232),
                             ("Metro Water District", 0.00007), ("Community College", 0.00048543),
                             ("Unified Schools", 0.00119605)]:
            cell(row, 1, placeholder[0])
            cell(row, 2, placeholder[1] / 100, fmt='0.00000000%')
            row += 1

    rate_end = row - 1
    cell(row, 1, "Total", True)
    cell(row, 2, f"=SUM(B{rate_start}:B{rate_end})", True, fmt="0.00000000%"); total_rate_row = row; row += 3

    # ── DIRECT ASSESSMENTS section ────────────────────────────────────────
    header_cell(row, 1, "DIRECT ASSESSMENTS"); apply_header_fill(row)
    apn_ref2_row = row
    c = ws.cell(row=row, column=2, value=f"=B{apn_row}"); c.font = bold_white; c.fill = dark_grey
    row += 1

    da_start = row
    if data["direct_assessments"]:
        for name, amount in data["direct_assessments"]:
            cell(row, 1, name)
            cell(row, 2, amount, fmt='#,##0.00')
            row += 1
    else:
        for placeholder in [("Safe Clean Water", 1251.53), ("LACo Vectr Cntrl", 19.27),
                             ("Flood Control", 503.20), ("City Lt Maint", 2682.74),
                             ("LA Stormwater", 401.16), ("Lacity Park Dist", 1995.95),
                             ("Rposd Measure A", 2544.42), ("Downtown Ind", 65346.56),
                             ("Trauma/Emerg Srv", 8392.60)]:
            cell(row, 1, placeholder[0])
            cell(row, 2, placeholder[1], fmt='#,##0.00')
            row += 1

    da_end = row - 1
    cell(row, 1, "Total", True)
    cell(row, 2, f"=SUM(B{da_start}:B{da_end})", True); total_da_row = row; row += 3

    # ── TAXABLE VALUE section ─────────────────────────────────────────────
    header_cell(row, 1, "TAXABLE VALUE"); apply_header_fill(row)
    c = ws.cell(row=row, column=2, value=f"=B{apn_ref2_row}"); c.font = bold_white; c.fill = dark_grey
    row += 1

    land_val = data["taxable_value"].get("land")
    impr_val = data["taxable_value"].get("improvements")
    pers_val = data["taxable_value"].get("pers_property")

    land_row = row
    cell(row, 1, "Land")
    if land_val is not None:
        cell(row, 2, land_val, fmt='#,##0')
    row += 1
    impr_row = row
    cell(row, 1, "Improvements")
    if impr_val is not None:
        cell(row, 2, impr_val, fmt='#,##0')
    row += 1
    pers_row = row
    cell(row, 1, "Pers Property")
    if pers_val is not None:
        cell(row, 2, pers_val, fmt='#,##0')
    row += 1
    # Total row
    tv_total_row = row
    cell(tv_total_row, 1, "Total", True)
    cell(tv_total_row, 2, f"=SUM(B{land_row}:B{pers_row})", True, '#,##0')
    row = tv_total_row + 3

    # ── Property Tax rows ─────────────────────────────────────────────────
    cell(row, 1, "Property Tax - Per Formula")
    cell(row, 2, f"=(B{total_rate_row}*B{tv_total_row})+B{total_da_row}", fmt='#,##0.00')
    row += 1
    cell(row, 1, "Property Tax - Hardcoded")
    hc = data.get("property_tax_hardcoded")
    if hc is not None:
        cell(row, 2, hc, fmt='#,##0.00')

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    # Clear column C (no data should be there)
    for r in range(1, row + 5):
        ws.cell(row=r, column=3).value = None
        ws.cell(row=r, column=3).fill = PatternFill(fill_type=None)
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50

    # ── Embed screenshot ──────────────────────────────────────────────────
    if screenshot_path and screenshot_path.exists():
        try:
            img = XLImage(str(screenshot_path))
            img.width  = 480
            img.height = 640
            ws.add_image(img, "D1")
            print("[9] Screenshot embedded in Excel")
        except Exception as e:
            print(f"[WARN] Could not embed screenshot: {e}")

    safe_apn = re.sub(r'[^0-9A-Za-z]', '', apn)
    xlsx_path = OUTPUT_DIR / f"tax_bill_{safe_apn}.xlsx"
    # Try saving — if file is locked/open, try an alternate filename
    for attempt in range(3):
        try:
            wb.save(str(xlsx_path))
            print(f"[10] Excel saved: {xlsx_path}")
            return xlsx_path
        except PermissionError:
            if attempt < 2:
                import time as _t
                alt = xlsx_path.with_stem(xlsx_path.stem + f"_v{attempt+2}")
                print(f"[WARN] File locked, trying: {alt.name}")
                xlsx_path = alt
                _t.sleep(1)
            else:
                raise RuntimeError(
                    f"Cannot save Excel — file is open or locked.\n"
                    f"Please close any open Excel files in your output folder and try again."
                )



# ──────────────────────────────────────────────────────────────────────────
#  HISTORY STORE  (SQLite — persists on Replit's filesystem)
# ──────────────────────────────────────────────────────────────────────────

import json
import sqlite3
import datetime
import zipfile
import pandas as pd

DB_PATH = Path(__file__).parent / "data" / "tax_history.db"
DB_PATH.parent.mkdir(parents=True, exist_ok=True)


def _conn():
    c = sqlite3.connect(str(DB_PATH))
    c.row_factory = sqlite3.Row
    return c


def init_db():
    with _conn() as c:
        c.execute(
            """CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                apn TEXT,
                source_filename TEXT,
                created_at TEXT,
                data_json TEXT,
                xlsx BLOB,
                screenshot BLOB
            )"""
        )


def save_record(name, apn, source_filename, data, xlsx_bytes, shot_bytes):
    with _conn() as c:
        cur = c.execute(
            "INSERT INTO bills (name, apn, source_filename, created_at, data_json, xlsx, screenshot) "
            "VALUES (?,?,?,?,?,?,?)",
            (name, apn, source_filename,
             datetime.datetime.now().isoformat(sep=" ", timespec="seconds"),
             json.dumps(data), xlsx_bytes, shot_bytes),
        )
        return cur.lastrowid


def search_records(q=""):
    with _conn() as c:
        if q.strip():
            like = f"%{q.strip()}%"
            return c.execute(
                "SELECT id, name, apn, source_filename, created_at FROM bills "
                "WHERE name LIKE ? OR apn LIKE ? OR source_filename LIKE ? ORDER BY id DESC",
                (like, like, like),
            ).fetchall()
        return c.execute(
            "SELECT id, name, apn, source_filename, created_at FROM bills ORDER BY id DESC"
        ).fetchall()


def get_record(rid):
    with _conn() as c:
        return c.execute("SELECT * FROM bills WHERE id=?", (rid,)).fetchone()


def rename_record(rid, new_name):
    with _conn() as c:
        c.execute("UPDATE bills SET name=? WHERE id=?", (new_name, rid))


def delete_record(rid):
    with _conn() as c:
        c.execute("DELETE FROM bills WHERE id=?", (rid,))


init_db()


# ──────────────────────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────────────────────

def collect_edited(apn_in, mill_df, da_df, land, impr, pers, original):
    """Rebuild the data dict from on-screen (possibly edited) values."""
    return {
        "apn": apn_in or original.get("apn"),
        "mill_rates": [
            (str(a), float(r)) for a, r in mill_df.itertuples(index=False)
            if str(a).strip() and pd.notna(r)
        ],
        "direct_assessments": [
            (str(n), float(amt)) for n, amt in da_df.itertuples(index=False)
            if str(n).strip() and pd.notna(amt)
        ],
        "taxable_value": {
            "land": land or None,
            "improvements": impr or None,
            "pers_property": pers or None,
        },
        "property_tax_hardcoded": original.get("property_tax_hardcoded"),
        "tax_year": original.get("tax_year"),
    }


def make_xlsx(edited, shot_path):
    """Build the Excel and return (bytes, filename)."""
    xlsx_path = build_excel(edited, Path(shot_path) if shot_path else None,
                            edited["apn"] or "unknown")
    return Path(xlsx_path).read_bytes(), Path(xlsx_path).name


BILL_WIDTH = 6   # 5 columns (A–E) + 1 spacer, for side-by-side layout


def write_bill(ws, data, apn, c0=0, screenshot_path=None):
    """Write one bill's block into worksheet `ws` starting at column 1+c0.
    Same layout as build_excel (labels col, value col, screenshot), but offset so
    multiple bills sit side by side. Returns the last row used."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    bold = Font(bold=True, name="Arial"); boldw = Font(bold=True, name="Arial", color="FFFFFF")
    normal = Font(name="Arial"); dark = PatternFill("solid", fgColor="404040")
    LC, VC = 1 + c0, 2 + c0
    VL = get_column_letter(VC)

    def cell(r, rel, value, is_bold=False, fmt=None):
        c = ws.cell(r, rel + c0, value); c.font = bold if is_bold else normal
        if fmt:
            c.number_format = fmt
        return c

    def hcell(r, rel, value):
        c = ws.cell(r, rel + c0, value); c.font = boldw; c.fill = dark; return c

    def hfill(r):
        for col in range(1 + c0, 6 + c0):
            ws.cell(r, col).fill = dark

    row = 1
    raw = re.sub(r"[^0-9]", "", apn)[:10]
    apnd = f"{raw[:4]}-{raw[4:7]}-{raw[7:10]}" if len(raw) >= 10 else raw
    hcell(row, 1, "MILL RATE"); hfill(row); apn_row = row
    ws.cell(row, VC, apnd).font = boldw; ws.cell(row, VC).fill = dark; row += 1
    rs = row
    for agency, rate in (data["mill_rates"] or []):
        cell(row, 1, agency); cell(row, 2, rate / 100, fmt="0.00000000%"); row += 1
    re_ = row - 1
    cell(row, 1, "Total", True); cell(row, 2, f"=SUM({VL}{rs}:{VL}{re_})", True, "0.00000000%"); trate = row; row += 3

    hcell(row, 1, "DIRECT ASSESSMENTS"); hfill(row)
    ws.cell(row, VC, f"={VL}{apn_row}").font = boldw; ws.cell(row, VC).fill = dark; row += 1
    ds = row
    for name, amt in (data["direct_assessments"] or []):
        cell(row, 1, name); cell(row, 2, amt, fmt="#,##0.00"); row += 1
    de = row - 1
    cell(row, 1, "Total", True); cell(row, 2, f"=SUM({VL}{ds}:{VL}{de})", True); tda = row; row += 3

    hcell(row, 1, "TAXABLE VALUE"); hfill(row)
    ws.cell(row, VC, f"={VL}{apn_row}").font = boldw; ws.cell(row, VC).fill = dark; row += 1
    tv = data["taxable_value"]; lr = row
    cell(row, 1, "Land"); (tv.get("land") is not None) and cell(row, 2, tv["land"], fmt="#,##0"); row += 1
    cell(row, 1, "Improvements"); (tv.get("improvements") is not None) and cell(row, 2, tv["improvements"], fmt="#,##0"); row += 1
    pr = row
    cell(row, 1, "Pers Property"); (tv.get("pers_property") is not None) and cell(row, 2, tv["pers_property"], fmt="#,##0"); row += 1
    ttv = row; cell(row, 1, "Total", True); cell(row, 2, f"=SUM({VL}{lr}:{VL}{pr})", True, "#,##0"); row = ttv + 3

    cell(row, 1, "Property Tax - Per Formula")
    cell(row, 2, f"=({VL}{trate}*{VL}{ttv})+{VL}{tda}", fmt="#,##0.00"); row += 1
    cell(row, 1, "Property Tax - Hardcoded")
    hc = data.get("property_tax_hardcoded")
    if hc is not None:
        cell(row, 2, hc, fmt="#,##0.00")

    ws.column_dimensions[get_column_letter(LC)].width = 30
    ws.column_dimensions[get_column_letter(VC)].width = 18
    if screenshot_path and Path(screenshot_path).exists():
        try:
            img = XLImage(str(screenshot_path)); img.width = 460; img.height = 620
            ws.add_image(img, f"{get_column_letter(4 + c0)}1")
        except Exception:
            pass
    return row


def _bill_year(data, fname=""):
    """Tax year for a bill: parsed from the PDF, else from the filename
    (e.g. 'Property Tax Bill - 2025-2026.pdf' -> 2025, '14BrookTax22.pdf' -> 2022)."""
    y = data.get("tax_year")
    if y:
        return int(y)
    stem = Path(fname or "").stem
    m = re.search(r'(20[123]\d)', stem)
    if m:
        return int(m.group(1))
    m = re.search(r'(\d{2})\D*$', stem)
    if m and 10 <= int(m.group(1)) <= 49:
        return 2000 + int(m.group(1))
    return None


def _apn_fmt(apn):
    raw = re.sub(r"[^0-9]", "", apn or "")[:10]
    return f"{raw[:4]}-{raw[4:7]}-{raw[7:10]}" if len(raw) >= 10 else (raw or "APN?")


def _union(names_per_bill):
    """Union of names across bills, preserving first-seen order. Returns
    (ordered display names, per-bill lookup dicts keyed by normalized name)."""
    def norm(n):
        return re.sub(r'\s+', ' ', str(n).strip().upper())
    order, seen = [], set()
    lookups = []
    for pairs in names_per_bill:
        lut = {}
        for name, val in pairs:
            k = norm(name)
            lut[k] = val
            if k not in seen:
                seen.add(k); order.append((k, name))
        lookups.append(lut)
    return order, lookups


def _write_combined(ws, bills):
    """Aligned grid: shared labels in col A, one value column per bill, and a
    COMBINED column that SUMS across the bills with live formulas."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    bold = Font(bold=True, name="Arial"); boldw = Font(bold=True, name="Arial", color="FFFFFF")
    normal = Font(name="Arial"); dark = PatternFill("solid", fgColor="404040")
    n = len(bills)
    bcols = list(range(2, 2 + n))          # one value column per bill
    ccol = 2 + n                           # COMBINED column
    B0, BL = get_column_letter(bcols[0]), get_column_letter(bcols[-1])
    CL = get_column_letter(ccol)

    def header(row, title, with_apns=True):
        ws.cell(row, 1, title).font = boldw
        for c in range(1, ccol + 1):
            ws.cell(row, c).fill = dark
        if with_apns:
            for i, col in enumerate(bcols):
                ws.cell(row, col, _apn_fmt(bills[i][1])).font = boldw
        ws.cell(row, ccol, "COMBINED").font = boldw

    def sum_row(row, fmt):
        c = ws.cell(row, ccol, f"=SUM({B0}{row}:{BL}{row})")
        c.font = normal; c.number_format = fmt
        return c

    # ── Year labels (which bill is which) ────────────────────────────────
    row = 1
    ws.cell(row, 1, "TAX YEAR").font = boldw; ws.cell(row, 1).fill = dark
    for i, col in enumerate(bcols):
        y = bills[i][3]
        c = ws.cell(row, col, f"{y % 100:02d}" if y else "?")
        c.font = boldw; c.fill = dark
    c = ws.cell(row, ccol, "COMBINED"); c.font = boldw; c.fill = dark
    row = 3

    # ── MILL RATE ─────────────────────────────────────────────────────────
    header(row, "MILL RATE"); row += 1
    rate_lists = [[(a, r) for a, r in (d["mill_rates"] or [])] for d, _, _, _ in bills]
    order, luts = _union(rate_lists)
    rs = row
    for k, name in order:
        ws.cell(row, 1, name).font = normal
        present = [lut[k] for lut in luts if k in lut]
        for i, col in enumerate(bcols):
            if k in luts[i]:
                c = ws.cell(row, col, luts[i][k] / 100); c.number_format = '0.00000000%'; c.font = normal
        # rates don't add — show in COMBINED only when every bill agrees
        if len(present) == n and len({round(p, 10) for p in present}) == 1:
            c = ws.cell(row, ccol, present[0] / 100); c.number_format = '0.00000000%'; c.font = normal
        row += 1
    re_ = row - 1
    ws.cell(row, 1, "Total").font = bold
    for col in bcols + [ccol]:
        L = get_column_letter(col)
        c = ws.cell(row, col, f"=SUM({L}{rs}:{L}{re_})"); c.font = bold; c.number_format = '0.00000000%'
    trate = row; row += 3

    # ── DIRECT ASSESSMENTS (summed) ──────────────────────────────────────
    header(row, "DIRECT ASSESSMENTS"); row += 1
    da_lists = [[(nm, amt) for nm, amt in (d["direct_assessments"] or [])] for d, _, _, _ in bills]
    order, luts = _union(da_lists)
    ds = row
    for k, name in order:
        ws.cell(row, 1, name).font = normal
        for i, col in enumerate(bcols):
            if k in luts[i]:
                c = ws.cell(row, col, luts[i][k]); c.number_format = '#,##0.00'; c.font = normal
        sum_row(row, '#,##0.00')
        row += 1
    de = row - 1
    ws.cell(row, 1, "Total").font = bold
    for col in bcols + [ccol]:
        L = get_column_letter(col)
        c = ws.cell(row, col, f"=SUM({L}{ds}:{L}{de})"); c.font = bold; c.number_format = '#,##0.00'
    tda = row; row += 3

    # ── TAXABLE VALUE (summed) ───────────────────────────────────────────
    header(row, "TAXABLE VALUE"); row += 1
    tv_rows = []
    for label, key in (("Land", "land"), ("Improvements", "improvements"), ("Pers Property", "pers_property")):
        ws.cell(row, 1, label).font = normal
        for i, col in enumerate(bcols):
            v = bills[i][0]["taxable_value"].get(key)
            if v is not None:
                c = ws.cell(row, col, v); c.number_format = '#,##0'; c.font = normal
        sum_row(row, '#,##0')
        tv_rows.append(row); row += 1
    ws.cell(row, 1, "Total").font = bold
    for col in bcols + [ccol]:
        L = get_column_letter(col)
        c = ws.cell(row, col, f"=SUM({L}{tv_rows[0]}:{L}{tv_rows[-1]})"); c.font = bold; c.number_format = '#,##0'
    ttv = row; row += 3

    # ── PROPERTY TAX (summed) ────────────────────────────────────────────
    ws.cell(row, 1, "Property Tax - Per Formula").font = normal
    for col in bcols:
        L = get_column_letter(col)
        c = ws.cell(row, col, f"=({L}{trate}*{L}{ttv})+{L}{tda}"); c.number_format = '#,##0.00'; c.font = normal
    sum_row(row, '#,##0.00'); row += 1
    ws.cell(row, 1, "Property Tax - Hardcoded").font = normal
    for i, col in enumerate(bcols):
        hc = bills[i][0].get("property_tax_hardcoded")
        if hc is not None:
            c = ws.cell(row, col, hc); c.number_format = '#,##0.00'; c.font = normal
    sum_row(row, '#,##0.00')

    ws.column_dimensions['A'].width = 30
    for col in bcols + [ccol]:
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ── Bill images (one per bill, aligned under its column) ─────────────
    from openpyxl.drawing.image import Image as XLImage
    img_row = row + 3
    ws.cell(img_row - 1, 1, "BILL IMAGES").font = boldw
    ws.cell(img_row - 1, 1).fill = dark
    for i, col in enumerate(bcols):
        shot = bills[i][2]
        y = bills[i][3]
        lbl = ws.cell(img_row, col, f"20{y % 100:02d}" if y else "?")
        lbl.font = bold
        if shot and Path(shot).exists():
            try:
                img = XLImage(str(shot)); img.width = 300; img.height = 400
                ws.add_image(img, f"{get_column_letter(col)}{img_row + 1}")
            except Exception:
                pass


def build_tax_into(wb, bills, prefix="S - Tax ", combined_name="W - RE Taxes"):
    """Write per-bill sheets (year-named, oldest→newest) + a combined sheet into an
    existing workbook. Returns metadata for the NEWEST year-bearing bill so other
    sheets can link its annual tax: {'sheet', 'hardcoded_cell', 'performula_cell'}."""
    bills = sorted(bills, key=lambda b: (b[3] is None, b[3] or 0))
    used, newest = set(), None
    for i, (data, apn, shot, year) in enumerate(bills):
        title = f"{prefix}{year}" if year else f"{prefix}Bill {i + 1}"
        while title in used:            # guard against duplicate years
            title += "*"
        used.add(title)
        last = write_bill(wb.create_sheet(title), data, apn, c0=0, screenshot_path=shot)
        if year is not None or newest is None:
            newest = {"sheet": title,
                      "hardcoded_cell": f"'{title}'!B{last}",
                      "performula_cell": f"'{title}'!B{last - 1}"}
    if len(bills) > 1:
        _write_combined(wb.create_sheet(combined_name), bills)
    return newest


def build_combined_workbook(bills):
    """bills: list of (data, apn, shot_path, year). Returns .xlsx bytes with one
    sheet per bill (named by tax year, oldest→newest) plus a Combined sheet: bills
    in aligned columns and a COMBINED column that adds them up (live SUM formulas)."""
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    build_tax_into(wb, bills)
    if len(bills) == 1:                 # standalone tool always shows the combined tab
        _write_combined(wb["W - RE Taxes"] if "W - RE Taxes" in wb.sheetnames
                        else wb.create_sheet("W - RE Taxes"),
                        sorted(bills, key=lambda b: (b[3] is None, b[3] or 0)))
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────
#  UI
# ──────────────────────────────────────────────────────────────────────────



def render():
    """Streamlit UI for the tax-bill parser (called by the sidebar router)."""

    with st.sidebar:
        st.header("🏠 APN Tax Bill Tool")
        view = st.radio("View", ["Process bills", "History"], label_visibility="collapsed")
        st.caption("Saved bills persist on Replit's filesystem.")


    # ─── PROCESS TAB ───────────────────────────────────────────────────────────
    if view == "Process bills":
        st.subheader("Process bills")
        st.caption("Upload one or more LA County tax bill PDFs. Review and correct the "
                   "extracted values, then save each to history and/or download the Excel.")

        files = st.file_uploader("Tax bill PDFs", type=["pdf"],
                                 accept_multiple_files=True, label_visibility="collapsed")

        ai_on = ai_available()
        if ai_on:
            always_ai = st.checkbox(
                "🤖 Double-check every bill with AI — Claude reads the PDF directly "
                "(slower + a small cost per bill, but you won't need to hand-verify).",
                value=False)
            st.caption("AI assist: **on** — any bill the built-in parser reads weakly is "
                       "automatically re-read by Claude.")
        else:
            always_ai = False
            st.caption("AI assist: off (set ANTHROPIC_API_KEY in the app secrets to enable it).")

        parsed = st.session_state.setdefault("parsed", {})

        # Parse any newly added files
        if files:
            seen = set()
            for f in files:
                key = f"{f.name}:{f.size}:{int(always_ai)}"
                seen.add(key)
                if key not in parsed:
                    pdf_path = OUTPUT_DIR / f.name
                    pdf_bytes = bytes(f.getbuffer())
                    pdf_path.write_bytes(pdf_bytes)
                    with st.spinner(f"Reading {f.name}…"):
                        data = parse_pdf(pdf_path)
                        used_ai = False
                        if ai_on and (always_ai or _weak_tax(data)):
                            try:
                                data = _merge_tax(data, extract_tax_bill(pdf_bytes))
                                used_ai = True
                            except Exception as e:  # noqa: BLE001
                                st.warning(f"AI read failed for {f.name}: {e}")
                        shot = pdf_to_screenshot(pdf_path)
                    if used_ai:
                        st.info(f"🤖 {f.name}: read/verified by Claude — spot-check the values against the preview.")
                    elif data.get("scanned_no_text"):
                        st.warning(
                            f"“{f.name}” looks scanned (no text layer) and OCR isn't available, "
                            "so fields may be blank. Install tesseract for OCR, or upload a text-based PDF.")
                    parsed[key] = {"fname": f.name, "data": data,
                                   "shot": str(shot) if shot else None}
            # Drop files the user removed from the uploader
            for key in list(parsed):
                if key not in seen:
                    del parsed[key]

        if not parsed:
            st.info("Upload tax bill PDFs to begin.")
        else:
            current = {}  # key -> (edited dict, shot path, source filename)

            for key, entry in parsed.items():
                data, shot, fname = entry["data"], entry["shot"], entry["fname"]
                with st.expander(f"📄 {fname}  —  APN {data.get('apn') or '?'}",
                                 expanded=len(parsed) == 1):
                    col_l, col_r = st.columns([3, 2], gap="large")
                    with col_l:
                        n1, n2 = st.columns(2)
                        save_name = n1.text_input("Save as", key=f"name_{key}",
                                                  value=data.get("apn") or Path(fname).stem)
                        apn_in = n2.text_input("APN", key=f"apn_{key}",
                                               value=data.get("apn") or "")
                        hc = data.get("property_tax_hardcoded")
                        st.metric("Property Tax (bill total)",
                                  f"${hc:,.2f}" if hc else "—")

                        st.markdown("**Mill Rates**")
                        mill_df = st.data_editor(
                            pd.DataFrame(data["mill_rates"], columns=["Agency", "Rate"]),
                            num_rows="dynamic", use_container_width=True, key=f"mill_{key}")

                        st.markdown("**Direct Assessments**")
                        da_df = st.data_editor(
                            pd.DataFrame(data["direct_assessments"], columns=["Assessment", "Amount"]),
                            num_rows="dynamic", use_container_width=True, key=f"da_{key}")

                        st.markdown("**Taxable Value**")
                        t1, t2, t3 = st.columns(3)
                        land = t1.number_input("Land", value=int(data["taxable_value"].get("land") or 0),
                                               step=1, key=f"land_{key}")
                        impr = t2.number_input("Improvements", value=int(data["taxable_value"].get("improvements") or 0),
                                               step=1, key=f"impr_{key}")
                        pers = t3.number_input("Pers Property", value=int(data["taxable_value"].get("pers_property") or 0),
                                               step=1, key=f"pers_{key}")
                    with col_r:
                        st.markdown("**Bill preview**")
                        if shot and Path(shot).exists():
                            st.image(shot, use_container_width=True)
                        else:
                            st.info("No preview.")

                    edited = collect_edited(apn_in, mill_df, da_df, land, impr, pers, data)
                    current[key] = (edited, shot, fname, save_name)

                    b1, b2 = st.columns(2)
                    if b1.button("💾 Save to history", key=f"save_{key}", use_container_width=True):
                        xb, xname = make_xlsx(edited, shot)
                        shot_bytes = Path(shot).read_bytes() if shot and Path(shot).exists() else None
                        rid = save_record(save_name, edited["apn"], fname, edited, xb, shot_bytes)
                        st.session_state[f"saved_{key}"] = (rid, xb, xname)
                        st.success(f"Saved as “{save_name}” (#{rid}).")

                    if f"saved_{key}" in st.session_state:
                        _, xb, xname = st.session_state[f"saved_{key}"]
                        b2.download_button("⬇ Download Excel", data=xb, file_name=xname,
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key=f"dl_{key}", use_container_width=True)

            # Batch actions — ONE workbook: a sheet per bill + a side-by-side Combined sheet
            st.divider()
            if st.button("📊 Save ALL & build combined workbook", type="primary",
                         use_container_width=True):
                bills = []
                for key, (edited, shot, fname, save_name) in current.items():
                    shot_bytes = Path(shot).read_bytes() if shot and Path(shot).exists() else None
                    save_record(save_name, edited["apn"], fname, edited, make_xlsx(edited, shot)[0], shot_bytes)
                    bills.append((edited, edited["apn"] or "unknown", shot, _bill_year(edited, fname)))
                st.session_state["batch_wb"] = build_combined_workbook(bills)
                st.success(f"Built one workbook: {len(bills)} bill sheet(s) + a Combined sheet.")

            if "batch_wb" in st.session_state:
                st.download_button("⬇ Download combined workbook (.xlsx)",
                                   data=st.session_state["batch_wb"],
                                   file_name="tax_bills_combined.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)


    # ─── HISTORY TAB ───────────────────────────────────────────────────────────
    else:
        st.subheader("History")
        q = st.text_input("Search by name, APN, or filename", placeholder="e.g. 5419 or Sunset Blvd")
        rows = search_records(q)

        if not rows:
            st.info("No saved bills match." if q else "No saved bills yet. Process some first.")
        else:
            st.caption(f"{len(rows)} result(s)")
            for r in rows:
                with st.expander(f"**{r['name']}**  ·  APN {r['apn'] or '?'}  ·  {r['created_at']}"):
                    rec = get_record(r["id"])
                    col_l, col_r = st.columns([3, 2], gap="large")
                    with col_l:
                        st.text(f"Source file: {rec['source_filename']}")
                        data = json.loads(rec["data_json"])
                        if data.get("mill_rates"):
                            st.markdown("**Mill Rates**")
                            st.dataframe(pd.DataFrame(data["mill_rates"], columns=["Agency", "Rate"]),
                                         use_container_width=True, hide_index=True)
                        if data.get("direct_assessments"):
                            st.markdown("**Direct Assessments**")
                            st.dataframe(pd.DataFrame(data["direct_assessments"], columns=["Assessment", "Amount"]),
                                         use_container_width=True, hide_index=True)

                        new_name = st.text_input("Rename", value=rec["name"], key=f"rn_{r['id']}")
                        a1, a2, a3 = st.columns(3)
                        if a1.button("Rename", key=f"rnb_{r['id']}", use_container_width=True):
                            rename_record(r["id"], new_name)
                            st.rerun()
                        a2.download_button("⬇ Excel", data=rec["xlsx"],
                                           file_name=f"{rec['name']}.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key=f"hdl_{r['id']}", use_container_width=True)
                        if a3.button("🗑 Delete", key=f"del_{r['id']}", use_container_width=True):
                            delete_record(r["id"])
                            st.rerun()
                    with col_r:
                        if rec["screenshot"]:
                            st.image(rec["screenshot"], use_container_width=True)
