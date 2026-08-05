"""Public UW — strip a finished underwriting workbook down to the tabs that go out.

A QCP workbook is laid out left-to-right: the final tabs (Sizer, Property Info,
Summary, Rent Roll, Pro Forma, Historicals), then a `Worksheets>>>` divider, then
the W- worksheets, the S- sources, and a Not Used pile. Only the left side ships.

Deleting the right side would leave every formula that pointed at it as #REF!, so
each of those cells is first replaced with the value Excel last calculated for it.
Formulas that only reference tabs that are staying are left live, so the public
workbook still recalculates the way the model does:

    ='W - Rent Roll'!L5      ->  27,450.00      (its source is leaving)
    =Summary!B12*12          ->  =Summary!B12*12 (its source is staying)
    =SUM(C4:C20)             ->  =SUM(C4:C20)    (same sheet)

The workbook is edited in place rather than rebuilt, so formatting, column widths,
merges, images and print setup come through untouched.
"""

import io
import re

# A cell can only keep its formula if every sheet it names is staying.
_STRINGS = re.compile(r'"[^"]*"')
_ERRORS = re.compile(r"#(?:REF|DIV/0|N/A|VALUE|NAME\?|NUM|NULL|GETTING_DATA|SPILL|CALC)!?")
_SHEET_REF = re.compile(r"(?:'((?:[^']|'')+)'|([A-Za-z0-9_.][A-Za-z0-9_. ]*))!")
_ERR_VALUE = re.compile(r"^#[A-Z/0-9_]+[!?]$")

# tabs like "Worksheets>>>" / "Source >>>" / "Not Used>>>" mark the end of the
# public section; the W-/S- prefixes are the fallback when a divider is missing
_DIVIDER = re.compile(r">{2,}\s*$")
_BACKUP = re.compile(r"^(?:W|S|WL|SL|PP?)\s*-\s", re.I)


def sheet_refs(formula: str) -> set:
    """Sheet names a formula reads from. Error tokens (#REF!) and text inside
    quotes are stripped first so they can't masquerade as sheet names."""
    body = _ERRORS.sub("", _STRINGS.sub("", formula or ""))
    return {(m.group(1) or m.group(2)).replace("''", "'").strip()
            for m in _SHEET_REF.finditer(body)}


def split_index(names) -> int:
    """Where the public section ends — the first divider tab, else the first
    worksheet/source tab, else everything is public."""
    for i, n in enumerate(names):
        if _DIVIDER.search(str(n).strip()):
            return i
    for i, n in enumerate(names):
        if _BACKUP.match(str(n).strip()):
            return i
    return len(names)


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


# ── column surgery ───────────────────────────────────────────────────────────
# Deleting or inserting a column is Excel's job normally: it rewrites every
# formula in the workbook. openpyxl moves cell values and styles and nothing
# else — not formulas, merges, widths, panes, filters, print areas or
# conditional formats — so all of that is done here.
_COL = r"\$?[A-Za-z]{1,3}"
_ROW = r"\$?[1-9][0-9]{0,6}"
_CELL = _COL + _ROW
_PART = rf"(?:{_CELL}(?::{_CELL})?|{_COL}:{_COL}|{_ROW}:{_ROW})"
_SHEET_Q = r"'(?:[^']|'')+'"
_SHEET_B = r"[A-Za-z_][A-Za-z0-9_.]*"
_TOKEN = re.compile(
    rf'(?P<str>"(?:[^"]|"")*")'
    rf'|(?<![A-Za-z0-9_$.!])(?:(?P<sheet>{_SHEET_Q}|{_SHEET_B})!)?'
    rf'(?P<ref>{_PART})(?![A-Za-z0-9_(!])'
)
_ONLY_ROWS = re.compile(rf"^{_ROW}:{_ROW}$")


def _split_col(part):
    """'$AB12' -> ('$AB', 12-ish tail). Returns (dollars, letters, rest)."""
    m = re.match(r"^(\$?)([A-Za-z]{1,3})(.*)$", part)
    return (m.group(1), m.group(2).upper(), m.group(3)) if m else None


def _map_col(c, at, delta):
    """Where column index c lands. None means the reference is destroyed."""
    if delta < 0:
        return None if c == at else (c - 1 if c > at else c)
    return c + 1 if c >= at else c


def _shift_ref(ref, at, delta):
    """One A1 / A1:B2 / A:B reference, shifted. '#REF!' when it's destroyed."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    if _ONLY_ROWS.match(ref):
        return ref                                   # whole-row ref: no columns
    ends = ref.split(":")
    parsed = [_split_col(e) for e in ends]
    if any(p is None for p in parsed):
        return ref
    cols = [column_index_from_string(p[1]) for p in parsed]
    if delta < 0 and len(cols) == 2:                 # a range shrinks around the cut
        s, e = cols
        ns = s if s <= at else s - 1
        ne = e if e < at else e - 1
        if ne < ns:
            return "#REF!"
        new = [ns, ne]
    else:
        new = [_map_col(c, at, delta) for c in cols]
        if any(n is None for n in new):
            return "#REF!"
    return ":".join(f"{p[0]}{get_column_letter(n)}{p[2]}" for p, n in zip(parsed, new))


def _sheet_name(tok):
    return tok[1:-1].replace("''", "'") if tok.startswith("'") else tok


def shift_formula(formula, host_sheet, target_sheet, at, delta):
    """Rewrite the column references in `formula` that point at `target_sheet`.
    A reference with no sheet prefix belongs to `host_sheet`."""
    def sub(m):
        if m.group("str"):
            return m.group(0)
        sheet = _sheet_name(m.group("sheet")) if m.group("sheet") else host_sheet
        if sheet != target_sheet:
            return m.group(0)
        shifted = _shift_ref(m.group("ref"), at, delta)
        return (m.group("sheet") + "!" if m.group("sheet") else "") + shifted
    return _TOKEN.sub(sub, formula)


def _shift_range_str(s, at, delta):
    """Shift every reference in a plain range string (print area, filters, …)."""
    if not s:
        return s
    return _TOKEN.sub(
        lambda m: m.group(0) if m.group("str") else
        (m.group("sheet") + "!" if m.group("sheet") else "") + _shift_ref(m.group("ref"), at, delta),
        str(s))


def _shift_geometry(ws, merged, at, delta):
    """Everything on the sheet that isn't a cell: merges, widths, panes,
    filters, print area, conditional formats, validations, image anchors.
    `merged` are the ranges captured before the move."""
    from openpyxl.utils import get_column_letter

    for r in merged:
        lo, hi = _map_col(r.min_col, at, delta), _map_col(r.max_col, at, delta)
        if delta < 0:
            lo = r.min_col if r.min_col <= at else r.min_col - 1
            hi = r.max_col if r.max_col < at else r.max_col - 1
        if lo is None or hi is None or hi < lo:
            continue
        ws.merge_cells(start_row=r.min_row, start_column=lo,
                       end_row=r.max_row, end_column=hi)

    dims = {}
    for key, dim in list(ws.column_dimensions.items()):
        try:
            from openpyxl.utils import column_index_from_string
            idx = column_index_from_string(key)
        except Exception:  # noqa: BLE001
            continue
        new = _map_col(idx, at, delta)
        if new is None:
            continue
        dim.min = dim.max = new
        dims[get_column_letter(new)] = dim
    ws.column_dimensions.clear()
    ws.column_dimensions.update(dims)

    if ws.freeze_panes:
        ws.freeze_panes = _shift_range_str(ws.freeze_panes, at, delta) or None
    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = _shift_range_str(ws.auto_filter.ref, at, delta)
    if ws.print_area:
        ws.print_area = _shift_range_str(str(ws.print_area), at, delta)

    for rng in list(ws.conditional_formatting):
        rng.sqref = _shift_range_str(str(rng.sqref), at, delta)
    for dv in ws.data_validations.dataValidation:
        dv.sqref = _shift_range_str(str(dv.sqref), at, delta)
    for img in getattr(ws, "_images", []):
        anc = getattr(img, "anchor", None)
        for corner in ("_from", "to"):
            spot = getattr(anc, corner, None)
            if spot is not None and hasattr(spot, "col"):
                new = _map_col(spot.col + 1, at, delta)     # anchors are 0-based
                spot.col = max(0, (new or 1) - 1)


def move_column(wb, sheet_name, keep, at=1, delta=-1):
    """Delete (delta=-1) or insert (delta=+1) a column on one sheet, rewriting
    every reference to it across the kept sheets."""
    from openpyxl.worksheet.cell_range import CellRange
    ws = wb[sheet_name]
    for name in keep:
        other = wb[name]
        for row in other.iter_rows():
            for c in row:
                if _is_formula(c.value):
                    c.value = shift_formula(c.value, name, sheet_name, at, delta)
    # Take the merges apart first: a merged range leaves read-only placeholder
    # cells behind, and any value the move lands on one of them is dropped.
    merged = [CellRange(str(r)) for r in ws.merged_cells.ranges]
    for r in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(r))
    if delta < 0:
        ws.delete_cols(at)
    else:
        ws.insert_cols(at)
    _shift_geometry(ws, merged, at, delta)


# ── the left edge: one blank spacer column, no classification tags ───────────
# "OpEx - Insurance", "Income - Other Income - Laundry" … the underwriting
# buckets the Historicals tab carries in column A to drive its SUMIFs. They are
# working notes, not part of the deliverable.
_CATEGORY = re.compile(r"^\s*(income|opex|capex|non-?opex)\s*-\s", re.I)
SPACER_WIDTH = 3.9


def _col_a_values(wsv):
    return [v for v in (wsv.cell(r, 1).value for r in range(1, (wsv.max_row or 0) + 1))
            if v not in (None, "")]


def is_category_column(values) -> bool:
    """Column A is a classification column when what's in it is overwhelmingly
    underwriting buckets — not when it happens to hold one such label."""
    if len(values) < 3:
        return False
    tagged = sum(1 for v in values if isinstance(v, str) and _CATEGORY.match(v))
    return tagged >= max(3, int(0.8 * len(values)))


def column_a_in_use(wb, keep: list, sheet: str) -> bool:
    """Does any formula still in the workbook read column A of `sheet`? Run this
    after the freeze: the SUMIFs that drive a classification column point into the
    dropped tabs and are literals by then, so a hit here means something on the
    public side genuinely depends on it."""
    from openpyxl.utils import column_index_from_string
    for name in keep:
        for row in wb[name].iter_rows():
            for c in row:
                if not _is_formula(c.value):
                    continue
                for m in _TOKEN.finditer(c.value):
                    if m.group("str"):
                        continue
                    target = _sheet_name(m.group("sheet")) if m.group("sheet") else name
                    if target != sheet:
                        continue
                    ref = m.group("ref")
                    if _ONLY_ROWS.match(ref):
                        continue
                    parts = [_split_col(e) for e in ref.split(":")]
                    if any(pt is None for pt in parts):
                        continue
                    cols = [column_index_from_string(pt[1]) for pt in parts]
                    if min(cols) <= 1 <= max(cols):
                        return True
    return False


def tidy_columns(wb, wbv, keep: list) -> list:
    """Give every kept tab exactly one empty spacer column A: drop a leading
    classification column if there is one, then insert a spacer if column A
    still holds content. Returns a line per tab describing what happened."""
    log = []
    widths = [wb[n].column_dimensions["A"].width for n in keep
              if not _col_a_values(wbv[n]) and wb[n].column_dimensions["A"].width]
    spacer = round(sum(widths) / len(widths), 2) if widths else SPACER_WIDTH
    for name in keep:
        vals = _col_a_values(wbv[name])
        if vals and is_category_column(vals):
            if column_a_in_use(wb, keep, name):
                log.append(f"{name}: left the classification column in place — a "
                           "formula on the public tabs still reads it")
            else:
                move_column(wb, name, keep, at=1, delta=-1)
                log.append(f"{name}: removed the classification column ({len(vals)} tags)")
                vals = [v for v in (wb[name].cell(r, 1).value
                                    for r in range(1, (wb[name].max_row or 0) + 1))
                        if v not in (None, "")]
        if vals:
            move_column(wb, name, keep, at=1, delta=+1)
            wb[name].column_dimensions["A"].width = spacer
            log.append(f"{name}: inserted a blank spacer column")
        elif (wb[name].column_dimensions["A"].width or 0) > 3 * spacer:
            wb[name].column_dimensions["A"].width = spacer
            log.append(f"{name}: narrowed the empty column A to a spacer")
    return log


def plan(data: bytes, keep: list) -> dict:
    """What building would do, without doing it: per-sheet counts of formulas
    kept vs hardcoded, plus the things worth telling the user about."""
    import openpyxl
    wbf = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wbv = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    keep_set = set(keep)
    rows, errors, stale = [], [], 0
    for name in keep:
        wsf, wsv = wbf[name], wbv[name]
        n_f = n_hard = n_live = 0
        for row in wsf.iter_rows():
            for c in row:
                # an Excel error anywhere on a public tab ships with the file,
                # whether its formula is frozen or left live — surface all of them
                cached = wsv[c.coordinate].value
                if isinstance(cached, str) and _ERR_VALUE.match(cached.strip()):
                    errors.append(f"{name}!{c.coordinate} = {cached.strip()}")
                if not _is_formula(c.value):
                    continue
                n_f += 1
                if sheet_refs(c.value) - keep_set:
                    n_hard += 1
                    # a formula whose result was "" caches as None — that's a real
                    # blank, not a missing value; anything else means no saved result
                    if cached is None and '""' not in c.value:
                        stale += 1
                else:
                    n_live += 1
        rows.append({"sheet": name, "formulas": n_f, "hardcoded": n_hard, "kept": n_live})

    columns = []                                    # preview of the left-edge tidy
    for name in keep:
        vals = _col_a_values(wbv[name])
        if vals and is_category_column(vals):
            columns.append(f"{name}: classification column removed ({len(vals)} tags)")
        elif vals:
            columns.append(f"{name}: blank spacer column inserted")
    return {"sheets": rows, "errors": errors, "stale": stale, "columns": columns,
            "dropped": [n for n in wbf.sheetnames if n not in keep_set],
            "total_hardcoded": sum(r["hardcoded"] for r in rows),
            "total_kept": sum(r["kept"] for r in rows)}


def build(data: bytes, keep: list, keep_vba: bool = False) -> bytes:
    """The public workbook: kept tabs only, formulas into the dropped tabs frozen
    at their last calculated value."""
    import openpyxl
    from openpyxl.utils.exceptions import IllegalCharacterError
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False, keep_vba=keep_vba)
    wbv = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    keep_set = set(keep)
    if not keep_set:
        raise ValueError("no tabs selected — nothing to build")

    for name in keep:
        ws, wsv = wb[name], wbv[name]
        for row in ws.iter_rows():
            for c in row:
                if not _is_formula(c.value) or not (sheet_refs(c.value) - keep_set):
                    continue
                v = wsv[c.coordinate].value
                fmt = c.number_format
                try:
                    c.value = v
                except (IllegalCharacterError, ValueError):
                    c.value = str(v)
                # a cached string that happens to start with "=" is text, not a
                # new formula — openpyxl would otherwise re-arm it
                if isinstance(v, str) and v.startswith("="):
                    c.data_type = "s"
                c.number_format = fmt

    # only now that the cross-tab formulas are literals is the classification
    # column dead weight — until this point its SUMIFs were still live
    tidy_columns(wb, wbv, keep)

    for name in list(wb.sheetnames):
        if name not in keep_set:
            del wb[name]

    # defined names that pointed into the dropped tabs would open as #REF!
    for dn in list(getattr(wb.defined_names, "keys", lambda: [])()):
        ref = str(getattr(wb.defined_names[dn], "attr_text", "") or "")
        if (sheet_refs(ref) - keep_set) or "#REF" in ref:
            del wb.defined_names[dn]

    wb.calculation.fullCalcOnLoad = True    # kept formulas recalc when it opens
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ───────────────────────────────────────────────────────────────────────
def render():
    import pandas as pd
    import streamlit as st

    st.header("📤 Public UW")
    st.caption("Upload a finished underwriting workbook. You get back just the tabs "
               "that go out — everything from the `Worksheets>>>` divider rightward is "
               "removed, and any formula that pointed into those tabs is frozen at the "
               "value Excel last calculated. Formulas between the public tabs stay live.")

    up = st.file_uploader("Finished workbook (.xlsx / .xlsm)", type=["xlsx", "xlsm"],
                          key="puw_file")
    if not up:
        st.info("Upload the completed underwriting file to begin.")
        return

    import openpyxl
    data = up.getvalue()
    try:
        names = openpyxl.load_workbook(io.BytesIO(data), read_only=True).sheetnames
    except Exception as e:  # noqa: BLE001
        st.error(f"Could not open that workbook: {e}")
        return

    cut = split_index(names)
    if cut == 0:
        st.warning("The first tab already looks like a worksheet/source tab — pick the "
                   "public tabs by hand below.")
    elif cut == len(names):
        st.warning("No `Worksheets>>>` divider or W-/S- tabs found, so every tab reads "
                   "as public. Check the selection below before building.")

    st.subheader("1 · Tabs that go out")
    keep = st.multiselect(
        "Kept in the public workbook (left-to-right order is preserved)",
        options=names, default=names[:cut], key="puw_keep",
        help="Defaults to everything left of the first divider tab. Anything not "
             "selected is deleted, and formulas pointing into it get hardcoded.")
    keep = [n for n in names if n in keep]          # keep workbook order
    if not keep:
        st.info("Select at least one tab.")
        return
    dropped = [n for n in names if n not in keep]
    st.caption(f"Dropping {len(dropped)}: " + (", ".join(f"`{d}`" for d in dropped) or "—"))

    st.subheader("2 · What will change")
    with st.spinner("Reading formulas…"):
        try:
            p = plan(data, keep)
        except Exception as e:  # noqa: BLE001
            st.error(f"Could not analyse the workbook: {e}")
            return
    st.dataframe(pd.DataFrame(p["sheets"]).rename(columns={
        "sheet": "Tab", "formulas": "Formulas", "hardcoded": "→ hardcoded",
        "kept": "→ stay live"}), use_container_width=True, hide_index=True)
    c1, c2 = st.columns(2)
    c1.metric("Formulas hardcoded", p["total_hardcoded"])
    c2.metric("Formulas kept live", p["total_kept"])
    if p["columns"]:
        st.caption("Left edge — every tab ends up with one blank spacer column A:\n\n" +
                   "\n".join(f"- {c}" for c in p["columns"]))

    if p["stale"]:
        st.warning(f"{p['stale']} formula(s) pointing into the dropped tabs have no "
                   "saved result, so they'll come out blank. That happens when a "
                   "workbook was written by a script and never opened in Excel — "
                   "open it in Excel, save, and re-upload to capture the values.")
    if p["errors"]:
        st.warning(f"{len(p['errors'])} cell(s) on the public tabs already show an Excel "
                   "error (#REF!, #DIV/0!, …) in the file you uploaded. They carry "
                   "through as-is — worth fixing before this goes out.")
        with st.expander(f"Show the {len(p['errors'])} error cell(s)"):
            st.code("\n".join(p["errors"][:300]) +
                    ("\n…" if len(p["errors"]) > 300 else ""))

    st.subheader("3 · Build")
    macro = up.name.lower().endswith(".xlsm")
    if st.button("📤  Build public workbook", type="primary"):
        with st.spinner("Freezing values and stripping tabs…"):
            try:
                out = build(data, keep, keep_vba=macro)
            except Exception as e:  # noqa: BLE001
                st.error(f"Build failed: {e}")
                return
        stem = re.sub(r"\.xls[xm]$", "", up.name, flags=re.I)
        ext = "xlsm" if macro else "xlsx"    # a macro workbook has to stay .xlsm
        mime = ("application/vnd.ms-excel.sheet.macroEnabled.12" if macro else
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.session_state["puw_out"] = (out, f"{stem} - Public.{ext}", mime)
        st.success(f"Built — {len(keep)} tab(s), {p['total_hardcoded']} value(s) frozen, "
                   f"{p['total_kept']} formula(s) still live.")

    if "puw_out" in st.session_state:
        out, fname, mime = st.session_state["puw_out"]
        st.download_button(f"⬇ Download public workbook (.{fname.rsplit('.', 1)[-1]})",
                           data=out, file_name=fname, mime=mime, use_container_width=True)
