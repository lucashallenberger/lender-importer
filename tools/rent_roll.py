"""Rent Roll Parser — Streamlit page.

Upload a rent-roll PDF (any property-management format). Get one workbook with:
  • Source     — a faithful PDF→Excel transcription of the rent roll
  • Worksheet  — the underwriting rent-roll format, with every shared value a live
                 formula back to the Source sheet (in-place rent, tenant, unit,
                 annual rent, totals, occupancy).

Extraction uses Claude (rent rolls vary too much for a regex parser) via the same
plumbing as the Historicals tool; degrades to a clear message if no API key.
"""

import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st

from tools import hist_llm

# ── styling constants (kept close to the reference screenshots) ─────────────
NAVY   = "203864"
GREEN  = "E2EFDA"
PEACH  = "FCE4D6"
BLUE   = "D9E1F2"

MONEY      = '"$"#,##0.00'
MONEY0     = '"$"#,##0'
MONEY_DASH = '"$"#,##0.00;-"$"#,##0.00;"-"'   # zero shows as "-"
PCT        = '0.0%'

UNIT_COLS = ["Unit", "Unit Type", "SF", "Monthly Rent", "Status", "Lease Name",
             "Lease Status", "Move-In Date", "Lease End", "Lease Term", "Notes"]


# ── extraction ──────────────────────────────────────────────────────────────
_EXTRACT_PROMPT = """This is a property RENT ROLL. Transcribe it faithfully — do
not compute, merge, or invent anything.

Return the property header and one entry per UNIT (include vacant units).

Header:
- property_name: building name/street (e.g. "2821 Sierra St"), or "" if absent
- city_state_zip: e.g. "Los Angeles, CA 90031", or ""
- apn: assessor parcel number if present, else null
- source_note: one line describing the report source exactly as printed
  (software, who generated it, dates), or ""

For each unit:
- unit: unit number/label exactly as printed (e.g. "01")
- unit_type: floor-plan/type exactly as printed (e.g. "1/1/d", "Single")
- unit_sf: the unit's square footage as a number, or null if not shown
- monthly_rent: the unit's monthly rent as a number. For a VACANT unit use the
  target/asking/market rent shown for it (or 0 if none is shown).
- status: "Occupied" or "Vacant"
- lease_name: tenant/lease name, or "" if vacant/blank
- lease_status: e.g. "Active", or ""
- move_in_date: move-in / lease start date as printed (e.g. "10/01/2013"), or ""
- lease_end: lease end as printed — a date or "MTM", or ""
- lease_type: the lease term exactly as printed (e.g. "MTM", "1 year",
  "MTM since 4/25", "Airbnb 2 months begin 6/15/26"), or ""

Use the primary rent column if several are shown. Keep the units in printed order."""

_SCHEMA = {
    "type": "object",
    "properties": {
        "property_name": {"type": "string"},
        "city_state_zip": {"type": "string"},
        "apn": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "source_note": {"type": "string"},
        "units": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "unit": {"type": "string"},
                    "unit_type": {"type": "string"},
                    "unit_sf": {"anyOf": [{"type": "number"}, {"type": "null"}]},
                    "monthly_rent": {"anyOf": [{"type": "number"}, {"type": "null"}]},
                    "status": {"type": "string"},
                    "lease_name": {"type": "string"},
                    "lease_status": {"type": "string"},
                    "move_in_date": {"type": "string"},
                    "lease_end": {"type": "string"},
                    "lease_type": {"type": "string"},
                },
                "required": ["unit", "unit_type", "unit_sf", "monthly_rent", "status",
                             "lease_name", "lease_status", "move_in_date",
                             "lease_end", "lease_type"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["property_name", "city_state_zip", "apn", "source_note", "units"],
    "additionalProperties": False,
}


def extract_rent_roll(pdf_bytes: bytes) -> dict:
    """Claude reads the rent-roll PDF and returns the structured dict above."""
    import base64
    resp = hist_llm._client().messages.create(
        model=hist_llm.MODEL,
        max_tokens=16000,
        thinking={"type": "adaptive"},
        messages=[{
            "role": "user",
            "content": [
                {"type": "document",
                 "source": {"type": "base64", "media_type": "application/pdf",
                            "data": base64.standard_b64encode(pdf_bytes).decode()}},
                {"type": "text", "text": _EXTRACT_PROMPT},
            ],
        }],
        output_config={"format": {"type": "json_schema", "schema": _SCHEMA}},
    )
    return hist_llm._json_response(resp)


# ── helpers ─────────────────────────────────────────────────────────────────
def _is_vacant(u: dict) -> bool:
    return "vac" in str(u.get("status", "")).lower()


def _norm_type(t: str) -> str:
    """Present a bed/bath type in '1 + 1' form; 'Studio'/'Single' -> 'Studio'."""
    if not t:
        return ""
    s = t.lower()
    if "studio" in s or "single" in s:
        return "Studio"
    m = re.match(r'\s*(\d+)\s*[/x+]\s*(\d+)', s)
    if m:
        return f"{m.group(1)} + {m.group(2)}"
    return t


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _beds(u):
    """Bedroom count from the unit type; Studio/Single -> 0, else leading int."""
    s = str(u.get("unit_type", "")).lower()
    if "studio" in s or "single" in s:
        return 0
    m = re.match(r'\s*(\d+)', s)
    return int(m.group(1)) if m else None


def _movein(u):
    s = str(u.get("move_in_date", "")).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _auto_notes(units):
    """Analytical flags per unit (biggest/smallest, highest/lowest rent, tenure,
    vacancy) — written into the Notes column instead of extra columns.
    Returns {index: note-string}. Only labels a superlative when it's unique."""
    n = len(units)
    tags = {i: [] for i in range(n)}

    for i, u in enumerate(units):
        if _is_vacant(u):
            tags[i].append("Vacant — target/asking rent")

    # biggest / smallest unit by bedroom count (only when there's a mix)
    beds = {i: _beds(u) for i, u in enumerate(units)}
    known = [v for v in beds.values() if v is not None]
    if known and len(set(known)) > 1:
        big = [i for i, v in beds.items() if v == max(known)]
        small = [i for i, v in beds.items() if v == min(known)]
        if len(big) == 1:
            tags[big[0]].append("Biggest unit")
        if len(small) == 1:
            tags[small[0]].append("Smallest unit")

    # highest / lowest in-place rent among occupied units with a real rent
    rents = {i: _num(u.get("monthly_rent")) for i, u in enumerate(units)
             if not _is_vacant(u) and _num(u.get("monthly_rent")) > 0}
    if len(rents) > 1:
        hi = max(rents, key=rents.get)
        lo = min(rents, key=rents.get)
        tags[hi].append("Highest rent")
        if lo != hi:
            tags[lo].append("Lowest rent")

    # longest-tenured / newest lease among occupied units with a parseable date
    dates = {i: _movein(u) for i, u in enumerate(units)
             if not _is_vacant(u) and _movein(u)}
    if len(dates) > 1:
        oldest = min(dates, key=dates.get)
        newest = max(dates, key=dates.get)
        tags[oldest].append("Longest-tenured")
        if newest != oldest:
            tags[newest].append("Newest lease")

    return {i: " · ".join(t) for i, t in tags.items()}


# ── workbook: Source sheet (faithful transcription) ─────────────────────────
def _build_source(ws, data):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr_font = Font(bold=True, color="FFFFFF")
    navy = PatternFill("solid", fgColor=NAVY)
    green = PatternFill("solid", fgColor=GREEN)
    peach = PatternFill("solid", fgColor=PEACH)
    blue = PatternFill("solid", fgColor=BLUE)
    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    sub = Font(size=11, color="595959")
    note = Font(italic=True, size=9, color="808080")

    units = data["units"]
    n = len(units)

    # title block
    ws.cell(1, 1, f"{data.get('property_name') or 'Property'} — Rent Roll").font = title
    subline = " · ".join(x for x in [data.get("city_state_zip"),
                                     f"APN {data['apn']}" if data.get("apn") else None,
                                     f"{n} Units"] if x)
    ws.cell(2, 1, subline).font = sub
    if data.get("source_note"):
        ws.cell(3, 1, f"Source: {data['source_note']}").font = note

    headers = ["Unit", "Type", "Monthly Rent", "Status", "Lease Name",
               "Lease Status", "Move-In Date"]
    # keep everything the PDF showed: append optional columns AFTER the fixed
    # seven so the Worksheet's A–G links never move
    has_sf = any(u.get("unit_sf") for u in units)
    has_end = any((u.get("lease_end") or "").strip() for u in units)
    has_term = any((u.get("lease_type") or "").strip() for u in units)
    extras = ([("SF", "unit_sf")] if has_sf else []) \
           + ([("Lease End", "lease_end")] if has_end else []) \
           + ([("Lease Term", "lease_type")] if has_term else [])
    headers += [h for h, _ in extras]
    ncol = len(headers)
    hrow = 5
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(hrow, c, h)
        cell.font = hdr_font; cell.fill = navy
        cell.alignment = Alignment(horizontal="center")

    ds = hrow + 1
    for i, u in enumerate(units):
        r = ds + i
        vac = _is_vacant(u)
        ws.cell(r, 1, u.get("unit", "")).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, u.get("unit_type", ""))
        rc = ws.cell(r, 3, _num(u.get("monthly_rent"))); rc.number_format = MONEY
        ws.cell(r, 4, "Vacant" if vac else "Occupied")
        ws.cell(r, 5, "" if vac else u.get("lease_name", ""))
        ws.cell(r, 6, "" if vac else u.get("lease_status", ""))
        ws.cell(r, 7, u.get("move_in_date", ""))
        for j, (_, k) in enumerate(extras):
            v = u.get(k)
            cell = ws.cell(r, 8 + j, float(v) if k == "unit_sf" and v else (v or ""))
            if k == "unit_sf" and v:
                cell.number_format = "#,##0"
        if vac:
            for c in range(1, ncol + 1):
                ws.cell(r, c).fill = peach
                ws.cell(r, c).font = Font(italic=True, color="C00000")
    de = ds + n - 1

    # TOTAL row
    tr = de + 1
    ws.cell(tr, 1, "TOTAL").font = bold
    ws.cell(tr, 2, f"{n} units").font = bold
    tc = ws.cell(tr, 3, f"=SUM(C{ds}:C{de})"); tc.font = bold; tc.number_format = MONEY
    if has_sf:
        sc = ws.cell(tr, 8, f"=SUM(H{ds}:H{de})"); sc.font = bold; sc.number_format = "#,##0"
    for c in range(1, ncol + 1):
        ws.cell(tr, c).fill = blue

    # Rent Summary block (live formulas)
    sr = tr + 2
    ws.cell(sr, 1, "Rent Summary").font = bold
    rows = [
        ("Units — Total", n, None),
        ("Units — Occupied", f'=COUNTIF(D{ds}:D{de},"Occupied")', None),
        ("Units — Vacant", f'=COUNTIF(D{ds}:D{de},"Vacant")', None),
        ("Occupancy %", f'=C{sr+2}/C{sr+1}', PCT),
        ("In-Place Monthly Rent (occupied only)", f'=SUMIF(D{ds}:D{de},"Occupied",C{ds}:C{de})', MONEY),
        ("Target Rent on Vacant Unit(s)", f'=SUMIF(D{ds}:D{de},"Vacant",C{ds}:C{de})', MONEY),
        ("Gross Potential Monthly Rent (in-place + target)", f'=SUM(C{ds}:C{de})', MONEY),
        ("Annualized Gross Potential Rent", f'=C{sr+7}*12', MONEY),
    ]
    for j, (label, val, fmt) in enumerate(rows, start=1):
        rr = sr + j
        ws.cell(rr, 1, label)
        cell = ws.cell(rr, 3, val)
        if fmt:
            cell.number_format = fmt
        for c in range(1, 8):
            ws.cell(rr, c).fill = green
    ws.cell(sr, 1).fill = green

    widths = [8, 12, 16, 12, 22, 14, 14] + [10] * len(extras)
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    src = ws.title
    return {
        "src_title": src,
        "data_start": ds, "data_end": de,
        "occupancy_cell": f"'{src}'!C{sr+4}",
        "total_cell": f"'{src}'!C{sr+1}",
        "occupied_cell": f"'{src}'!C{sr+2}",
        "vacant_cell": f"'{src}'!C{sr+3}",
        "gpr_monthly_cell": f"'{src}'!C{sr+7}",
        "gpr_annual_cell": f"'{src}'!C{sr+8}",
    }


# ── workbook: Worksheet sheet (underwriting format, sourced) ────────────────
# Columns: A UNIT#  B TENANT NAME  C UNIT TYPE  D SIZE(SF)  E SIZE(%)
#          F IN-PLACE $/SF  G IN-PLACE $  H MARKET $/SF  I MARKET $
#          J LEASE TERM  K LEASE START  L LEASE END  M IN-PLACE VACANCY
#          N LOSS TO LEASE  O NOTES
def _build_worksheet(ws, data, refs):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_font = Font(bold=True, color="FFFFFF")
    grey = PatternFill("solid", fgColor="808080")
    green = PatternFill("solid", fgColor=GREEN)
    bold = Font(bold=True)
    redi = Font(italic=True, color="C00000")
    note = Font(italic=True, size=9, color="808080")
    hairline = Border(bottom=Side(style="hair", color="404040"))
    center = Alignment(horizontal="center", vertical="center")

    # ── 3-row grouped header ─────────────────────────────────────────────
    top = {1: "UNIT", 2: "TENANT", 3: "UNIT", 4: "UNIT SIZE", 6: "IN PLACE RENT",
           8: "MARKET RENT", 10: "LEASE", 11: "LEASE", 12: "LEASE",
           13: "IN PLACE", 14: "LOSS", 15: "NOTES"}
    mid = {1: "#", 2: "NAME", 3: "TYPE", 10: "TERM", 11: "START", 12: "END",
           13: "VACANCY", 14: "TO LEASE"}
    sub = {4: "(SF)", 5: "(%)", 6: "($/SF)", 7: "($)", 8: "($/SF)", 9: "($)"}
    for c, v in top.items():
        ws.cell(1, c, v)
    for c, v in mid.items():
        ws.cell(2, c, v)
    for c, v in sub.items():
        ws.cell(3, c, v)
    for a, b in ((4, 5), (6, 7), (8, 9)):     # group labels span their two columns
        ws.merge_cells(start_row=1, start_column=a, end_row=1, end_column=b)
    for r in (1, 2, 3):
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.font = hdr_font; cell.fill = grey; cell.alignment = center

    units = data["units"]
    n = len(units)
    ds = refs["data_start"]
    S = refs["src_title"]
    auto_notes = _auto_notes(units)
    d0, d1 = 4, 3 + n                        # data rows

    for i, u in enumerate(units):
        r = d0 + i
        sr = ds + i                          # matching Source row
        vac = _is_vacant(u)
        sf = u.get("unit_sf")
        ws.cell(r, 1, f"='{S}'!A{sr}").alignment = center
        ws.cell(r, 2, "VACANT" if vac else f"='{S}'!E{sr}")
        ws.cell(r, 3, _norm_type(u.get("unit_type", "")))
        if sf:
            ws.cell(r, 4, float(sf)).number_format = "#,##0"
        c = ws.cell(r, 5, f'=IF(OR(D{r}="",SUM($D${d0}:$D${d1})=0),"",D{r}/SUM($D${d0}:$D${d1}))')
        c.number_format = "0.0%"
        ws.cell(r, 6, f'=IF(OR(D{r}="",D{r}=0,G{r}=""),"",G{r}/D{r})').number_format = MONEY
        g = ws.cell(r, 7, 0 if vac else f"='{S}'!C{sr}"); g.number_format = MONEY
        ws.cell(r, 8, f'=IF(OR(D{r}="",D{r}=0,I{r}=""),"",I{r}/D{r})').number_format = MONEY
        # market $ defaults: vacant -> target/asking from Source; else = in-place (edit over it)
        mi = ws.cell(r, 9, f"='{S}'!C{sr}" if vac else f"=G{r}"); mi.number_format = MONEY
        ws.cell(r, 10, u.get("lease_type", ""))
        ws.cell(r, 11, f"='{S}'!G{sr}").alignment = Alignment(horizontal="right")
        ws.cell(r, 12, u.get("lease_end", "")).alignment = Alignment(horizontal="right")
        # user-specified formulas, verbatim (row-adjusted)
        ws.cell(r, 13, f'=IF(B{r}="Vacant",I{r},0)').number_format = MONEY_DASH
        ws.cell(r, 14, f"=IF(M{r}>0,0,I{r}-G{r})").number_format = MONEY_DASH
        ws.cell(r, 15, u["notes"] if "notes" in u else auto_notes.get(i, ""))
        for c_ in range(1, 16):
            ws.cell(r, c_).border = hairline
        if vac:
            for c_ in (1, 2, 3, 7, 9):
                ws.cell(r, c_).font = redi

    # ── totals ───────────────────────────────────────────────────────────
    mt = d1 + 2
    ws.cell(mt, 1, "Monthly Total").font = bold
    for col in (7, 9, 13, 14):
        L = get_column_letter(col)
        cell = ws.cell(mt, col, f"=SUM({L}{d0}:{L}{d1})")
        cell.font = bold
        cell.number_format = MONEY_DASH if col in (13, 14) else MONEY0
    at = mt + 1
    ws.cell(at, 1, "Annual Total").font = bold
    for col in (7, 9, 13):
        L = get_column_letter(col)
        cell = ws.cell(at, col, f"={L}{mt}*12"); cell.font = bold; cell.number_format = MONEY0

    # ── unit summary (linked to Source) ──────────────────────────────────
    us = at + 3
    summ = [("Total Units", refs["total_cell"], "0"),
            ("Occupied Units", refs["occupied_cell"], "0"),
            ("Vacant Units", refs["vacant_cell"], "0"),
            ("Occupancy", refs["occupancy_cell"], PCT)]
    for j, (label, ref, fmt) in enumerate(summ):
        rr = us + j
        ws.cell(rr, 2, label).font = bold
        cell = ws.cell(rr, 6, f"={ref}"); cell.number_format = fmt
        for c_ in range(2, 8):
            ws.cell(rr, c_).fill = green

    if data.get("source_note"):
        ws.cell(us + 5, 2, f"Source: {data['source_note']}  ·  "
                           "Market Rent defaults to in-place — edit; vacant unit shows target/asking rent.").font = note

    widths = [6, 18, 9, 9, 7, 9, 12, 9, 12, 26, 11, 10, 13, 12, 30]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"


def build_into(wb, data: dict, name="W - RR", src_name="S - RR", prop_info=None) -> dict:
    """Write the Worksheet + Source pair (and an optional Property Info cover
    sheet) into an existing workbook. Returns the Source refs for linking."""
    if prop_info:
        from tools import prop_info as PI
        PI.build_sheet(wb.create_sheet("Property Info"), prop_info)
    ws = wb.create_sheet(name)
    src = wb.create_sheet(src_name)
    refs = _build_source(src, data)
    _build_worksheet(ws, data, refs)
    return refs


def build_workbook(data: dict, prop_info=None) -> bytes:
    """Return .xlsx bytes: optional Property Info cover, the Worksheet tab
    (underwriting format, sourced), and the Source tab it links back to."""
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    build_into(wb, data, prop_info=prop_info)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── UI ───────────────────────────────────────────────────────────────────────
def _df_from_units(units):
    notes = _auto_notes(units)
    return pd.DataFrame([{
        "Unit": u.get("unit", ""), "Unit Type": u.get("unit_type", ""),
        "SF": (float(u["unit_sf"]) if u.get("unit_sf") else None),
        "Monthly Rent": _num(u.get("monthly_rent")),
        "Status": "Vacant" if _is_vacant(u) else "Occupied",
        "Lease Name": u.get("lease_name", ""), "Lease Status": u.get("lease_status", ""),
        "Move-In Date": u.get("move_in_date", ""), "Lease End": u.get("lease_end", ""),
        "Lease Term": u.get("lease_type", ""), "Notes": notes.get(i, ""),
    } for i, u in enumerate(units)], columns=UNIT_COLS)


def _units_from_df(df):
    out = []
    for row in df.itertuples(index=False):
        d = dict(zip(UNIT_COLS, row))
        if not str(d["Unit"]).strip():
            continue
        sf = d.get("SF")
        out.append({
            "unit": str(d["Unit"]).strip(), "unit_type": str(d["Unit Type"]).strip(),
            "unit_sf": (float(sf) if sf is not None and pd.notna(sf) and _num(sf) > 0 else None),
            "monthly_rent": _num(d["Monthly Rent"]), "status": str(d["Status"]).strip(),
            "lease_name": str(d["Lease Name"]).strip(), "lease_status": str(d["Lease Status"]).strip(),
            "move_in_date": str(d["Move-In Date"]).strip(), "lease_end": str(d["Lease End"]).strip(),
            "lease_type": str(d["Lease Term"]).strip(), "notes": str(d["Notes"]).strip(),
        })
    return out


def render():
    st.header("🏘️ Rent Roll Parser")
    st.caption("Upload a rent-roll PDF. You get one workbook: a **Source** tab (faithful "
               "PDF→Excel transcription) and a **Worksheet** tab (underwriting format) whose "
               "shared values link by formula back to the Source tab.")

    if not hist_llm.available():
        st.warning("This tool reads rent rolls with Claude — set **ANTHROPIC_API_KEY** in the "
                   "app secrets to enable it. (Rent-roll layouts vary too much for a fixed parser.)")
        return

    files = st.file_uploader("Rent roll PDF(s)", type=["pdf"], accept_multiple_files=True)
    if not files:
        st.info("Upload a rent-roll PDF to begin.")
        return

    parsed = st.session_state.setdefault("rr_parsed", {})
    for f in files:
        key = f"{f.name}:{f.size}"
        if key not in parsed:
            with st.spinner(f"Reading {f.name} with Claude…"):
                try:
                    parsed[key] = extract_rent_roll(f.getvalue())
                except Exception as e:  # noqa: BLE001
                    st.error(f"Could not read {f.name}: {e}")
                    continue
    # drop removed files
    live = {f"{f.name}:{f.size}" for f in files}
    for k in list(parsed):
        if k not in live:
            del parsed[k]

    for f in files:
        key = f"{f.name}:{f.size}"
        data = parsed.get(key)
        if not data:
            continue
        with st.expander(f"📄 {f.name}  —  {data.get('property_name') or '?'}",
                         expanded=len(files) == 1):
            c1, c2, c3 = st.columns(3)
            pname = c1.text_input("Property", value=data.get("property_name") or "", key=f"pn_{key}")
            csz = c2.text_input("City / State / ZIP", value=data.get("city_state_zip") or "", key=f"csz_{key}")
            apn = c3.text_input("APN", value=data.get("apn") or "", key=f"apn_{key}")
            snote = st.text_input("Source note", value=data.get("source_note") or "", key=f"sn_{key}")

            # ── Property Info cover sheet (web lookup by APN) ────────────
            from tools import prop_info as PI
            if st.button("🔎 Look up property info on the web (by APN)", key=f"pib_{key}",
                         disabled=not apn.strip()):
                zc = re.search(r"\b(\d{5})\b", csz or "")
                with st.spinner(f"Researching APN {apn} on the web…"):
                    try:
                        st.session_state[f"pinfo_{key}"] = PI.fetch(
                            apn, f"{pname} {csz}", zip_code=zc.group(1) if zc else "")
                    except Exception as e:  # noqa: BLE001
                        st.error(f"Lookup failed: {e}")
            pinfo = st.session_state.get(f"pinfo_{key}")
            if pinfo and pinfo.get("verified_address"):
                st.info(f"📍 Verified as: **{pinfo['verified_address']}** — wrong property? "
                        "Fix the City/State/ZIP above and re-run the lookup.")
            if pinfo is not None:
                st.markdown("**Property Info** (from web — verify & edit; blanks stay "
                            "blank on the sheet)")
                pi_edit = {}
                left, mid_, right_ = st.columns(3)
                cols3 = [left, mid_, right_]
                fields = PI.PROP_FIELDS[:2] + [("", "address_line2")] + PI.PROP_FIELDS[2:] + PI.BLDG_FIELDS
                for j, (label, k) in enumerate(fields):
                    v = pinfo.get(k)
                    pi_edit[k] = cols3[j % 3].text_input(label or "Address line 2",
                                                         value="" if v is None else str(v),
                                                         key=f"pif_{k}_{key}")
                # numbers back to numbers where possible
                for k, v in pi_edit.items():
                    v = v.strip()
                    if v == "":
                        pi_edit[k] = None
                    else:
                        try:
                            pi_edit[k] = float(v) if "." in v else int(v)
                        except ValueError:
                            pi_edit[k] = v
                st.session_state[f"pinfo_edit_{key}"] = pi_edit

            st.markdown("**Units** (edit any cell to correct the extraction)")
            df = st.data_editor(_df_from_units(data["units"]), num_rows="dynamic",
                                use_container_width=True, key=f"units_{key}")

            edited = {"property_name": pname, "city_state_zip": csz, "apn": apn or None,
                      "source_note": snote, "units": _units_from_df(df)}

            if edited["units"]:
                xlsx = build_workbook(edited, prop_info=st.session_state.get(f"pinfo_edit_{key}"))
                stem = re.sub(r'[^0-9A-Za-z]+', "_", (pname or f.name)).strip("_") or "rent_roll"
                st.download_button("⬇ Download workbook (.xlsx)", data=xlsx,
                                   file_name=f"{stem}_RentRoll.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key=f"dl_{key}", use_container_width=True)
            else:
                st.info("No units yet — add at least one row to build the workbook.")
