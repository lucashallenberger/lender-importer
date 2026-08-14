"""Spreadsheet ingestion — rent rolls and P&Ls arrive as .xlsx at least as often
as they arrive as PDFs, and the Anthropic API can't read a workbook the way it
reads a PDF page.

Everything here turns a workbook (or CSV) into a plain-text rendering that reads
like the printed report: one line per row, cells separated by ' | ', one banner
per sheet, blank rows/columns dropped, merged cells repeated across their span,
and formulas replaced by their cached values. That text feeds Claude in place of
a document block, and the deterministic classifier greps it exactly as it greps
PDF text.

Format is detected from the bytes (magic number), not the filename, so callers
never have to thread a name through.
"""

import csv
import io
import re
from datetime import date, datetime, time

# accepted upload extensions, for st.file_uploader(type=...)
UPLOAD_TYPES = ["pdf", "xlsx", "xlsm", "xls", "csv", "tsv"]

MAX_ROWS = 4000          # per sheet
MAX_CHARS = 180_000      # whole rendering (~45k tokens)

_PDF = b"%PDF"
_ZIP = b"PK\x03\x04"     # xlsx/xlsm are zip archives
_OLE = b"\xd0\xcf\x11\xe0"   # legacy .xls (OLE2 compound file)


def kind(data) -> str:
    """'pdf' | 'xlsx' | 'xls' | 'text' — from the leading bytes."""
    if not isinstance(data, (bytes, bytearray)):
        return "pdf"                      # paths only ever come from the PDF tools
    head = bytes(data[:8])
    if head.startswith(_PDF):
        return "pdf"
    if head.startswith(_ZIP):
        return "xlsx"
    if head.startswith(_OLE):
        return "xls"
    return "text"


def is_spreadsheet(data) -> bool:
    """True for anything that isn't a PDF (workbook, CSV, plain text)."""
    return kind(data) != "pdf"


def sheet_names(data) -> list:
    """Visible sheet titles in workbook order, or [] for anything that isn't a
    workbook. A workbook often holds one statement per sheet — a year per tab —
    and each has to be read on its own, so callers need the list first."""
    k = kind(data)
    if k == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(bytes(data)), read_only=True)
        try:
            return [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        finally:
            wb.close()
    if k == "xls":
        import pandas as pd
        return [str(n) for n in pd.read_excel(io.BytesIO(bytes(data)), sheet_name=None,
                                              header=None, nrows=0)]
    return []


def split_sheets(text: str) -> list:
    """[(sheet name, that sheet's slice of the rendering)]. to_text() banners
    every sheet, so a caller can weigh them one at a time without re-reading the
    file — cheaper than a round trip per tab to find out what each one holds."""
    parts = re.split(r'^=== SHEET: (.*) ===$', text or "", flags=re.M)
    return [(parts[i].strip(), parts[i + 1]) for i in range(1, len(parts) - 1, 2)]


def to_text(data, only=None) -> str:
    """Render a workbook / CSV as text. `only` renders that one sheet — the rest
    are still read, so cross-sheet references still resolve. Raises with a
    readable message if the file can't be opened."""
    k = kind(data)
    if k == "pdf":
        raise ValueError("to_text() got a PDF — use the PDF path instead")
    out = (_xlsx_text(data, only) if k == "xlsx" else
           _xls_text(data, only) if k == "xls" else _csv_text(data))
    if len(out) > MAX_CHARS:
        out = out[:MAX_CHARS] + "\n… [rendering truncated — file is very large]"
    return out


# ── value formatting ────────────────────────────────────────────────────────
def _fmt(v, nf="General") -> str:
    """Cell value as it would read on screen. Money-formatted numbers keep two
    decimals (the statement regexes key off that); plain integers stay plain so
    'Year Built 1923' doesn't become '1,923.00'."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, time) and not isinstance(v, datetime):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%m/%d/%Y %H:%M") if (v.hour or v.minute) else v.strftime("%m/%d/%Y")
    if isinstance(v, date):
        return v.strftime("%m/%d/%Y")
    if isinstance(v, (int, float)):
        f = float(v)
        if f != f or f in (float("inf"), float("-inf")):   # NaN / inf
            return ""
        # follow the cell's own format: thousands separators only where Excel
        # shows them, so "Year Built 2021" doesn't come out as "2,021"
        nf = nf or ""
        grouped = "," in nf or "$" in nf
        if "%" in nf:
            return f"{f * 100:,.2f}".rstrip("0").rstrip(".") + "%"
        if "$" in nf or "0.00" in nf or f != int(f):
            return f"{f:,.2f}" if grouped else f"{f:.2f}"
        return f"{int(f):,}" if grouped else str(int(f))
    return re.sub(r"\s+", " ", str(v)).strip()


def _render_sheet(name: str, rows: list) -> str:
    """rows: list of lists of already-formatted strings."""
    ncol = max((len(r) for r in rows), default=0)
    if not ncol:
        return ""
    keep = [c for c in range(ncol) if any(len(r) > c and r[c] for r in rows)]
    if not keep:
        return ""
    lines, blanks = [], 0
    for r in rows:
        cells = [(r[c] if len(r) > c else "") for c in keep]
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            blanks += 1
            if blanks == 1:
                lines.append("")
            continue
        blanks = 0
        lines.append(" | ".join(cells))
    while lines and not lines[0]:
        lines.pop(0)
    while lines and not lines[-1]:
        lines.pop()
    return f"=== SHEET: {name} ===\n" + "\n".join(lines) if lines else ""


def _join(parts: list) -> str:
    body = "\n\n".join(p for p in parts if p)
    if not body.strip():
        raise ValueError("the spreadsheet is empty — no readable cells")
    return body


# ── readers ─────────────────────────────────────────────────────────────────
# A cell whose whole formula is one reference: ='S - RR'!C6, =Source!C6, =$C$6.
_REF_HERE = re.compile(r"^=\s*\$?([A-Z]{1,3})\$?(\d+)\s*$")
_REF_SHEET = re.compile(r"^=\s*(?:'([^']+)'|([^'!*?\[\]/\\]+))!\$?([A-Z]{1,3})\$?(\d+)\s*$")


def _xlsx_text(data, only=None) -> str:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(bytes(data)), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"could not open this .xlsx ({e})") from e
    # A workbook written by a script (never opened in Excel) carries no cached
    # formula results, so data_only gives None everywhere. Keep the formulas so
    # the plain references among them can be followed to their real values.
    try:
        wbf = openpyxl.load_workbook(io.BytesIO(bytes(data)), data_only=False)
    except Exception:  # noqa: BLE001
        wbf = None

    # every sheet is read (a hidden one is often the data behind the visible
    # links), but only the visible ones are rendered
    grids, pending, visible = {}, [], []      # title -> rows; [(title, ri, ci, formula)]
    for ws in wb.worksheets:
        if ws.sheet_state == "visible" and (only is None or ws.title == only):
            visible.append(ws.title)
        wsf = wbf[ws.title] if (wbf is not None and ws.title in wbf.sheetnames) else None
        nrow = min(ws.max_row or 0, MAX_ROWS)
        rows = []
        for ri, row in enumerate(ws.iter_rows(max_row=nrow)):
            out = []
            for ci, cell in enumerate(row):
                txt = _fmt(cell.value, cell.number_format)
                if not txt and wsf is not None:
                    f = wsf.cell(cell.row, cell.column).value
                    if isinstance(f, str) and f.startswith("="):
                        txt = f
                        pending.append((ws.title, ri, ci, f))
                out.append(txt)
            rows.append(out)
        if (ws.max_row or 0) > MAX_ROWS:
            rows.append([f"… [{ws.max_row - MAX_ROWS} more rows not shown]"])
        grids[ws.title] = rows
    _resolve_refs(grids, pending)
    return _join([_render_sheet(t, grids[t]) for t in visible])


def _resolve_refs(grids, pending):
    """Replace '=Source!C6'-style cells with the value they point at. Workbooks
    this app writes are all cross-sheet links, and without cached results they
    would otherwise reach Claude as a page of formulas."""
    from openpyxl.utils import column_index_from_string

    def at(title, r, c):                       # 1-based sheet coords -> text
        rows = grids.get(title)
        if rows is None or r > len(rows):
            return None
        row = rows[r - 1]
        return row[c - 1] if c <= len(row) else ""

    for _pass in range(3):                     # follow short chains of links
        moved = False
        for title, ri, ci, f in pending:
            cur = grids[title][ri][ci]
            if not cur.startswith("="):
                continue
            m = _REF_HERE.match(f)
            tgt, col, num = (title, m.group(1), m.group(2)) if m else (None, None, None)
            if not m:
                m = _REF_SHEET.match(f)
                if not m:
                    continue
                tgt, col, num = (m.group(1) or m.group(2)).strip(), m.group(3), m.group(4)
            val = at(tgt, int(num), column_index_from_string(col))
            if val is None or val.startswith("="):
                continue                       # unknown sheet, or still a formula
            grids[title][ri][ci] = val
            moved = True
        if not moved:
            break


def _xls_text(data, only=None) -> str:
    try:
        import pandas as pd
        sheets = pd.read_excel(io.BytesIO(bytes(data)), sheet_name=None,
                               header=None, dtype=object)
    except ImportError as e:
        raise ValueError("reading legacy .xls files needs the 'xlrd' package "
                         "(pip install xlrd) — or re-save the file as .xlsx") from e
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"could not open this .xls ({e}) — try re-saving it as .xlsx") from e
    parts = []
    for name, df in sheets.items():
        if only is not None and str(name) != only:
            continue
        rows = [[_fmt(v) for v in rec] for rec in df.head(MAX_ROWS).itertuples(index=False)]
        if len(df) > MAX_ROWS:
            rows.append([f"… [{len(df) - MAX_ROWS} more rows not shown]"])
        parts.append(_render_sheet(str(name), rows))
    return _join(parts)


def _decode(data) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return bytes(data).decode(enc)
        except UnicodeDecodeError:
            continue
    return bytes(data).decode("utf-8", "replace")


def _csv_text(data) -> str:
    txt = _decode(data)
    try:
        dialect = csv.Sniffer().sniff(txt[:4000], delimiters=",;\t|")
    except Exception:  # noqa: BLE001
        dialect = csv.excel
    rows = []
    for i, r in enumerate(csv.reader(io.StringIO(txt), dialect)):
        if i >= MAX_ROWS:
            rows.append(["… [more rows not shown]"])
            break
        rows.append([re.sub(r"\s+", " ", (c or "")).strip() for c in r])
    return _join([_render_sheet("Sheet1", rows)])

# A development budget spreads its cost lines across a monthly draw schedule —
# sixty to eighty columns of numbers that say nothing about what a cost IS.
# Sending them to Claude costs more than the rest of the document put together.
_PERIOD_HEAD = re.compile(
    r"^\s*(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"          # 09/01/2026
    r"|\d{4}-\d{2}"                                    # 2026-09
    r"|(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s-]*\d{2,4}"
    r"|\d{1,3}(?:\.\d+)?%?)\s*$", re.I)


def drop_period_columns(text: str, keep_first: int = 20, min_run: int = 8) -> str:
    """Drop a long run of period columns from a rendered sheet. Only a run that
    is unbroken and sits to the right of the labelled columns goes — a table that
    is genuinely wide keeps everything."""
    rows = [line.split(" | ") for line in text.split("\n")]
    width = max((len(r) for r in rows), default=0)
    if width <= keep_first + min_run:
        return text
    header = max(rows, key=lambda r: sum(1 for c in r if c.strip()), default=[])
    run, best = [], []
    for i in range(keep_first, width):
        head = header[i].strip() if i < len(header) else ""
        col = [r[i].strip() for r in rows if i < len(r)]
        numeric = sum(1 for v in col if v and re.fullmatch(r"-?[\d,.$()%]+", v))
        filled = sum(1 for v in col if v)
        if _PERIOD_HEAD.match(head) or (filled and numeric / filled > 0.9):
            run.append(i)
        else:
            if len(run) > len(best):
                best = run
            run = []
    if len(run) > len(best):
        best = run
    if len(best) < min_run:
        return text
    drop = set(best)
    out = []
    for r in rows:
        kept = [c for i, c in enumerate(r) if i not in drop]
        while kept and not kept[-1].strip():
            kept.pop()
        out.append(" | ".join(kept))
    note = (f"\n\n[{len(drop)} period/schedule columns were left out of this "
            f"rendering — they hold a month-by-month spread of the same costs.]")
    return "\n".join(out) + note
