"""Historicals engine: parse property income statements (summary or transaction
detail) and assemble a workbook — one source tab per statement + a combined tab
where every value is a formula linking back to its source, categories are aligned
across years (union), missing-year lines shown in red, and each line classified."""

import io
import re
import datetime
import difflib

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tools import xl as XL   # safe/unique sheet titles

CUR = '#,##0.00'
RED = Font(color='C00000'); REDB = Font(color='C00000', bold=True)
BOLD = Font(bold=True); PLAIN = Font()
HDR = Font(bold=True, color='FFFFFF'); HFILL = PatternFill('solid', fgColor='4F46E5')
TOTAL_FILL = PatternFill('solid', fgColor='EEF0FF')   # very light indigo band
TOTAL_TOP = Border(top=Side(style='thin', color='9CA3AF'))

# Amounts may or may not carry a "$": QuickBooks prints it only on bold/total
# lines, so requiring it silently dropped every ordinary line item.
_ROW = re.compile(r'^(.*?[A-Za-z].*?)\s+((?:-?\$?[\d,]+\.\d{2})(?:\s+-?\$?[\d,]+\.\d{2})*)\s*$')
_AMT = re.compile(r'-?\$?[\d,]+\.\d{2}')
# type may be multi-word: "Journal Entry", "Credit Memo", "Prepayment Application"
_TXN = re.compile(r'^([A-Za-z][A-Za-z ]{0,24}?)\s+(\d{1,2}/\d{1,2}/\d{4})\s+.*?(-?\$[\d,]+\.\d{2})\s+(-?\$[\d,]+\.\d{2})\s*$')
_HEADER = re.compile(r'^(.+?)\s*\((\d{3,6})\)\s*$')
_TOTAL = re.compile(r'^Total\s+(.+?)\s+(-?\$?[\d,]+\.\d{2})\s*$')
_MONEY = re.compile(r'-?\$[\d,]+\.\d{2}')
_PERIOD = re.compile(r'(\d{1,2})/(\d{1,2})/(\d{4})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{4})')


def _money(m):
    return float(m.replace('$', '').replace(',', ''))


def _doubled(s):
    return len(s) >= 2 and len(s) % 2 == 0 and all(s[i] == s[i + 1] for i in range(0, len(s), 2))


def _text(data):
    """Plain text of a source document: PDF text layer, or the sheet-by-sheet
    rendering for workbooks/CSVs (so classification greps both the same way)."""
    from tools import tabular
    if tabular.is_spreadsheet(data):
        return tabular.to_text(data)
    with pdfplumber.open(io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data) as pdf:
        return '\n'.join((p.extract_text() or '') for p in pdf.pages)


def detect_kind(data):
    from tools import tabular
    if tabular.is_spreadsheet(data):
        return 'summary'      # spreadsheet statements go to the LLM as summaries
    t = _text(data)
    return 'detail' if ('Cash Flow Detail' in t or _doubled('CCaasshh') and 'DDeettaaiill' in t or
                        sum(bool(_TXN.match(l.strip())) for l in t.splitlines()) > 15) else 'summary'


# ---------------------------------------------------------------- parsers
_SECTION_WORDS = {'income', 'expenses', 'expense', 'other income', 'other expenses',
                  'revenue', 'cost of goods sold', 'operating expenses'}


def parse_summary(data):
    lines = [l.strip() for l in _text(data).splitlines() if l.strip()]
    # A parent/section line carries no amount of its own, but the report prints a
    # "Total for <parent>" for it — use that to tell real sections apart from the
    # page title and footer, which also have no amount.
    parents = set()
    for l in lines:
        m = re.match(r'^Total\s+for\s+(.+?)(?:\s+-?\$?[\d,]+\.\d{2})?\s*$', l)
        if m:
            parents.add(canon(m.group(1)))
    rows = []
    for s in lines:
        m = _ROW.match(s)
        if m:
            label = m.group(1).strip()
            amounts = [_money(x) for x in _AMT.findall(m.group(2))]   # keep EVERY column
            if not amounts:
                continue
            low = label.lower()
            rows.append({'label': label, 'amount': amounts[0], 'amounts': amounts,
                         'total': low.startswith('total') or low.startswith('gross'),
                         'net': low.startswith('net')})
        elif s.lower() in _SECTION_WORDS or canon(s) in parents:
            rows.append({'label': s, 'amount': None, 'section': True})
    return rows


def parse_detail(data):
    t = _text(data)
    cats, totals, cur = {}, {}, None
    seen_dates = []
    for raw in t.splitlines():
        s = raw.strip()
        if not s or _doubled(s):
            continue
        mt = _TOTAL.match(s)
        if mt:
            totals[mt.group(1).strip()] = _money(mt.group(2)); cur = None; continue
        mtx = _TXN.match(s)
        if mtx and cur:
            d = datetime.datetime.strptime(mtx.group(2), '%m/%d/%Y').date()
            seen_dates.append(d)
            cats.setdefault(cur, {}).setdefault(f'{d.year}-{d.month:02d}', 0.0)
            cats[cur][f'{d.year}-{d.month:02d}'] += _money(mtx.group(4)); continue
        mh = _HEADER.match(s)
        if mh and not _MONEY.search(s):
            cur = mh.group(1).strip()
    # month columns: prefer the printed "Report Period: m/d/y - m/d/y" (header text
    # may be char-doubled, so also try the de-doubled line), widened to cover any
    # transaction dates; fall back to the transaction date range alone.
    period = None
    for raw in t.splitlines()[:40]:
        period = _PERIOD.search(raw) or _PERIOD.search(raw[::2])
        if period:
            break
    starts, ends = [], []
    if period:
        g = period.groups()
        starts.append(datetime.date(int(g[2]), int(g[0]), 1))
        ends.append(datetime.date(int(g[5]), int(g[3]), 1))
    if seen_dates:
        starts.append(datetime.date(min(seen_dates).year, min(seen_dates).month, 1))
        ends.append(datetime.date(max(seen_dates).year, max(seen_dates).month, 1))
    months = []
    if starts:
        d, end = min(starts), max(ends)
        while d <= end:
            months.append((f'{d.year}-{d.month:02d}', d.strftime('%b-%y')))
            d = (d.replace(day=28) + datetime.timedelta(days=7)).replace(day=1)
    return {'cats': cats, 'totals': totals, 'months': months}


# -------------------------------------------- period columns -> detail / summary
_MONTH_WORDS = {m: i for i, m in enumerate(
    ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec'], 1)}
_TOTALISH = re.compile(r'\b(total|ytd|year[\s-]?to[\s-]?date|annual|sum)\b', re.I)


def parse_period(s):
    """A column header -> 'YYYY-MM', or None if it isn't a month.
    Handles 'Jan-25', 'January 2025', '01/2025', '2025-01', 'Jan 2025'."""
    s = str(s or '').strip()
    if not s:
        return None
    m = re.search(r'([A-Za-z]{3,9})[\s\-./]*(\d{2,4})', s)
    if m and m.group(1)[:3].lower() in _MONTH_WORDS:
        mo, yr = _MONTH_WORDS[m.group(1)[:3].lower()], int(m.group(2))
        return f'{2000 + yr if yr < 100 else yr}-{mo:02d}'
    m = re.fullmatch(r'\s*(\d{1,2})[/\-](\d{2,4})\s*', s)      # 01/2025
    if m and 1 <= int(m.group(1)) <= 12:
        yr = int(m.group(2))
        return f'{2000 + yr if yr < 100 else yr}-{int(m.group(1)):02d}'
    m = re.fullmatch(r'\s*(\d{4})[-/](\d{1,2})\s*', s)         # 2025-01
    if m and 1 <= int(m.group(2)) <= 12:
        return f'{m.group(1)}-{int(m.group(2)):02d}'
    return None


def _month_columns(periods):
    """[(column index, 'YYYY-MM')] for the month columns, chronological, deduped."""
    seen, keep = set(), []
    for i, p in enumerate(periods or []):
        k = parse_period(p)
        if k and k not in seen:
            seen.add(k)
            keep.append((i, k))
    keep.sort(key=lambda t: t[1])
    return keep


def _amt(a):
    return float(a) if isinstance(a, (int, float)) and not isinstance(a, bool) else None


def detail_from_periods(periods, rows, min_months=3):
    """A monthly grid (T-12) -> the same {'cats','totals','months'} shape parse_detail
    produces for QuickBooks detail PDFs, so it flows into the existing month-column
    builder. Returns None when the statement isn't monthly."""
    cols = _month_columns(periods)
    if len(cols) < min_months:
        return None
    months = [(k, datetime.date(int(k[:4]), int(k[5:]), 1).strftime('%b-%y')) for _, k in cols]
    # 'spine' keeps the printed line order — sections and totals included — so a
    # workbook built from monthly statements alone still reads like the statement
    cats, totals, spine = {}, {}, []
    for r in rows or []:
        label = str(r.get('label') or '').strip()
        if not label:
            continue
        amounts = r.get('amounts') or []
        vals = {k: _amt(amounts[i]) for i, k in cols
                if i < len(amounts) and _amt(amounts[i]) is not None}
        kind = r.get('kind')
        if kind == 'section':
            spine.append({'label': label, 'amount': None, 'section': True})
        elif kind in ('total', 'net'):
            if vals:      # a blank printed total is missing, not zero
                # parse_detail keys totals by the category name, not "Total <name>"
                amount = round(sum(vals.values()), 2)
                totals[re.sub(r'^total\s+', '', label, flags=re.I).strip()] = amount
                spine.append({'label': label, 'amount': amount,
                              'total': kind == 'total', 'net': kind == 'net'})
        elif kind == 'item' and vals:
            cats[label] = vals
            spine.append({'label': label, 'amount': round(sum(vals.values()), 2),
                          'total': False, 'net': False})
    return {'cats': cats, 'totals': totals, 'months': months, 'spine': spine} if cats else None


def summary_from_periods(periods, rows):
    """Flatten a multi-column statement to parse_summary's row shape. The headline
    amount is the printed Total/YTD column when there is one, else the sum of the
    month columns, else the first value — never January standing in for the year."""
    cols = _month_columns(periods)
    tot_i = next((i for i, p in enumerate(periods or [])
                  if _TOTALISH.search(str(p or '')) and not parse_period(p)), None)
    out = []
    for r in rows or []:
        label = str(r.get('label') or '').strip()
        if not label:
            continue
        amounts = r.get('amounts') or []
        vals = [_amt(a) for a in amounts]
        kept = [v for v in vals if v is not None]
        if r.get('kind') == 'section':
            out.append({'label': label, 'amount': None, 'section': True})
            continue
        if tot_i is not None and tot_i < len(vals) and vals[tot_i] is not None:
            amount = vals[tot_i]
        elif len(cols) >= 3:
            amount = round(sum(vals[i] for i, _ in cols
                               if i < len(vals) and vals[i] is not None), 2)
        else:
            amount = kept[0] if kept else None
        out.append({'label': label, 'amount': amount, 'amounts': kept or None,
                    'total': r.get('kind') == 'total', 'net': r.get('kind') == 'net'})
    return out


# ---------------------------------------------------------------- normalize + classify
def canon(s):
    s = (s or '').lower().replace('–', '-').replace(':', ' ').replace('.', '')
    s = s.replace('-', ' ').replace('/', ' ').replace('&', 'and')
    s = re.sub(r'\s+', ' ', s).strip()
    out = []
    for w in s.split():
        if not out or out[-1] != w:
            out.append(w)
    s = ' '.join(out)
    al = {'city registration fee': 'city reg', 'dwp general': 'dwp',
          'plumbing drain stoppage': 'drain stoppage', 'repairs repair supplies': 'repair supplies'}
    return al.get(s, s)


_CLASS_RULES = [
    (('management fee',), 'OpEx - Management Fee'),
    (('property tax', 're tax'), 'OpEx - RE Taxes'),
    (('registration', 'permits', 'business license', 'admin'), 'OpEx - G&A'),
    (('mortgage', 'debt service', 'interest paid'), 'Non-OpEx - Mortgage Interest'),
    (('legal', 'professional', 'accounting', 'forensic', 'engineering'), 'Non-OpEx - Legal & Professional'),
    (('escrow', 'prepayment', 'deposit', 'security', 'settlement', 'gain', 'cash back'), 'Non-OpEx - Other'),
    (('leasing fee', 'leasing commission'), 'CapEx - Leasing Commissions'),
    (('dwp', 'gas', 'water', 'sewage', 'electric', 'trash', 'utilit', 'comcast'), 'OpEx - Utilities'),
    (('rent',), 'Income - Rental Income'),
    (('scep', 'city reg', 'move out', 'other income'), 'Income - Other Income'),
    (('depreciation', 'amortization'), 'Non-OpEx - Other'),
]


def classify(label):
    n = canon(label)
    for keys, cls in _CLASS_RULES:
        if any(k in n for k in keys):
            return cls
    return 'OpEx - R&M'


# ---------------------------------------------------------------- assemble
# Why the Claude pass last fell back to the deterministic rules, or None. A
# failure here still yields a workbook — just an unaligned, unclassified one —
# so callers read this after building and say so rather than letting a broken
# combined tab pass for a good one.
LAST_LLM_ERROR = None


def _as_details(detail):
    """The detail argument, as a list. Callers may pass one monthly statement, a
    list of them, or None — every statement that came in as a month-by-month grid
    keeps its months, so two T-12s stay two T-12s."""
    if not detail:
        return []
    ds = [detail] if isinstance(detail, dict) else list(detail)
    return sorted(ds, key=lambda d: str(d.get('label') or ''))


def _detail_spine(d):
    """Row spine from a monthly statement, for a build with no summary statement
    to take one from: its own printed line order where the reader recorded one,
    else its categories followed by its totals."""
    if d.get('spine'):
        return d['spine']
    rows = [{'label': k, 'amount': round(sum(v.values()), 2), 'total': False, 'net': False}
            for k, v in d['cats'].items()]
    return rows + [{'label': k, 'amount': v, 'total': True, 'net': False}
                   for k, v in (d['totals'] or {}).items()]


def _llm_refine(summaries, details):
    """Ask Claude to (a) align variant names onto the newest summary's labels,
    (b) classify all line items, and (c) flag which labels are total/subtotal/net
    rows. Returns (alias_map canon->canon, class_map canon->class, total_keys set
    of canon labels). Empty when no API key / any failure."""
    global LAST_LLM_ERROR
    LAST_LLM_ERROR = None
    try:
        from tools import hist_llm
        if not hist_llm.available():
            return {}, {}, set()          # switched off, not broken
        spine_labels = [r['label'] for r in (summaries[-1]['rows'] if summaries else [])
                        if r.get('amount') is not None and not r.get('total') and not r.get('net')]
        spine_keys = {canon(l) for l in spine_labels}
        others = []
        for sm in summaries[:-1]:
            others += [r['label'] for r in sm['rows']
                       if r.get('amount') is not None and not r.get('total') and not r.get('net')]
        for d in details:
            others += list(d['cats'].keys())
        unmatched = sorted({l for l in others if canon(l) not in spine_keys})
        amap = {}
        for src, tgt in hist_llm.match_labels(unmatched, spine_labels).items():
            if tgt:
                amap[canon(src)] = canon(tgt)
        all_items = sorted({*spine_labels, *others})
        cmap = {canon(l): c for l, c in hist_llm.classify_labels(all_items).items()}
        # every label (incl. totals/sections) so Claude can flag total rows to bold
        every = sorted({r['label'] for sm in summaries for r in sm['rows']}
                       | {l for d in details for l in (*d['cats'], *(d['totals'] or {}))})
        tset = {canon(l) for l, role in hist_llm.label_roles(every).items() if role == 'total'}
        return amap, cmap, tset
    except Exception as e:  # noqa: BLE001
        LAST_LLM_ERROR = f"{type(e).__name__}: {e}"
        return {}, {}, set()


def build_into(wb, summaries, detail, use_llm=True, combined_title='W - Historicals',
               src_prefix='S - '):
    """Write the combined + source sheets into an existing workbook. Returns
    metadata for cross-sheet linking (classification SUMIFs on the combined tab).
    summaries: [{'label','rows'}] oldest->newest. detail: one
    {'label','cats','totals','months'}, a list of them, or None — each gets its
    own month-column block, so a year that arrived month-by-month stays that way.
    use_llm: when an ANTHROPIC_API_KEY is available, Claude aligns variant names,
    classifies line items, and flags total rows to bold; otherwise the
    deterministic rules run alone."""
    details = _as_details(detail)
    amap, cmap, tset = _llm_refine(summaries, details) if use_llm else ({}, {}, set())

    def C(s):
        k = canon(s)
        return amap.get(k, k)

    spine_src = (summaries[-1]['rows'] if summaries
                 else _detail_spine(details[-1]) if details else [])
    spine = [dict(r) for r in spine_src]
    spine_items = {C(r['label']) for r in spine_src if r.get('amount') is not None and not r['total'] and not r.get('net')}

    extras = []
    seen = set(spine_items)
    for sm in summaries[:-1]:
        for r in sm['rows']:
            if r.get('amount') is not None and not r['total'] and not r.get('net') and C(r['label']) not in seen:
                extras.append(dict(r, _only=sm['label'])); seen.add(C(r['label']))
    for d in details:
        for name in d['cats']:
            if C(name) not in seen:
                extras.append({'label': name, 'amount': 0.0, '_only': d['label']}); seen.add(C(name))

    # Insert each extra NEXT TO the spine items it belongs with (same classification,
    # else same top-level group) instead of dumping them all at the bottom. Keeps a
    # line like "Computer Expenses" among the expenses rather than below the totals.
    def _cls(lbl):
        return cmap.get(canon(lbl)) or classify(lbl)

    def _grp(lbl):
        return _cls(lbl).split(' - ')[0]

    def _is_break(r):        # totals/nets/sections are not item rows
        return bool(r.get('total') or r.get('net') or r.get('section'))

    for ex in extras:
        cls, grp = _cls(ex['label']), _grp(ex['label'])
        pos = None
        for i, r in enumerate(spine):        # after the last item sharing the class
            if not _is_break(r) and _cls(r['label']) == cls:
                pos = i + 1
        if pos is None:                       # else after the last item in the group
            for i, r in enumerate(spine):
                if not _is_break(r) and _grp(r['label']) == grp:
                    pos = i + 1
        if pos is None:                       # nothing related in the spine (e.g. non-op
            pos = len(spine)                  # items like mortgage/deposits) -> append;
        spine.insert(pos, ex)                 # same-class followers cluster after it

    # one record per monthly statement — its own months and lookups now, its own
    # column block and source tab below
    dets = [{'d': d, 'months': d['months'], 'nM': len(d['months']),
             'cats_c': {C(k): k for k in d['cats']},
             'tot_c': {C(k): k for k in (d['totals'] or {})}} for d in details]

    def _near(key, table):
        if key in table:
            return table[key]
        m = difflib.get_close_matches(key, list(table), n=1, cutoff=0.86)
        return table[m[0]] if m else None

    def ytd_item(dt, label):
        return _near(C(label), dt['cats_c'])

    def ytd_total(dt, label):
        return _near(re.sub(r'^total\s+', '', C(label)), dt['tot_c'])

    # per-summary lookups (canon -> full row) + how many value columns each has
    smaps = [{C(r['label']): r for r in sm['rows'] if r.get('amount') is not None} for sm in summaries]
    ncols = [max((len(r.get('amounts') or [0]) for r in sm['rows'] if r.get('amount') is not None), default=1)
             for sm in summaries]

    # Labels come from filenames, so titles can be illegal ("P&L 2024/25"),
    # too long, or duplicated — take the title new_sheet actually used so the
    # cross-sheet formulas below can never point at a renamed sheet.
    comb, combined_title = XL.new_sheet(wb, combined_title)
    stab_names, stabs = [], []
    for si, sm in enumerate(summaries):
        ws, title = XL.new_sheet(wb, src_prefix + str(sm['label']))
        ws.append(['Line', sm['label']] + [f'Col {j}' for j in range(2, ncols[si] + 1)])
        stab_names.append(title); stabs.append((sm, ws))
    for dt in dets:
        dt['tab'], dt['tab_name'] = XL.new_sheet(wb, src_prefix + str(dt['d']['label']))
        dt['tab'].append(['Line'] + [l for _, l in dt['months']] + ['YTD Total'])

    def put(ws, row, col, val, red=False, bold=False):
        c = ws.cell(row, col, val)
        c.font = REDB if (red and bold) else RED if red else BOLD if bold else PLAIN
        return c

    # column layout — between sections: blank | solid gray divider | blank
    blanks, grays = set(), set()
    col = 2

    def _sep():
        nonlocal col
        blanks.add(col); grays.add(col + 1); blanks.add(col + 2)
        col += 3

    CLS = 1
    sum_val_cols = []
    sum_blocks = []          # (summary index, Lcol, Vcol)
    for si, sm in enumerate(summaries):
        sum_blocks.append((si, col, col + 1)); sum_val_cols.append(col + 1); col += 2
        _sep()
    for dt in dets:                  # months, then that statement's own YTD + recon
        dt['L'], dt['M0'] = col, col + 1
        dt['T'] = dt['M0'] + dt['nM']
        col = dt['T'] + 1
        _sep()
        dt['RECON'] = col; col += 1
        _sep()
    # compact year-over-year block on the far right: values only, one column per
    # source, so big differences jump out without scanning the wide blocks
    CMP = []                 # (compact col, source value col on this sheet, header)
    for si, sm in enumerate(summaries):
        CMP.append((col, sum_val_cols[si], sm['label'])); col += 1
    for dt in dets:
        CMP.append((col, dt['T'], dt['d']['label'])); col += 1

    # header
    comb.cell(1, CLS, 'Classification')
    for si, Lc, Vc in sum_blocks:
        comb.cell(1, Lc, summaries[si]['label'])
    for dt in dets:
        comb.cell(1, dt['L'], dt['d']['label'])
        for j, (_, lbl) in enumerate(dt['months']):
            comb.cell(1, dt['M0'] + j, lbl)
        comb.cell(1, dt['T'], 'YTD Total')
        comb.cell(1, dt['RECON'], 'Recon')
    for cc, _srcc, hdr in CMP:
        comb.cell(1, cc, hdr)

    total_rows = []
    for i, r in enumerate(spine):
        row = i + 2
        label = r['label']
        is_tot = bool(r.get('total') or r.get('net') or canon(label) in tset); is_sec = bool(r.get('section'))
        if is_tot:
            total_rows.append(row)

        # write source tabs (every parsed value column, not just the first)
        for si, (sm, ws) in enumerate(stabs):
            m = smaps[si].get(C(label))
            ws.cell(row, 1, m['label'] if m else label)
            # a section header has no amount of its own — blank, not 0.00 (a
            # missing *line item*, though, is a real 0 for the combined tab)
            amts = (m.get('amounts') or [m['amount']]) if m else ([None] if is_sec else [0])
            for j in range(ncols[si]):
                ws.cell(row, 2 + j, amts[j] if j < len(amts) else None)
        for dt in dets:
            cm = None if is_sec else ytd_item(dt, label)
            tm = ytd_total(dt, label) if is_tot else None
            dt['cm'], dt['tm'] = cm, tm
            tab, nM, cats = dt['tab'], dt['nM'], dt['d']['cats']
            # tm is the totals *key* ("Income"), never the printed wording — a
            # total row shows its own label, not the category it sums
            tab.cell(row, 1, cm or label)
            if is_tot and tm:
                tab.cell(row, 1 + nM + 1, (dt['d']['totals'] or {}).get(tm))
            elif cm:
                for j, (mk, _) in enumerate(dt['months']):
                    tab.cell(row, 2 + j, round(cats[cm].get(mk, 0.0), 2))
                tab.cell(row, 1 + nM + 1, round(sum(cats[cm].values()), 2))

        if is_sec:
            for si, Lc, Vc in sum_blocks:
                put(comb, row, Lc, label, bold=True)
            for dt in dets:
                put(comb, row, dt['L'], label, bold=True)
            continue
        if not is_tot:
            c = comb.cell(row, CLS, cmap.get(canon(label)) or classify(label))
            c.font = Font(size=9, color='6B7280')

        # summary blocks
        for si, Lc, Vc in sum_blocks:
            present = C(label) in smaps[si] or r.get('_only') == summaries[si]['label']
            put(comb, row, Lc, f"='{stab_names[si]}'!A{row}", red=not present, bold=is_tot)
            put(comb, row, Vc, f"='{stab_names[si]}'!B{row}", red=not present, bold=is_tot).number_format = CUR
        # detail blocks — one per monthly statement
        dpres = []
        for dt in dets:
            present = bool(dt['cm']) or bool(dt['tm']) or r.get('_only') == dt['d']['label']
            dpres.append(present)
            dl, nM = dt['tab_name'], dt['nM']
            put(comb, row, dt['L'], f"='{dl}'!A{row}", red=not present, bold=is_tot)
            for j in range(nM):
                put(comb, row, dt['M0'] + j, f"='{dl}'!{get_column_letter(2 + j)}{row}", red=not present, bold=is_tot).number_format = CUR
            put(comb, row, dt['T'], f"='{dl}'!{get_column_letter(2 + nM)}{row}", red=not present, bold=is_tot).number_format = CUR
            if not is_tot:
                TL, M0L = get_column_letter(dt['T']), get_column_letter(dt['M0'])
                MNL = get_column_letter(dt['M0'] + nM - 1)
                comb.cell(row, dt['RECON'],
                          f'=IF(ABS({TL}{row}-SUM({M0L}{row}:{MNL}{row}))<0.01,"ok","CHECK")')
        # compact year-over-year mirror (same-sheet references, keeps red cues)
        for k, (cc, srcc, _hdr) in enumerate(CMP):
            if k < len(summaries):
                redf = not (C(label) in smaps[k] or r.get('_only') == summaries[k]['label'])
            else:
                redf = not dpres[k - len(summaries)]
            put(comb, row, cc, f"={get_column_letter(srcc)}{row}", red=redf, bold=is_tot).number_format = CUR

    _format(comb, sum_blocks, [dt['L'] for dt in dets], total_rows, blanks, grays)
    for si, (_, ws) in enumerate(stabs):
        _format_src(ws, list(range(2, 2 + ncols[si])))
    for dt in dets:
        _format_src(dt['tab'], list(range(2, 2 + dt['nM'] + 1)))

    # the SUMIF block downstream reads one value column — the newest monthly
    # statement when there is one, else the newest summary
    last = dets[-1] if dets else None
    return {
        'title': combined_title,
        'first_row': 2, 'last_row': len(spine) + 1,
        'cls_letter': 'A',
        'val_letter': (get_column_letter(last['T']) if last
                       else (get_column_letter(sum_val_cols[-1]) if sum_val_cols else None)),
        'val_months': last['nM'] if last else 12,
        'val_label': last['d']['label'] if last else (summaries[-1]['label'] if summaries else ''),
    }


def build_workbook(summaries, detail, use_llm=True):
    """Standalone: build a fresh workbook and return .xlsx bytes."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    build_into(wb, summaries, detail, use_llm)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _format_src(ws, vcols):
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).font = HDR; ws.cell(1, c).fill = HFILL
    ws.freeze_panes = 'B2'; ws.column_dimensions['A'].width = 30
    for r in range(2, ws.max_row + 1):
        for c in vcols:
            ws.cell(r, c).number_format = CUR
    for c in vcols:
        ws.column_dimensions[get_column_letter(c)].width = 11


def _format(comb, sum_blocks, det_label_cols=(), total_rows=(), blanks=(), grays=()):
    blanks, grays = set(blanks), set(grays)
    skip = blanks | grays
    GRAY = PatternFill('solid', fgColor='BFBFBF')
    for c in range(1, comb.max_column + 1):
        hc = comb.cell(1, c)
        if c in blanks:
            comb.column_dimensions[get_column_letter(c)].width = 1.5
        elif c in grays:
            comb.column_dimensions[get_column_letter(c)].width = 2
        else:
            hc.font = HDR; hc.fill = HFILL; hc.alignment = Alignment(horizontal='center')
            comb.column_dimensions[get_column_letter(c)].width = 11
    # solid gray divider bars, full height of the sheet
    for c in grays:
        for r in range(1, comb.max_row + 1):
            comb.cell(r, c).fill = GRAY
    comb.freeze_panes = 'B2'
    comb.row_dimensions[1].height = 22
    comb.column_dimensions['A'].width = 18
    for si, Lc, Vc in sum_blocks:
        comb.column_dimensions[get_column_letter(Lc)].width = 22
    for c in det_label_cols:
        comb.column_dimensions[get_column_letter(c)].width = 24

    # Emphasize total/subtotal/net rows: a light band + a thin top rule so they
    # read as summary lines (bold is already applied per-cell during the build).
    maxc = comb.max_column
    for r in total_rows:
        for c in range(1, maxc + 1):
            if c in skip:
                continue
            cell = comb.cell(r, c)
            cell.fill = TOTAL_FILL
            cell.border = TOTAL_TOP
