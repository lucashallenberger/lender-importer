"""
Read the lender Excel and create Deal Source records in Salesforce (new format).

Sheet columns expected: Account, Contact, Notes, Interest.
The record Name is computed as "<LenderName> - <Property>", the Deal is chosen at
run time, and Contacts are matched by name scoped to their Account.

Usage:
    ./.venv/bin/python import_lenders.py "/path/to/list.xlsx" \
        --deal "Sample - 1234 Main" --property "1234 Main" --dry-run

    # then drop --dry-run for the real upload.

ALWAYS run --dry-run first. It resolves everything and prints what WOULD be
created without writing. If a Contact/Account match is ambiguous and you run in a
real terminal, it will ask you to pick; piped/non-interactive runs leave it blank
and report it instead.
"""

import argparse
import sys

import openpyxl

from src.sf_client import connect, load_config, get_object_api_name
from mapping import (
    MAPPING, DEFAULTS, NAME_FIELD, NAME_JOINER,
    CONTACT_COLUMN, ACCOUNT_COLUMN, ACCOUNT_PREFER_TYPE, INTEREST_COLUMN,
    DECLINED_INTEREST_VALUE, DECLINED_REASON_FIELD, DECLINED_REASON_DEFAULT,
    FUZZY_AUTO_THRESHOLD, FUZZY_SUGGEST_THRESHOLD,
)
from aliases import ACCOUNT_ALIASES, CONTACT_ALIASES
from fuzzy import closest_matches, closest_contacts_in_account, is_clean_partial
from learned_store import (
    load_learned, remember_account, lookup_account, remember_contact, lookup_contact,
)

DEAL_FIELD = "ascendix__Deal__c"  # required lookup on the object, set via --deal


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def clean_text(value):
    """Tidy Excel text: drop the _x000D_ carriage-return artifacts openpyxl leaves."""
    if value is None:
        return None
    s = str(value).replace("_x000D_", "").replace("\r\n", "\n").replace("\r", "\n")
    s = s.strip()
    return s or None


def abbreviate_lender(account_name):
    """Lender name for the record Name: strip leading purely-numeric tokens.
    '3650 REIT' -> 'REIT';  'A10 Capital' -> 'A10 Capital';  '1234 Capital' -> 'Capital'."""
    if not account_name:
        return ""
    tokens = account_name.split()
    while tokens and tokens[0].isdigit():
        tokens.pop(0)
    return " ".join(tokens).strip() or account_name


def _chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def read_excel(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        ws = wb.active
    elif str(sheet).isdigit():
        idx = int(sheet) - 1  # 1-based: --sheet 2 -> second sheet
        if idx < 0 or idx >= len(wb.sheetnames):
            sys.exit(f"ERROR: sheet {sheet} out of range; sheets are {wb.sheetnames}")
        ws = wb[wb.sheetnames[idx]]
    elif sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        sys.exit(f"ERROR: no sheet named '{sheet}'; sheets are {wb.sheetnames}")
    print(f"Using sheet: '{ws.title}'")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"ERROR: {path} has no rows.")
    headers = [clean_text(h) for h in rows[0]]
    records = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        records.append({headers[i]: r[i] for i in range(len(headers)) if headers[i]})
    return headers, records


def describe_object(sf, api_name):
    try:
        return getattr(sf, api_name).describe()
    except Exception as exc:  # noqa: BLE001
        sys.exit(f"ERROR: could not describe object '{api_name}': {exc}")


def validate_mapping(desc, headers):
    """Check every mapped Excel column + SF field exists. Exit on problems."""
    fields_by_name = {f["name"]: f for f in desc["fields"]}
    problems = []
    for col, (sf_field, kind) in MAPPING.items():
        if col not in headers:
            problems.append(f"  Excel column '{col}' not found in the spreadsheet headers ({headers}).")
        if sf_field not in fields_by_name:
            problems.append(f"  Salesforce field '{sf_field}' (for column '{col}') does not exist.")
            continue
        f = fields_by_name[sf_field]
        if kind == "lookup" and f["type"] != "reference":
            problems.append(f"  Field '{sf_field}' mapped as lookup but is type '{f['type']}'.")
        if kind == "picklist" and f["type"] not in ("picklist", "multipicklist"):
            problems.append(f"  Field '{sf_field}' mapped as picklist but is type '{f['type']}'.")
    if problems:
        print("MAPPING PROBLEMS - fix mapping.py using fields.txt:\n" + "\n".join(problems))
        sys.exit(1)
    return fields_by_name


# --------------------------------------------------------------------------- #
# Lookup index building
# --------------------------------------------------------------------------- #
def _lookup_exact(sf, obj, name):
    rows = sf.query_all(f"SELECT Id, Name FROM {obj} WHERE Name = '{name.replace(chr(39), chr(92)+chr(39))}' LIMIT 1")["records"]
    return (rows[0]["Name"], rows[0]["Id"]) if rows else (None, None)


def build_account_index(sf, records, learned):
    """Resolve account names -> AccountId. Exact match first (preferring
    Type=ACCOUNT_PREFER_TYPE on dupes), then learned picks, aliases, typo-correction.

    Returns (index, real_name, corrections, suggestions, unmatched) where
      index[typed_lower]      = AccountId
      real_name[typed_lower]  = the actual Salesforce Account name (used for record Name)
      corrections             = [(typed, matched_name, score, how)]
      suggestions             = {typed: [(name, id, score), ...]}  (needs confirmation)
    """
    names = {clean_text(r.get(ACCOUNT_COLUMN)) for r in records if clean_text(r.get(ACCOUNT_COLUMN))}
    index, real_name, dupes = {}, {}, set()
    # exact (case-insensitive) pass
    for chunk in _chunks(sorted(names), 200):
        quoted = ", ".join("'" + n.replace("'", "\\'") + "'" for n in chunk)
        soql = f"SELECT Id, Name, Type FROM Account WHERE Name IN ({quoted})"
        for row in sf.query_all(soql)["records"]:
            key = row["Name"].strip().lower()
            if key in index and index[key] != row["Id"]:
                dupes.add(row["Name"])
                if row.get("Type") == ACCOUNT_PREFER_TYPE:
                    index[key] = row["Id"]
                    real_name[key] = row["Name"]
                continue
            index.setdefault(key, row["Id"])
            real_name.setdefault(key, row["Name"])
    if dupes:
        print(f"  NOTE: duplicate Account names (preferred Type='{ACCOUNT_PREFER_TYPE}'): "
              f"{', '.join(sorted(dupes))}")

    # learned + alias + fuzzy pass for whatever didn't match exactly
    corrections, suggestions = [], {}
    for typed in sorted(n for n in names if n.lower() not in index):
        key = typed.lower()
        remembered = lookup_account(learned, typed)
        if remembered:
            index[key], real_name[key] = remembered["id"], remembered["name"]
            corrections.append((typed, remembered["name"], 1.0, "learned"))
            continue
        alias = ACCOUNT_ALIASES.get(key)
        if alias:
            aname, aid = _lookup_exact(sf, "Account", alias)
            if aid:
                index[key], real_name[key] = aid, aname
                corrections.append((typed, aname, 1.0, "alias"))
                continue
        matches = closest_matches(sf, "Account", typed)
        # "clean" = typed is the account minus only company-suffix words (Asgore -> Asgore Partners)
        clean = [m for m in matches if is_clean_partial(typed, m[0])]
        runner_up = matches[1][2] if len(matches) > 1 else 0
        if len(clean) == 1 and clean[0][2] >= FUZZY_SUGGEST_THRESHOLD:
            name, rid, score = clean[0]
            index[key], real_name[key] = rid, name
            corrections.append((typed, name, score, "abbrev"))
        elif matches and matches[0][2] >= FUZZY_AUTO_THRESHOLD and runner_up < FUZZY_AUTO_THRESHOLD:
            name, rid, score = matches[0]
            index[key], real_name[key] = rid, name
            corrections.append((typed, name, score, "typo"))
        elif matches and matches[0][2] >= FUZZY_SUGGEST_THRESHOLD:
            suggestions[typed] = matches[:3]  # ambiguous -> confirm (interactive prompt below)

    unmatched = sorted(n for n in names if n.lower() not in index)
    return index, real_name, corrections, suggestions, unmatched


def build_contact_indexes(sf, records):
    """Return (scoped, by_name):
       scoped  = {(AccountId, name_lower): [ContactId, ...]}
       by_name = {name_lower: [ContactId, ...]}  (for fallback / ambiguity checks)"""
    names = {clean_text(r.get(CONTACT_COLUMN)) for r in records if clean_text(r.get(CONTACT_COLUMN))}
    scoped, by_name = {}, {}
    for chunk in _chunks(sorted(names), 200):
        quoted = ", ".join("'" + n.replace("'", "\\'") + "'" for n in chunk)
        soql = f"SELECT Id, Name, AccountId FROM Contact WHERE Name IN ({quoted})"
        for row in sf.query_all(soql)["records"]:
            key = row["Name"].strip().lower()
            by_name.setdefault(key, []).append(row["Id"])
            if row.get("AccountId"):
                scoped.setdefault((row["AccountId"], key), []).append(row["Id"])
    return scoped, by_name


# --------------------------------------------------------------------------- #
# Deal resolution (the deal is given at run time and applied to every row)
# --------------------------------------------------------------------------- #
def resolve_target_deal(sf, desc, deal_arg, deal_id_arg):
    fields_by_name = {f["name"]: f for f in desc["fields"]}
    target = fields_by_name[DEAL_FIELD]["referenceTo"][0]
    if deal_id_arg:
        rows = sf.query_all(f"SELECT Id, Name FROM {target} WHERE Id = '{deal_id_arg}'")["records"]
        if not rows:
            sys.exit(f"ERROR: no {target} record with Id '{deal_id_arg}'.")
        chosen = rows[0]
    else:
        text = deal_arg.replace("'", "")
        rows = sf.query_all(
            f"SELECT Id, Name FROM {target} WHERE Name LIKE '%{text}%' ORDER BY Name LIMIT 25"
        )["records"]
        if not rows:
            sys.exit(f"ERROR: no {target} matches '{deal_arg}'. Try different text or --deal-id.")
        chosen = rows[0] if len(rows) == 1 else _pick("Multiple Deals match", rows)
    print(f"\n==> All rows attach to Deal: '{chosen['Name']}'  (Id {chosen['Id']})\n")
    return chosen["Id"], chosen["Name"]


def _pick(prompt, rows, allow_skip=False):
    """Interactive chooser; requires a real terminal. Returns the chosen row, or
    None if allow_skip and the user skips. Quits the program on 'q'."""
    print(f"\n{prompt}:")
    for n, r in enumerate(rows, 1):
        print(f"  [{n}] {r['Name']}")
    if not sys.stdin.isatty():
        if allow_skip:
            return None
        sys.exit("  (non-interactive run; re-run with --deal-id to choose exactly.)")
    extra = "/s to skip (leave blank)" if allow_skip else ""
    while True:
        pick = input(f"Type the number{extra} (or q to quit): ").strip().lower()
        if pick == "q":
            sys.exit("Cancelled.")
        if allow_skip and pick == "s":
            return None
        if pick.isdigit() and 1 <= int(pick) <= len(rows):
            return rows[int(pick) - 1]
        print("  invalid choice, try again.")


# --------------------------------------------------------------------------- #
# Record building
# --------------------------------------------------------------------------- #
def _match_picklist(val, allowed):
    for label, value in allowed.items():
        if val.lower() == label.lower() or val.lower() == value.lower():
            return value
    return None


def resolve_contact(sf, cname, acct_id, scoped, by_name, info, rownum, interactive, learned):
    """Return a ContactId or None. Prefer a remembered pick, then the contact in
    this row's Account, a unique global name, alias/typo within the account; ask
    (or report) when genuinely ambiguous - and remember any interactive choice."""
    key = cname.lower()
    # 0) a choice you made before for this name at this account
    remembered = lookup_contact(learned, acct_id, cname)
    if remembered:
        info["by_account"] += 1
        return remembered["id"]
    # 1) scoped to the row's account
    if acct_id and (acct_id, key) in scoped:
        info["by_account"] += 1
        return scoped[(acct_id, key)][0]
    ids = by_name.get(key, [])
    # 2) unique name anywhere
    if len(ids) == 1:
        info["by_unique_name"] += 1
        return ids[0]
    # 3) not found -> alias, then fuzzy WITHIN this row's account (safe)
    if not ids:
        alias = CONTACT_ALIASES.get(key)
        if alias:
            scoped_ids = scoped.get((acct_id, alias.lower())) if acct_id else None
            pick = (scoped_ids or by_name.get(alias.lower()) or [None])[0]
            if pick:
                info["corrections"].append((rownum, cname, alias, 1.0, "alias"))
                return pick
        # fuzzy/abbreviation/first-name match WITHIN this row's account (safe)
        fz = closest_contacts_in_account(sf, cname, acct_id)
        runner_up = fz[1][2] if len(fz) > 1 else 0
        if fz and fz[0][2] >= FUZZY_AUTO_THRESHOLD and runner_up < FUZZY_AUTO_THRESHOLD:
            name, rid, score = fz[0]
            info["corrections"].append((rownum, cname, name, score, "match"))
            return rid
        if fz and fz[0][2] >= FUZZY_SUGGEST_THRESHOLD:
            # e.g. a first name matching several people in the account
            if interactive:
                rows = [{"Name": f"{n}  ({int(s*100)}%)", "Id": i, "_name": n} for n, i, s in fz[:6]]
                chosen = _pick(f"row {rownum}: Contact '{cname}' - pick the right person", rows, allow_skip=True)
                if chosen:
                    remember_contact(learned, acct_id, cname, chosen["_name"], chosen["Id"])
                    info["corrections"].append((rownum, cname, chosen["_name"], 0.0, "chosen"))
                    return chosen["Id"]
            else:
                info["suggestions"].append((rownum, cname, [(n, i, s) for n, i, s in fz[:3]]))
            return None
        info["unmatched"].append((rownum, cname))
        return None
    # 4) ambiguous: several contacts with this exact name, none in this account
    if interactive:
        rows = _contact_candidates(sf, ids)
        chosen = _pick(f"row {rownum}: which '{cname}'?", rows, allow_skip=True)
        if chosen:
            remember_contact(learned, acct_id, cname, chosen["Name"], chosen["Id"])
            info["by_account"] += 1
            return chosen["Id"]
    info["ambiguous"].append((rownum, cname, ids))
    return None


def _contact_candidates(sf, ids):
    quoted = ", ".join(f"'{i}'" for i in ids)
    out = []
    for r in sf.query_all(f"SELECT Id, Name, Account.Name FROM Contact WHERE Id IN ({quoted})")["records"]:
        acct = (r.get("Account") or {}).get("Name") or "no account"
        out.append({"Id": r["Id"], "Name": f"{r['Name']} @ {acct}"})
    return out


def build_records(sf, records, desc, account_index, real_name, scoped, by_name,
                  deal_id, property_text, interactive, learned):
    fields_by_name = {f["name"]: f for f in desc["fields"]}
    built, skipped = [], []
    info = {"by_account": 0, "by_unique_name": 0, "unmatched": [], "ambiguous": [],
            "blank": 0, "corrections": [], "suggestions": []}

    for i, rec in enumerate(records, start=2):
        out = dict(DEFAULTS)
        out[DEAL_FIELD] = deal_id
        skip_reason = None

        # Computed Name = "<abbreviated account> - <property>". Use the real
        # Salesforce account name when we resolved one (fixes typos in the Name too).
        acct_name = clean_text(rec.get(ACCOUNT_COLUMN))
        lender_src = real_name.get(acct_name.lower(), acct_name) if acct_name else None
        out[NAME_FIELD] = abbreviate_lender(lender_src) + NAME_JOINER + property_text if lender_src else property_text

        for col, (sf_field, kind) in MAPPING.items():
            raw = rec.get(col)
            if col == ACCOUNT_COLUMN:
                aid = account_index.get(acct_name.lower()) if acct_name else None
                if aid:
                    out[sf_field] = aid
                continue
            if col == CONTACT_COLUMN:
                cname = clean_text(raw)
                if not cname:
                    info["blank"] += 1
                    continue
                aid = account_index.get(acct_name.lower()) if acct_name else None
                cid = resolve_contact(sf, cname, aid, scoped, by_name, info, i, interactive, learned)
                if cid:
                    out[sf_field] = cid
                continue
            if kind == "text":
                val = clean_text(raw)
                if val is not None:
                    out[sf_field] = val
            elif kind == "picklist":
                val = clean_text(raw)
                if val is not None:
                    allowed = {v["label"]: v["value"]
                               for v in fields_by_name[sf_field]["picklistValues"] if v["active"]}
                    match = _match_picklist(val, allowed)
                    if match is None:
                        skip_reason = f"Interest '{val}' is not a valid option {list(allowed)}"
                        break
                    out[sf_field] = match
                    if col == INTEREST_COLUMN and match == DECLINED_INTEREST_VALUE:
                        out[DECLINED_REASON_FIELD] = DECLINED_REASON_DEFAULT

        if skip_reason:
            skipped.append((i, skip_reason))
        else:
            built.append((i, out))
    return built, skipped, info


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Bulk-create Deal Source records from Excel (new format).")
    ap.add_argument("excel", help="Path to the lender list .xlsx file")
    ap.add_argument("--deal", help="Deal name to search for (applied to every row).")
    ap.add_argument("--deal-id", help="Exact Deal record Id (no ambiguity).")
    ap.add_argument("--property", dest="property_text", required=True,
                    help="Property text for the record Name: '<Lender> - <Property>'.")
    ap.add_argument("--sheet", help="Worksheet name or 1-based number (e.g. --sheet 2). Default: first sheet.")
    ap.add_argument("--dry-run", action="store_true", help="Preview only; write nothing.")
    args = ap.parse_args()

    if not (args.deal or args.deal_id):
        sys.exit("ERROR: provide --deal \"name\" or --deal-id <Id> (the deal is required).")

    cfg = load_config()
    api_name = get_object_api_name(cfg)
    if not api_name:
        sys.exit("ERROR: set [object] api_name in config.ini first (run discover.py).")

    headers, records = read_excel(args.excel, args.sheet)
    print(f"Read {len(records)} data rows from {args.excel}\n")

    sf = connect(cfg)
    print(f"Connected as {cfg['salesforce']['username']}; target object: {api_name}\n")

    desc = describe_object(sf, api_name)
    validate_mapping(desc, headers)
    print("Mapping validated. Resolving deal, accounts and contacts...\n")

    learned = load_learned()
    deal_id, deal_name = resolve_target_deal(sf, desc, args.deal, args.deal_id)
    account_index, real_name, acct_corr, acct_sugg, acct_unmatched = build_account_index(sf, records, learned)
    scoped, by_name = build_contact_indexes(sf, records)

    interactive = sys.stdin.isatty() and not args.dry_run

    # Let the user resolve ambiguous Accounts up front (real terminal runs only).
    if interactive and acct_sugg:
        for typed in list(acct_sugg.keys()):
            rows = [{"Name": f"{n}  ({int(s*100)}%)", "Id": i, "_name": n}
                    for n, i, s in acct_sugg[typed]]
            chosen = _pick(f"Account '{typed}' didn't match exactly - pick the right one", rows, allow_skip=True)
            if chosen:
                account_index[typed.lower()] = chosen["Id"]
                real_name[typed.lower()] = chosen["_name"]
                remember_account(learned, typed, chosen["_name"], chosen["Id"])
                acct_corr.append((typed, chosen["_name"], 0.0, "chosen"))
                del acct_sugg[typed]

    built, skipped, info = build_records(sf, records, desc, account_index, real_name, scoped, by_name,
                                         deal_id, args.property_text, interactive, learned)

    # --- Typo / correction reports ---
    if acct_corr:
        print("\nACCOUNT auto-corrections (review!):")
        for typed, matched, score, how in acct_corr:
            print(f"    '{typed}' -> '{matched}'  ({how}, {int(score*100)}% match)")
    if acct_sugg:
        print("\nACCOUNT not matched - did you mean? (left blank, confirm to use):")
        for typed, cands in acct_sugg.items():
            opts = "; ".join(f"{n} ({int(s*100)}%)" for n, _i, s in cands)
            print(f"    '{typed}' -> {opts}")
    if info["corrections"]:
        print("\nCONTACT auto-corrections (review!):")
        for r, typed, matched, score, how in info["corrections"]:
            print(f"    row {r}: '{typed}' -> '{matched}'  ({how}, {int(score*100)}% match)")
    if info["suggestions"]:
        print("\nCONTACT not matched - did you mean? (left blank, confirm to use):")
        for r, typed, cands in info["suggestions"]:
            opts = "; ".join(f"{n} ({int(s*100)}%)" for n, _i, s in cands)
            print(f"    row {r}: '{typed}' -> {opts}")

    # Reports
    if acct_unmatched:
        print(f"\nUNMATCHED Accounts ({len(acct_unmatched)}) - left blank, Name still uses the text: "
              f"{', '.join(acct_unmatched)}")
    print(f"\nContacts matched in-account: {info['by_account']}; by unique name: {info['by_unique_name']}; "
          f"corrected: {len(info['corrections'])}; ambiguous (blank): {len(info['ambiguous'])}; "
          f"not found (blank): {len(info['unmatched'])}; no contact in row: {info['blank']}")
    for r, n, ids in info["ambiguous"]:
        print(f"    row {r}: Contact '{n}' ambiguous ({len(ids)} candidates, none in this account) -> left blank")
    for r, n in info["unmatched"]:
        print(f"    row {r}: Contact '{n}' not found -> left blank")

    print(f"\n{len(built)} rows ready to create, {len(skipped)} skipped.")
    if skipped:
        for rownum, reason in skipped:
            print(f"  row {rownum}: {reason}")

    if args.dry_run:
        print("\n--- DRY RUN: first 5 records that WOULD be created ---")
        for rownum, rec in built[:5]:
            print(f"  row {rownum}: Name={rec.get(NAME_FIELD)!r} | "
                  f"Interest={rec.get('ascendix__Interest__c')} | "
                  f"Account={'set' if rec.get('ascendix__Account__c') else 'BLANK'} | "
                  f"Contact={'set' if rec.get('ascendix__Contact__c') else 'BLANK'}"
                  + (f" | DeclinedReason={rec.get(DECLINED_REASON_FIELD)}" if rec.get(DECLINED_REASON_FIELD) else ""))
        print("\nNothing was written. Re-run without --dry-run to create the records.")
        return

    if not built:
        print("\nNothing to create. Exiting.")
        return

    payload = [rec for _, rec in built]
    if interactive:
        if input(f"\nProceed to create {len(payload)} records on '{deal_name}'? (y/n): ").strip().lower() != "y":
            sys.exit("Cancelled - nothing written.")
    print(f"\nCreating {len(payload)} records in Salesforce...")
    results = getattr(sf.bulk, api_name).insert(payload, batch_size=200)
    ok = sum(1 for r in results if r.get("success"))
    fail = len(results) - ok
    print(f"\nDONE: {ok} created, {fail} failed.")
    if fail:
        print("\nFailures:")
        for (rownum, _), res in zip(built, results):
            if not res.get("success"):
                print(f"  row {rownum}: {res.get('errors')}")


if __name__ == "__main__":
    main()
