"""Property Info — a cover sheet with PROPERTY INFORMATION / BUILDING INFORMATION
blocks (user's exact format: blue headers, gold value cells), plus a web lookup:
give an APN (and optional name/address hint) and Claude researches the county
assessor / ZIMAS / listing sites via web search and fills what it can find.
Unknown fields stay blank (gold = fill-in)."""

import json
import re

from tools import hist_llm

# (row label, dict key) in exact sheet order
PROP_FIELDS = [
    ("Property Name", "property_name"),
    ("Property Address", "address_line1"),      # line 2 goes on the next row
    ("County", "county"),
    ("APN", "apn"),
    ("Zoning", "zoning"),
    ("Land Area (acres)", "land_acres"),
    ("Land Area (SF)", "land_sf"),
    ("Parking", "parking"),
    ("Number of Spaces", "parking_spaces"),
]
BLDG_FIELDS = [
    ("Property Type", "property_type"),
    ("Year Built", "year_built"),
    ("Number of Buildings", "num_buildings"),
    ("Number of Stories", "num_stories"),
    ("Number of Units", "num_units"),
    ("Gross SF", "gross_sf"),
    ("Leasable SF", "leasable_sf"),
]
ALL_KEYS = [k for _, k in PROP_FIELDS] + ["address_line2"] + [k for _, k in BLDG_FIELDS]


# ── building the info dict from whatever we have ────────────────────────────
def blank() -> dict:
    return dict.fromkeys(ALL_KEYS + ["verified_address"])


def _scalar(v):
    """The research reply is free-form JSON, so a field can come back as a list
    ("parking": ["Surface","Covered"]) or a nested object. openpyxl refuses to
    write those — flatten to something a cell can hold."""
    if v is None or isinstance(v, (str, int, float)) and not isinstance(v, bool):
        return v
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (list, tuple, set)):
        return ", ".join(str(x) for x in v if x not in (None, "")) or None
    if isinstance(v, dict):
        return ", ".join(f"{k}: {x}" for k, x in v.items() if x not in (None, "")) or None
    return str(v)


def has_any(info) -> bool:
    """True when there's something worth putting on a cover sheet."""
    return bool(info) and any(info.get(k) not in (None, "") for k in ALL_KEYS)


def merge(*layers) -> dict:
    """Later layers win, but only where they actually carry a value — so a web
    lookup fills the gaps in what the documents already told us instead of
    blanking them out when the lookup comes back thin."""
    out = blank()
    for layer in layers:
        for k, v in (layer or {}).items():
            if k in out and v not in (None, ""):
                out[k] = v
    return out


def from_rent_roll(rr: dict) -> dict:
    """What the rent roll itself already told us about the property. This is
    what guarantees a Property Info tab even when the web lookup isn't run."""
    out = blank()
    if not rr:
        return out
    name = str(rr.get("property_name") or "").strip()
    if name:
        out["property_name"] = name
        if re.match(r"\s*\d", name):        # "2821 N Sierra St" — already an address
            out["address_line1"] = name
    csz = str(rr.get("city_state_zip") or "").strip()
    if csz:
        out["address_line2"] = csz
    if rr.get("apn"):
        out["apn"] = rr["apn"]
    units = rr.get("units") or []
    if units:
        out["num_units"] = len(units)
        sf = sum(float(u["unit_sf"]) for u in units if u.get("unit_sf"))
        if sf:
            out["leasable_sf"] = round(sf)
    return out

_PROMPT = """Research this property using web search and report what you find.

APN (Assessor's Parcel Number): {apn}
{zip_line}{hint}

Check the county assessor's parcel records (for Los Angeles County: the LA County
Assessor portal, and ZIMAS zimas.lacity.org for zoning), plus listing/data sites
(Redfin, Zillow, LoopNet, PropertyShark) as needed.

IDENTITY RULES — getting the RIGHT property matters more than filling fields:
- The APN is the unique parcel id within its county: treat it as ground truth.
- Street addresses are ambiguous (the same "14 Brooks Ave" exists in many
  cities). Only use an address-based source if it matches the APN, or the full
  address INCLUDING the ZIP code given above. If a source's city/ZIP conflicts
  with the given ZIP, discard that source entirely — do not mix properties.
- When sources disagree, prefer: county assessor > city planning (ZIMAS) >
  listing sites.
- If you cannot confirm a value is for THIS parcel, return null for it.

When done, output ONLY a JSON object (no other text after it) with these keys —
use null for anything you could not verify; do NOT guess:
- verified_address: the full address (street, city, ZIP) you confirmed for this
  APN — the user checks this to make sure you found the right property
- property_name: short name, usually the street address (e.g. "2821 Sierra")
- address_line1: street address (e.g. "2821 N SIERRA ST")
- address_line2: city, state zip (e.g. "Lincoln Heights, CA 90031")
- county: e.g. "Los Angeles"
- apn: echo the APN, digits/dashes as commonly written
- zoning: zoning code (e.g. "[Q]R1-1D-HCR")
- land_acres: number
- land_sf: number
- parking: e.g. "Surface", "Garage", or null
- parking_spaces: number or null
- property_type: e.g. "Multifamily", "Office", "Retail"
- year_built: e.g. 1923 or "1923/2026" if renovated
- num_buildings: number
- num_stories: number
- num_units: number
- gross_sf: number
- leasable_sf: number or null"""


def fetch(apn: str, hint: str = "", zip_code: str = "") -> dict:
    """Web-research the APN and return the property-info dict (missing keys None).
    zip_code disambiguates address-based sources (many cities share street names);
    the result includes 'verified_address' so the user can confirm the match.
    Web search is variable run to run — retries once if the first pass finds
    little, and returns whichever attempt filled more fields."""
    best = None
    for _attempt in range(2):
        out = _fetch_once(apn, hint, zip_code)
        n = sum(1 for k in ALL_KEYS if out.get(k) is not None)
        if best is None or n > best[0]:
            best = (n, out)
        if best[0] >= 6:                      # good enough — stop
            break
    return best[1]


def _fetch_once(apn: str, hint: str, zip_code: str) -> dict:
    resp = hist_llm._client().messages.create(
        model=hist_llm.MODEL,
        max_tokens=12000,
        thinking={"type": "adaptive"},
        tools=[{"type": "web_search_20260209", "name": "web_search", "max_uses": 8}],
        messages=[{"role": "user", "content": _PROMPT.format(
            apn=apn.strip() or "unknown — determine it from the full address below "
                               "(then treat that parcel as the subject)",
            zip_line=f"ZIP code: {zip_code.strip()}\n" if zip_code.strip() else "",
            hint=f"Known name/address hint: {hint}" if hint.strip() else "")}],
    )
    text = "".join(b.text for b in resp.content if b.type == "text")
    info = {k: _scalar(v) for k, v in _last_json(text).items()}
    if not info and text.strip():
        # The research happened but the closing JSON never landed (long web-search
        # turns get truncated, or the model answered in prose). Don't throw the
        # findings away — have it restate them as JSON, no searching this time.
        info = _restate(text)
    out = {k: info.get(k) for k in ALL_KEYS}
    out["verified_address"] = info.get("verified_address")
    return out


_RESTATE_SCHEMA = {
    "type": "object",
    "properties": {k: {"anyOf": [{"type": "string"}, {"type": "number"}, {"type": "null"}]}
                   for k in ALL_KEYS + ["verified_address"]},
    "required": ALL_KEYS + ["verified_address"],
    "additionalProperties": False,
}


def _restate(text: str) -> dict:
    """Convert a prose/partial research write-up into the info dict."""
    try:
        resp = hist_llm._client().messages.create(
            model=hist_llm.MODEL, max_tokens=4000,
            messages=[{"role": "user", "content":
                       "Below is a research write-up about one property. Report the facts it "
                       "states, as JSON. Use null for anything it does not state — do not "
                       "guess or add anything of your own.\n\n" + text[-12000:]}],
            output_config={"format": {"type": "json_schema", "schema": _RESTATE_SCHEMA}},
        )
        return hist_llm._json_response(resp) or {}
    except Exception:  # noqa: BLE001
        return {}


def _balanced_json(text: str, want: set, min_hits: int) -> dict:
    """Extract the final JSON object from a reply: scan '{' positions rightmost-
    first, take each balanced {...} span, and accept the first that parses to a
    dict containing at least min_hits of the wanted keys (guards against
    fragments and nested sub-objects)."""
    starts = [m.start() for m in re.finditer(r"\{", text)]
    for s in reversed(starts):
        depth = 0
        for i in range(s, len(text)):
            if text[i] == "{":
                depth += 1
            elif text[i] == "}":
                depth -= 1
                if depth == 0:
                    try:
                        d = json.loads(text[s:i + 1])
                        if isinstance(d, dict) and len(set(d) & want) >= min_hits:
                            return d
                    except Exception:
                        pass
                    break
    return {}


def _last_json(text: str) -> dict:
    return _balanced_json(text, set(ALL_KEYS) | {"verified_address"}, 3)


# ── address -> candidate properties (the "which 14 Brooks Ave?" dropdown) ────
_CAND_PROMPT = """Find the U.S. property or properties matching this street
address using web search.

Street address: {address}
{zip_line}
The same street address can exist in many different cities. List EVERY distinct
real property you can find matching this street address{zipnote}. For each:
- address: the street address as commonly written
- city_state_zip: e.g. "Venice, CA 90291"
- county: county name, or null
- apn: the parcel's APN if you can determine it, else null
- note: one short identifying hint (e.g. "12-unit multifamily on LoopNet"), or null

Output ONLY a JSON object (no text after it): {{"candidates": [ ... ]}}
Max 6 candidates, most likely first."""


def candidates(address: str, zip_code: str = "") -> list:
    """Web-search an ambiguous street address; returns candidate property dicts
    (address / city_state_zip / county / apn / note) for the user to pick from."""
    resp = hist_llm._client().messages.create(
        model=hist_llm.MODEL,
        max_tokens=6000,
        thinking={"type": "adaptive"},
        tools=[{"type": "web_search_20260209", "name": "web_search", "max_uses": 6}],
        messages=[{"role": "user", "content": _CAND_PROMPT.format(
            address=address,
            zip_line=f"ZIP code (if known): {zip_code.strip()}\n" if zip_code.strip() else "",
            zipnote=" (prioritize the given ZIP)" if zip_code.strip() else "")}],
    )
    text = "".join(b.text for b in resp.content if b.type == "text")
    d = _balanced_json(text, {"candidates"}, 1)
    return [c for c in (d.get("candidates") or []) if isinstance(c, dict) and c.get("address")]


def build_sheet(ws, info: dict):
    """Write the Property Info sheet in the user's exact layout."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    blue = PatternFill("solid", fgColor="305496")
    gold = PatternFill("solid", fgColor="FFC000")
    hdr = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="404040")
    box = Border(top=thin, bottom=thin, left=thin, right=thin)
    right = Alignment(horizontal="right", vertical="center")

    def header(row, c0, c1, title):
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
        c = ws.cell(row, c0, title); c.font = hdr; c.fill = blue
        for cc in range(c0, c1 + 1):
            ws.cell(row, cc).fill = blue; ws.cell(row, cc).border = box

    def field(row, lc, vc, label, val, fmt=None):
        ws.cell(row, lc, label).border = box
        for cc in range(lc + 1, vc + 1):
            cell = ws.cell(row, cc)
            cell.fill = gold; cell.border = box
        c = ws.cell(row, vc)
        val = _scalar(val)
        if val is not None and val != "":
            c.value = val
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt
        c.alignment = right

    NUMFMT = {"land_acres": "0.00", "land_sf": "#,##0.00", "parking_spaces": "0",
              "year_built": "0", "num_buildings": "0", "num_stories": "0",
              "num_units": "0", "gross_sf": "#,##0", "leasable_sf": "#,##0"}

    # left block — PROPERTY INFORMATION (labels A, values C)
    header(1, 1, 3, "PROPERTY INFORMATION")
    r = 2
    for label, key in PROP_FIELDS:
        field(r, 1, 3, label, info.get(key), NUMFMT.get(key))
        r += 1
        if key == "address_line1":               # second address line on its own row
            field(r, 1, 3, "", info.get("address_line2"))
            r += 1

    # right block — BUILDING INFORMATION (labels E, values G)
    header(1, 5, 7, "BUILDING INFORMATION")
    r = 2
    for label, key in BLDG_FIELDS:
        field(r, 5, 7, label, info.get(key), NUMFMT.get(key))
        r += 1

    for col, w in (("A", 20), ("B", 6), ("C", 24), ("D", 3), ("E", 22), ("F", 8), ("G", 18)):
        ws.column_dimensions[col].width = w
