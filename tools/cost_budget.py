"""Cost Budget — one property's cost basis, as a two-tab workbook.

One sponsor document in, one workbook out holding exactly its two tabs:

  S - Cost Budget - 3210    the sponsor's own cost basis, transcribed as printed
                            (col B = trade, col C = amount, section + TOTALS rows)
  W - Cost Budget - 3210    the underwriting sheet

The W sheet pairs every category down the left — `SOFT COSTS - SPONSOR` then
`SOFT COSTS - QCP` — and the right-hand QCP Budget column (J–P) rolls up off the
QCP blocks. Each SPONSOR row reads one row of the S sheet by formula:

    B40  ='S - Cost Budget - 3210'!C14      Architect

so the wiring is by line-item name. Building a property means: transcribe the
sponsor's document onto the S tab, clone the W template beside it, and re-point
every sponsor row at the row its trade actually landed on — a sponsor's document
never lists the trades in the template's order, and rarely lists all of them.

Only the sponsor's side carries numbers. Every QCP block is cleared — labels,
formatting, column headers and subtotal formulas stay, so each section reads zero
until it's underwritten — and so are the standalone input blocks, whose lender,
loan balance, rate and maturity date would otherwise be inherited from whichever
property was cloned. The QCP Budget column (J-P) is pure formula and is left
exactly as it is.
"""

import difflib
import io
import re
from copy import copy

from tools import hist_llm

# a tab pair is "<W|S> - Cost Budget - <id>"
_TAB = re.compile(r"^([WS])\s*-\s*Cost Budget\s*-\s*(.+?)\s*$", re.I)
# section banners in column A: "SOFT COSTS - SPONSOR", "FINANCING COSTS - QCP (REMAINING)"
_BANNER = re.compile(r"\s*[-(]?\s*\b(SPONSOR|QCP)\b.*$", re.I)
_TOTAL_ROW = re.compile(r"^\s*(total|subtotal)\b", re.I)
_SREF = re.compile(r"'(?P<sheet>S\s*-\s*Cost Budget\s*-\s*[^']*)'!\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)")

SOURCE_COL, LABEL_COL, AMOUNT_COL = 1, 2, 3        # S sheet: A unused, B trade, C amount


# ── the workbook's tab pairs ─────────────────────────────────────────────────
def pairs(wb) -> list:
    """[(id, W title, S title)] for every complete pair, in workbook order."""
    seen = {}
    for name in wb.sheetnames:
        m = _TAB.match(name)
        if m:
            seen.setdefault(m.group(2), {})[m.group(1).upper()] = name
    return [(pid, d["W"], d["S"]) for pid, d in seen.items() if "W" in d and "S" in d]


def sponsor_categories(wb, template_w: str) -> list:
    """The SPONSOR block names on a template — the categories a line can land in."""
    return [name for _s, _e, name, side in blocks(wb[template_w]) if side == "SPONSOR"]


def _norm(label) -> str:
    s = re.sub(r"[^a-z0-9 ]+", " ", str(label or "").lower())
    return re.sub(r"\s+", " ", s).strip()


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


# ── reading the sponsor's document ───────────────────────────────────────────
_EXTRACT_PROMPT = """This is a property COST BASIS / cost budget from a project
sponsor — the money spent and committed on a development so far. Transcribe it
faithfully, top to bottom, exactly as printed. Do not compute, merge, reorder or
invent anything.

A messy budget carries several number columns — unit cost, quantity, a monthly
draw schedule, spent-to-date, remaining balance. Pick the ONE column that states
each line's total committed cost (headed Amount, Total, Total Cost, SPENT, Budget
or similar) and read every line from that same column. Never take a number from a
per-month, per-SF, per-unit or percentage column. Report which column you used in
amount_header, exactly as its heading is printed — "" if the document has only one
number column and no heading for it.

Return one entry per line of the document:
- label: the line exactly as printed (a trade name, a section title, a total line)
- amount: that line's number from the column you chose, or null if it has none.
  Strip currency symbols and commas; a figure in parentheses is negative.
- kind: "section" for a block title (ACQUISITION COST, SOFT COST, HARD COST,
  LEASING/TI, CAPITALIZED CARRY …), "header" for a column-heading line (TRADE,
  Total, SPENT), "item" for a normal cost line, "total" for a TOTALS/subtotal line

Also return property_name: the property this budget is for, as printed (e.g.
"3210 SUNSET BL"), or "" if it isn't stated.

- category: which of the underwriting categories below this cost belongs in,
  judged by what the cost IS — a GMP trade is a hard cost, title insurance and
  legal at closing are acquisition costs, a broker fee on the loan is a financing
  cost. Use "" for section, header and total lines, and for anything that fits
  none of them.

CATEGORIES:
{categories}

Keep every line, including ones with no amount. Scanned or messy documents: read
what is actually there — never guess a number you cannot see."""

_SCHEMA = {
    "type": "object",
    "properties": {
        "property_name": {"type": "string"},
        "amount_header": {"type": "string"},
        "rows": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "label": {"type": "string"},
                    "amount": {"anyOf": [{"type": "number"}, {"type": "null"}]},
                    "kind": {"type": "string", "enum": ["section", "header", "item", "total"]},
                    "category": {"type": "string"},
                },
                "required": ["label", "amount", "kind", "category"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["property_name", "amount_header", "rows"],
    "additionalProperties": False,
}


def extract_source(data: bytes, categories=None) -> dict:
    """Claude reads the sponsor's cost basis — PDF (however messy) or spreadsheet
    — and tags each line with the category it belongs in, in the same pass. The
    monthly draw schedule a development budget carries is dropped before sending:
    it is most of the document by volume and says nothing about what a cost is."""
    from tools import tabular
    prompt = _EXTRACT_PROMPT.format(
        categories="\n".join(f"- {c}" for c in (categories or [])) or "(none given)")
    with hist_llm._client().messages.stream(
        model=hist_llm.MODEL,
        max_tokens=24000,
        thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": hist_llm.content_blocks(
            data, prompt, transform=tabular.drop_period_columns)}],
        output_config={"format": {"type": "json_schema", "schema": _SCHEMA}},
    ) as stream:
        return hist_llm._json_response(stream.get_final_message())


# ── writing the S tab ────────────────────────────────────────────────────────
def _row_styles(ws) -> dict:
    """Style of a representative row of each kind, harvested from the template S
    tab so a new one looks identical without hand-built formatting."""
    kinds = {}
    for r in range(1, (ws.max_row or 0) + 1):
        label = ws.cell(r, LABEL_COL).value
        if label in (None, ""):
            continue
        text = str(label).strip()
        if r <= 3:
            kinds.setdefault("title", r)
        elif _TOTAL_ROW.match(text) or text.upper().startswith("TOTALS"):
            kinds.setdefault("total", r)
        elif text.upper() in ("TRADE", "TOTAL", "SPENT"):
            kinds.setdefault("header", r)
        elif text.isupper() and ws.cell(r, AMOUNT_COL).value in (None, ""):
            kinds.setdefault("section", r)
        else:
            kinds.setdefault("item", r)
    return kinds


def write_source_tab(wb, template_s: str, sheet_title: str, parsed: dict) -> dict:
    """Clone the template S tab and write the transcription onto it. Returns the
    index the W sheet aims at: line items by section, flat by label, and the full
    list for Claude to reason over."""
    ws = _clone(wb, wb[template_s], sheet_title)
    styles = _row_styles(wb[template_s])
    tmpl = wb[template_s]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1):
        for c in row:
            c.value = None

    def style_from(kind, r_out, source_row):
        if not source_row:
            return
        for col in (LABEL_COL, AMOUNT_COL):
            ws.cell(r_out, col)._style = copy(tmpl.cell(source_row, col)._style)

    title = str(parsed.get("property_name") or "").strip()
    r = 2
    if title:
        ws.cell(r, LABEL_COL, title)
        style_from("title", r, styles.get("title"))
        r += 2

    # Line items are indexed per section as well as flat: "Hazardous Materials"
    # appears under both SOFT and HARD cost on plenty of budgets, and a W row
    # must take the one from its own category.
    by_section, flat, items, cats, section = {}, {}, [], {}, ""
    for item in parsed.get("rows") or []:
        label = str(item.get("label") or "").strip()
        if not label:
            continue
        kind = item.get("kind") or "item"
        if kind == "section":
            if _norm(label) == _norm(title):
                continue                             # the sponsor's own title again
            section = label
            if r > 2:
                r += 1                               # a blank line before each block
        ws.cell(r, LABEL_COL, label)
        amount = item.get("amount")
        if amount is not None:
            ws.cell(r, AMOUNT_COL, float(amount))
        style_from(kind, r, styles.get(kind if kind != "header" else "header"))
        if kind == "item":
            spot = (r, AMOUNT_COL)
            by_section.setdefault(_norm(section), {}).setdefault(_norm(label), spot)
            flat.setdefault(_norm(label), spot)
            items.append((spot, section or "", label, amount))
            if str(item.get("category") or "").strip():
                cats[spot] = str(item["category"]).strip()
        r += 1
    return {"by_section": by_section, "flat": flat, "items": items, "category": cats}


# ── the S tab: the sponsor's document as given ───────────────────────────────
def paste_source_sheet(wb, src_ws, title: str):
    """Copy the sponsor's own sheet in cell for cell — values, formats, widths,
    merges. The S tab is the document as they sent it, not a tidied version of
    it: that is what the existing tabs in the workbook are, and it's what makes
    the W sheet auditable back to the source."""
    ws = wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for c in row:
            if c.value is None and not c.has_style:
                continue
            out = ws.cell(c.row, c.column, c.value)
            if c.has_style:
                out.font = copy(c.font)
                out.fill = copy(c.fill)
                out.border = copy(c.border)
                out.alignment = copy(c.alignment)
                out.number_format = c.number_format
                out.protection = copy(c.protection)
    for key, dim in src_ws.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[key].width = dim.width
        ws.column_dimensions[key].hidden = dim.hidden
    for idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            ws.row_dimensions[idx].height = dim.height
    for rng in src_ws.merged_cells.ranges:
        ws.merge_cells(str(rng))
    ws.freeze_panes = src_ws.freeze_panes
    ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    return ws


def _amount_near(ws, row, label_col, amount, max_col):
    """The cell on this row holding that amount — the sponsor's sheets put it
    anywhere from the next column to thirty columns over."""
    if amount is None:
        return None
    for c in range(label_col + 1, max_col + 1):
        v = ws.cell(row, c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool) \
                and abs(float(v) - float(amount)) < 0.01:
            return c
    return None


def _amount_column(ws, header: str):
    """The column Claude read its numbers from, found by its heading — so a line
    is pinned to the right column even when the same figure appears in a monthly
    draw or a spent-to-date column too."""
    if not str(header or "").strip():
        return None
    want = _norm(header)
    for r in range(1, min(ws.max_row or 0, 30) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _norm(v) == want:
                return c
    return None


def locate_items(ws, parsed: dict) -> dict:
    """Find where each line Claude read actually sits on the pasted sheet, so the
    W sheet can point at the real cell. Returns the same index shape as a
    transcribed tab, but with (row, column) instead of just a row."""
    max_row, max_col = ws.max_row or 0, ws.max_column or 0
    pref = _amount_column(ws, parsed.get("amount_header"))
    where = {}
    for r in range(1, max_row + 1):
        for c in range(1, min(max_col, 8) + 1):        # labels live on the left
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                where.setdefault(_norm(v), []).append((r, c))

    _blocks, block_of = document_blocks(parsed)
    by_section, flat, items, cats, section, taken = {}, {}, [], {}, "", set()
    for idx, item in enumerate(parsed.get("rows") or []):
        label = str(item.get("label") or "").strip()
        if not label:
            continue
        if item.get("kind") == "section":
            section = label
            continue
        section = block_of[idx] or section
        if item.get("kind") != "item" or _TOTAL_ROW.match(label):
            continue                          # a subtotal is not a line item
        spot = None
        for r, c in where.get(_norm(label), []):
            if (r, c) in taken:
                continue
            acol = None
            if pref and pref > c:                      # the column Claude read
                v = ws.cell(r, pref).value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    acol = pref
            if acol is None:
                acol = _amount_near(ws, r, c, item.get("amount"), max_col)
            if acol:
                spot = (r, acol)
                taken.add((r, c))
                break
        if not spot:
            continue                                   # couldn't be pinned down
        by_section.setdefault(_norm(section), {}).setdefault(_norm(label), spot)
        flat.setdefault(_norm(label), spot)
        items.append((spot, section or "", label, item.get("amount")))
        if str(item.get("category") or "").strip():
            cats[spot] = str(item["category"]).strip()
    return {"by_section": by_section, "flat": flat, "items": items, "category": cats}


# ── cloning a worksheet ──────────────────────────────────────────────────────
def _clone(wb, src, title: str):
    """copy_worksheet keeps values, styles, widths, merges and row heights —
    everything except the images, which are re-attached here. The copy needs its
    own bytes: two sheets sharing one image object exhaust the same file handle
    and the workbook then refuses to save."""
    from openpyxl.drawing.image import Image as XLImage
    ws = wb.copy_worksheet(src)
    ws.title = title
    for img in list(getattr(src, "_images", [])):
        try:
            # read the bytes without consuming them — img._data() hands the
            # buffer to PIL, which closes it, and then neither sheet can save
            ref = getattr(img, "ref", None)
            raw = ref.getvalue() if hasattr(ref, "getvalue") else img._data()
            fresh = XLImage(io.BytesIO(raw))
            fresh.anchor = copy(img.anchor)
            for attr in ("width", "height"):
                if getattr(img, attr, None):
                    setattr(fresh, attr, getattr(img, attr))
            ws._images.append(fresh)
        except Exception:  # noqa: BLE001
            pass                    # a missing logo beats an unsaveable workbook
    return ws



# ── build the sponsor blocks out of the sponsor's own lines ──────────────────
_RANGE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)(?::\$?[A-Z]{1,3}\$?(\d+))?")


def covered_rows(ws, start: int, end: int) -> list:
    """The rows a block's own totals add up. Writing anywhere else means the
    number is on the sheet but not in the total — the template has one such row
    (an acquisition "Closing Costs" line that sits outside its SUM), and that
    quirk shouldn't be inherited by every property built from it."""
    totals, covered = [], set()
    for r in range(start, end + 1):
        label = str(ws.cell(r, 1).value or "")
        if _TOTAL_ROW.match(label):
            totals.append(r)
    for r in totals:
        f = ws.cell(r, 2).value
        if not _is_formula(f):
            continue
        for m in _RANGE.finditer(f):
            lo = int(m.group(2))
            hi = int(m.group(3) or m.group(2))
            covered.update(range(lo, hi + 1))
    slots = [r for r in sorted(covered)
             if start <= r <= end and r not in totals
             and not str(ws.cell(r, 1).value or "").strip().lower().startswith("total")]
    if slots:
        return slots
    # Some blocks carry no working total at all — Developer Costs has no rows,
    # Interest Reserve's total is blank. Fall back to the block's body: from the
    # column-header row to the first total, or the end of the block.
    head = next((r for r in range(start, end + 1)
                 if str(ws.cell(r, 2).value or "").strip().lower() == "total"), start + 1)
    stop = next((r for r in totals if r > head), end + 1)
    return [r for r in range(head + 1, stop)]


_ASSIGN_PROMPT = """Sort a sponsor's cost lines into the categories an
underwriting sheet keeps them in.

CATEGORIES:
{cats}

THE SPONSOR'S LINES (section | line | amount):
{lines}

Put every line in exactly one category, using what the cost IS. A GMP trade is a
hard cost; title insurance and legal at closing are acquisition costs; a broker
fee on the loan is a financing cost; the interest reserve is its own category if
one exists. Leave a line out only if it belongs in none of them.

Echo each line exactly as given."""

_ASSIGN_SCHEMA = {
    "type": "object",
    "properties": {
        "assignments": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "category": {"type": "string"},
                    "lines": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["category", "lines"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["assignments"],
    "additionalProperties": False,
}


def assign_to_categories(categories, items) -> dict:
    """{category: [spot, ...]} — which of the sponsor's lines belong in each
    sponsor block, in the order the sponsor listed them."""
    if not categories or not items:
        return {}
    key = {f"{sec} | {lbl} | {'' if a is None else format(a, ',.0f')}": spot
           for spot, sec, lbl, a in items}
    loose = {f"{sec} | {lbl}": spot for spot, sec, lbl, _a in items}
    resp = hist_llm._client().messages.create(
        model=hist_llm.MODEL, max_tokens=16000, thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": _ASSIGN_PROMPT.format(
            cats="\n".join(f"- {c}" for c in categories),
            lines="\n".join(key))}],
        output_config={"format": {"type": "json_schema", "schema": _ASSIGN_SCHEMA}},
    )
    order = {spot: i for i, spot in enumerate(key.values())}
    out, taken = {}, set()
    valid = {_norm(c): c for c in categories}
    for a in hist_llm._json_response(resp).get("assignments") or []:
        cat = valid.get(_norm(a.get("category", "")))
        if not cat:
            continue
        picked = []
        for line in a.get("lines") or []:
            line = str(line).strip()
            spot = key.get(line) or loose.get(" | ".join(line.split(" | ")[:2]))
            if spot is not None and spot not in taken:
                taken.add(spot)
                picked.append(spot)
        if picked:
            out.setdefault(cat, []).extend(picked)
    for cat in out:
        out[cat].sort(key=lambda sp: order.get(sp, 0))
    return out


# ── the W tab ────────────────────────────────────────────────────────────────
def blocks(ws) -> list:
    """[(first row, last row, category, side)] for the SPONSOR/QCP banners in col A.
    The side marker is written several ways — 'SOFT COSTS - SPONSOR',
    'FINANCING COSTS - QCP (TO DATE ONLY)', 'EXISTING LOAN ASSUMPTIONS (QCP)' —
    so match on the keyword rather than a fixed shape."""
    marks = []
    for r in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(r, 1).value
        if not (isinstance(v, str) and len(v.strip()) > 4 and v.strip() == v.strip().upper()):
            continue
        t = v.strip()
        side = "SPONSOR" if "SPONSOR" in t else ("QCP" if "QCP" in t else None)
        if side:
            marks.append((r, _BANNER.sub("", t).strip(" -"), side))
    out = []
    for i, (r, name, side) in enumerate(marks):
        end = marks[i + 1][0] - 1 if i + 1 < len(marks) else (ws.max_row or r)
        out.append((r, end, name, side))
    return out


def _blank_qcp_blocks(ws) -> int:
    """Everything that isn't the sponsor's side starts empty. Line-item cells are
    cleared — including the standalone input blocks, whose lender name, loan
    balance, rate and maturity date would otherwise be inherited from the
    template property. Labels, formatting, column headers and the subtotal
    formulas stay, so each section reads zero until it's underwritten."""
    cleared = 0
    for start, end, name, side in blocks(ws):
        if side == "SPONSOR":
            continue
        for r in range(start, end + 1):
            if r == start:
                continue                              # the block's banner stays
            label = ws.cell(r, 1).value
            if isinstance(label, str) and _TOTAL_ROW.match(label.strip()):
                continue                              # totals keep name and =SUM
            for col in (1, 2, 3, 4):                  # line-item labels go too
                if ws.cell(r, col).value is not None:
                    ws.cell(r, col).value = None
                    cleared += 1
    return cleared


NOTES_COL = 6


def clear_notes(ws) -> int:
    """Empty the Notes column. What it holds is carried over from the property
    that was cloned — "Sponsor" against rows this sponsor may not even report,
    and a leftover instruction to whoever underwrote that deal. The "Notes"
    column heading stays."""
    cleared = 0
    for r in range(1, (ws.max_row or 0) + 1):
        c = ws.cell(r, NOTES_COL)
        if c.value in (None, "") or str(c.value).strip().lower() == "notes":
            continue
        c.value = None
        cleared += 1
    return cleared


def _zero_hardcoded(ws) -> list:
    """Anything still typed into the sheet outside the sponsor's side is a value
    inherited from the template property, not something this deal knows yet."""
    zeroed = []
    for start, end, name, side in blocks(ws):
        if side == "SPONSOR":
            continue
        for r in range(start, end + 1):
            for col in (2, 3, 4):
                c = ws.cell(r, col)
                if c.value not in (None, "") and not _is_formula(c.value) \
                        and str(ws.cell(r, 1).value or "").strip():
                    zeroed.append(f"{ws.cell(r, 1).value or ''} {c.coordinate}={c.value}".strip())
                    c.value = None
    for r in range(1, (ws.max_row or 0) + 1):        # the QCP Budget column, J–P
        for col in range(10, 17):
            c = ws.cell(r, col)
            if isinstance(c.value, (int, float)) and not isinstance(c.value, bool) and c.value:
                zeroed.append(f"QCP Budget {c.coordinate}={c.value}")
                c.value = 0
    return zeroed


def document_blocks(parsed: dict):
    """Split the sponsor's document into cost blocks the way it actually reads:
    a block runs until the total that closes it. Via Roca prints its soft costs,
    totals them, then carries on listing TI and leasing under the same heading —
    comparing that heading's total against everything beneath it is what made a
    section look 362% covered.

    A total that closes nothing (a restatement like "Total Hard Costs" right
    after "Subtotal Hard Costs", or a grand total) is not a block of its own.

    Returns (blocks, block name per row of parsed["rows"])."""
    grand = re.compile(r"\b(development|project|all|grand)\b", re.I)
    rows = parsed.get("rows") or []
    blocks, names = [], [""] * len(rows)
    section = ""
    cur = {"name": "", "printed": None, "items": 0.0, "n": 0, "rows": []}

    used = set()

    def close(printed, label):
        nonlocal cur
        # an unclosed run of zero-value rows is nothing to reconcile
        if cur["n"] and not (printed is None and abs(cur["items"]) < 0.01):
            name = re.sub(r"^\s*(sub)?totals?\s*(for)?\s*", "", label or "", flags=re.I).strip()
            name = name.strip(":- ") or cur["name"] or section or "(unnamed)"
            while name in used:              # two blocks can't share a name — the
                name += " (cont.)"           # tie-out keys off it
            used.add(name)
            cur["name"] = name
            cur["printed"] = printed
            blocks.append(cur)
            for i in cur["rows"]:
                names[i] = cur["name"]
        cur = {"name": section, "printed": None, "items": 0.0, "n": 0, "rows": []}

    for i, row in enumerate(rows):
        label = str(row.get("label") or "").strip()
        amt = row.get("amount")
        kind = row.get("kind")
        if not label:
            continue                       # a recap figure with no line name
        if kind == "section":
            close(None, cur["name"])
            section = label
            cur["name"] = section
            continue
        is_total = kind == "total" or _TOTAL_ROW.match(label)
        if is_total and isinstance(amt, (int, float)):
            # a roll-up restates a total already accounted for: it closes nothing
            # ("Total Hard Costs" straight after "Subtotal Hard Costs"), or it is
            # the grand total of the whole budget
            if grand.search(label) or cur["n"] == 0 \
                    or (abs(cur["items"]) < 0.01 and abs(float(amt)) > 0.01):
                continue
            close(float(amt), label)
        elif kind == "item" and isinstance(amt, (int, float)):
            cur["items"] += float(amt)
            cur["n"] += 1
            cur["rows"].append(i)
    close(None, cur["name"])
    return blocks, names



def fill_from_source(ws, s_title: str, index: dict, use_llm: bool = True) -> dict:
    """Rebuild every SPONSOR block out of the sponsor's own lines: their labels,
    their order, each wired to the cell it came from. A trade the sponsor never
    mentioned simply isn't a row — the template's own list is a starting point,
    not a form to be filled in with zeros.

    The paired QCP block takes the same labels so the two sides line up, but no
    numbers; the QCP Budget column's labels follow it."""
    from openpyxl.utils import get_column_letter
    found = blocks(ws)
    sponsor = [(st, en, name) for st, en, name, side in found if side == "SPONSOR"]
    items = index.get("items", [])
    # the extraction already tagged each line with its category — no second pass
    assign, valid = {}, {_norm(n): n for _s, _e, n in sponsor}
    for spot, _sec, _lbl, _a in items:
        cat = valid.get(_norm(index.get("category", {}).get(spot, "")))
        if cat:
            assign.setdefault(cat, []).append(spot)
    if not assign and use_llm and hist_llm.available() and items:
        try:                                   # older parse with no tags
            assign = assign_to_categories([n for _s, _e, n in sponsor], items)
        except Exception:  # noqa: BLE001
            assign = {}
    if not assign:                       # fall back to the sponsor's own headings
        for spot, sec, _l, _a in items:
            near = difflib.get_close_matches(_norm(sec), [_norm(n) for _s, _e, n in sponsor],
                                             n=1, cutoff=0.45)
            if near:
                name = next(n for _s, _e, n in sponsor if _norm(n) == near[0])
                assign.setdefault(name, []).append(spot)

    label_of = {spot: lbl for spot, _s, lbl, _a in items}
    amount_of = {spot: a for spot, _s, _l, a in items}
    placed, overflow, wiring, seen = [], [], [], set()
    for st, en, name in sponsor:
        slots = covered_rows(ws, st, en)
        # A leftover line item still points at the property we cloned from, so it
        # has to go — but the block's furniture stays: its banner, the
        # Total/To Date/Remaining column headers, and the subtotal rows. Wiping
        # those was what collapsed the sheet into one undivided block.
        header = next((r for r in range(st, en + 1)
                       if str(ws.cell(r, 2).value or "").strip().lower() == "total"), st)
        for r in range(header + 1, en + 1):
            label = str(ws.cell(r, 1).value or "").strip()
            if r in slots or not label or _TOTAL_ROW.match(label):
                continue
            for col in (1, 2, 3, 4):
                ws.cell(r, col).value = None
        picks = [sp for sp in assign.get(name, []) if not (sp in seen or seen.add(sp))]
        for i, row in enumerate(slots):
            spot = picks[i] if i < len(picks) else None
            if spot is None:
                for col in (1, 2, 3, 4):
                    ws.cell(row, col).value = None
            else:
                r_, c_ = spot
                ws.cell(row, 1, label_of.get(spot, ""))
                ws.cell(row, 2, f"='{s_title}'!{get_column_letter(c_)}{r_}")
                ws.cell(row, 3, f"=B{row}")
                ws.cell(row, 4, 0)
                placed.append(spot)
                wiring.append((label_of.get(spot, ""), [spot]))
        if len(picks) > len(slots):
            overflow += [(label_of.get(sp, ""), amount_of.get(sp), name)
                         for sp in picks[len(slots):]]
        # a block whose total was left blank never adds up what we just put in it
        used_rows = [slots[i] for i in range(min(len(picks), len(slots)))]
        if used_rows:
            for tr in range(st, en + 1):
                if not _TOTAL_ROW.match(str(ws.cell(tr, 1).value or "")):
                    continue
                if ws.cell(tr, 2).value in (None, ""):
                    for col, letter in ((2, "B"), (3, "C"), (4, "D")):
                        ws.cell(tr, col, f"=SUM({letter}{used_rows[0]}:{letter}{used_rows[-1]})")
                break

    return {"linked": wiring, "placed": placed, "overflow": overflow,
            "used_rows": set(placed)}


# ── does the W sheet still add up to the sponsor's document? ─────────────────
def reconcile(parsed: dict, index: dict, wiring_spots) -> dict:
    """Two checks per block of the sponsor's document.

    read   — the lines we transcribed, added up, against the total the sponsor
             printed to close that block. A gap means the document was misread.
    wired  — the same lines against what actually reached the W sheet. A gap
             means costs were dropped (no template row) or double-counted."""
    blocks, _names = document_blocks(parsed)
    where = {spot: sec for spot, sec, _l, _a in index.get("items", [])}
    amount = {spot: float(a or 0.0) for spot, _s, _l, a in index.get("items", [])}
    wired = {}
    for spot in wiring_spots:
        name = where.get(spot, "")
        wired[name] = wired.get(name, 0.0) + amount.get(spot, 0.0)

    rows, seen = [], set()
    for b in blocks:
        got = wired.get(b["name"], 0.0)
        seen.add(b["name"])
        rows.append({
            "section": b["name"],
            "printed": b["printed"],
            "read": round(b["items"], 2),
            "on_w_sheet": round(got, 2),
            "read_gap": None if b["printed"] is None else round(b["items"] - b["printed"], 2),
            "wired_gap": round(got - b["items"], 2),
        })
    for name, got in wired.items():        # wired from somewhere with no block
        if name not in seen and got:
            rows.append({"section": name or "(unsectioned)", "printed": None,
                         "read": 0.0, "on_w_sheet": round(got, 2),
                         "read_gap": None, "wired_gap": round(got, 2)})
    return {
        "sections": rows,
        "read_total": round(sum(r["read"] for r in rows), 2),
        "wired_total": round(sum(r["on_w_sheet"] for r in rows), 2),
        "difference": round(sum(r["wired_gap"] for r in rows), 2),
        "misread": [r["section"] for r in rows
                    if r["read_gap"] is not None and abs(r["read_gap"]) >= 1],
    }


# ── Claude reads the finished job back ───────────────────────────────────────
_REVIEW_PROMPT = """A sponsor's cost budget has been read off their document and
wired onto an underwriting template. Check the result and report what looks wrong.

THE SPONSOR'S LINES, as read (section | line | amount):
{source}

HOW THE TEMPLATE WAS WIRED (template row <- the sponsor line(s) it now sums):
{wiring}

TEMPLATE ROWS LEFT EMPTY (no sponsor line was matched to them):
{missing}

SPONSOR LINES NOT USED ANYWHERE:
{unused}

SECTION TIE-OUT (section | printed on the document | lines read | reached the sheet):
{tie}

Look for:
- a cost wired into the wrong category (a hard cost sitting in soft costs)
- a subtotal or section total wired in as if it were a line item, which
  double-counts everything beneath it
- a line that plainly belongs to an empty template row but was left out
- a sponsor line used for a template row that means something different
- a section whose lines don't add to its printed total, where the cause looks
  like a misread number rather than the sponsor nesting other sections inside it
- anything in the unused list large enough that leaving it off materially
  understates the budget

Report only real problems, most serious first. Say what you'd change. If the job
looks right, return an empty list — do not invent work. Judge whether each number
is in the right place, not whether you would have named the rows differently."""

_REVIEW_SCHEMA = {
    "type": "object",
    "properties": {
        "issues": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "where": {"type": "string"},
                    "problem": {"type": "string"},
                    "fix": {"type": "string"},
                    "severity": {"type": "string", "enum": ["high", "medium", "low"]},
                },
                "required": ["where", "problem", "fix", "severity"],
                "additionalProperties": False,
            },
        },
        "verdict": {"type": "string"},
    },
    "required": ["issues", "verdict"],
    "additionalProperties": False,
}


def review(report: dict, parsed: dict) -> dict:
    """A second pass over the finished sheet. Reading a budget is judgement as
    much as transcription, so the result gets read back before anyone relies
    on it."""
    def money(a):
        return "—" if not isinstance(a, (int, float)) else f"{a:,.0f}"
    section, lines = "", []
    for row in parsed.get("rows") or []:
        label = str(row.get("label") or "").strip()
        if row.get("kind") == "section":
            section = label
        elif label and row.get("kind") == "item":
            lines.append(f"{section} | {label} | {money(row.get('amount'))}")
    wiring = "\n".join(
        f"{w} <- " + " + ".join(f"{l} ({money(a)})" for l, a in srcs)
        for w, srcs in report.get("wiring", []))
    tie = "\n".join(
        f"{r['section']} | {money(r['printed'])} | {money(r['read'])} | {money(r['on_w_sheet'])}"
        for r in (report.get("check") or {}).get("sections", []))
    resp = hist_llm._client().messages.create(
        model=hist_llm.MODEL, max_tokens=8000, thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": _REVIEW_PROMPT.format(
            source="\n".join(lines) or "(none)",
            wiring=wiring or "(nothing wired)",
            missing=", ".join(m[0] for m in report.get("missing", [])) or "(none)",
            unused="\n".join(f"{l} | {money(a)}" for l, a in report.get("unused", [])) or "(none)",
            tie=tie or "(none)")}],
        output_config={"format": {"type": "json_schema", "schema": _REVIEW_SCHEMA}},
    )
    return hist_llm._json_response(resp)


def score(report: dict) -> dict:
    """A headline number with its parts shown, because the parts mean different
    things: money that never reached the sheet may be perfectly correct (the
    template has no row for it), while a section that doesn't add up is a real
    reading error."""
    chk = report.get("check") or {}
    read, wired = chk.get("read_total") or 0.0, chk.get("wired_total") or 0.0
    coverage = min(1.0, wired / read) if read else 0.0
    secs = chk.get("sections") or []
    ties = sum(1 for r in secs
               if r["read_gap"] is None or abs(r["read_gap"]) < 1)
    fidelity = (ties / len(secs)) if secs else 1.0
    issues = report.get("review", {}).get("issues", [])
    # money carried is the strongest signal; a section that doesn't add up is
    # often the sponsor nesting one section inside another rather than a misread,
    # so it counts for less. Flagged issues take it down from there.
    weight = {"high": 0.08, "medium": 0.03, "low": 0.005}
    penalty = min(0.25, sum(weight.get(i.get("severity", "low"), 0.005) for i in issues))
    value = max(0.0, min(1.0, 0.50 * coverage + 0.20 * fidelity + 0.30) - penalty)
    return {
        "score": round(value * 100),
        "coverage": round(coverage * 100, 1),
        "fidelity": round(fidelity * 100, 1),
        "issues": len(issues),
        "high": sum(1 for i in issues if i.get("severity") == "high"),
    }


def write_tie_out(ws, report: dict, at_col: int = 18):
    """A tie-out block on the worksheet, clear of the model: every total the
    document printed, what reached the sheet, and the percentage that ties."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    chk = report.get("check") or {}
    rows = chk.get("sections") or []
    if not rows:
        return
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="203864")
    money = '"$"#,##0'
    r = 1
    ws.cell(r, at_col, "TIE-OUT TO THE SPONSOR'S DOCUMENT").font = hdr
    for c in range(at_col, at_col + 5):
        ws.cell(r, c).fill = fill
        ws.cell(r, c).font = hdr
    r += 1
    for i, head in enumerate(["Section", "Printed", "On this sheet", "Difference", "% tied"]):
        cell = ws.cell(r, at_col + i, head)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right" if i else "left")
    r += 1
    first = r
    for row in rows:
        printed = row["printed"] if row["printed"] is not None else row["read"]
        ws.cell(r, at_col, row["section"])
        ws.cell(r, at_col + 1, printed).number_format = money
        ws.cell(r, at_col + 2, row["on_w_sheet"]).number_format = money
        ws.cell(r, at_col + 3, round(row["on_w_sheet"] - printed, 2)).number_format = money
        pct = ws.cell(r, at_col + 4,
                      f"={get_column_letter(at_col + 2)}{r}/{get_column_letter(at_col + 1)}{r}")
        pct.number_format = "0.0%"
        r += 1
    A, B, C, D, E = (get_column_letter(at_col + i) for i in range(5))
    ws.cell(r, at_col, "TOTAL").font = Font(bold=True)
    for col, letter in ((at_col + 1, B), (at_col + 2, C), (at_col + 3, D)):
        cell = ws.cell(r, col, f"=SUM({letter}{first}:{letter}{r - 1})")
        cell.font = Font(bold=True)
        cell.number_format = money
    tot = ws.cell(r, at_col + 4, f"={C}{r}/{B}{r}")
    tot.font = Font(bold=True)
    tot.number_format = "0.0%"
    for i, w in enumerate((34, 16, 16, 16, 10)):
        ws.column_dimensions[get_column_letter(at_col + i)].width = w


# ── the whole job ────────────────────────────────────────────────────────────
def _pick_source_sheet(source_bytes):
    """The sponsor's own worksheet, when the document is a workbook — the busiest
    sheet in it. PDFs have no grid to paste, so they get a transcription."""
    import openpyxl
    from tools import tabular
    if not source_bytes or tabular.kind(source_bytes) not in ("xlsx", "xls"):
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(source_bytes), data_only=True)
    except Exception:  # noqa: BLE001
        return None
    best, score = None, -1
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        n = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ""))
        if n > score:
            best, score = ws, n
    return best


def add_property(wb, template_w: str, template_s: str, new_id: str, parsed: dict,
                 use_llm: bool = True, source_bytes: bytes = None) -> dict:
    """Write the S and W tabs for one property. Returns a report."""
    s_title, w_title = f"S - Cost Budget - {new_id}", f"W - Cost Budget - {new_id}"
    for t in (s_title, w_title):
        if t in wb.sheetnames:
            raise ValueError(f"'{t}' already exists in this workbook")

    src_ws = _pick_source_sheet(source_bytes)
    if src_ws is not None:
        index = locate_items(paste_source_sheet(wb, src_ws, s_title), parsed)
    else:
        index = write_source_tab(wb, template_s, s_title, parsed)
    ws = _clone(wb, wb[template_w], w_title)
    report = fill_from_source(ws, s_title, index, use_llm=use_llm)
    report["cleared"] = _blank_qcp_blocks(ws) + clear_notes(ws)
    report["zeroed"] = _zero_hardcoded(ws)

    # spell the wiring out for review: which source line(s) each row now reads
    by_spot = {spot: (lbl, amt) for spot, _sec, lbl, amt in index.get("items", [])}
    report["wiring"] = [(w_label, [by_spot.get(spot, (str(spot), None)) for spot in picks])
                        for w_label, picks in report["linked"]]

    items = [(str(i.get("label")).strip(), i.get("amount"))
             for i in (parsed.get("rows") or []) if i.get("kind") == "item"]
    report["missing"] = report.get("overflow", [])
    report["unused"] = [(lbl, a) for spot, _sec, lbl, a in index.get("items", [])
                        if spot not in report["used_rows"]]
    report["check"] = reconcile(parsed, index,
                                [spot for _w, picks in report["linked"] for spot in picks])
    write_tie_out(ws, report)
    if use_llm and hist_llm.available():
        try:
            report["review"] = review(report, parsed)
        except Exception as e:  # noqa: BLE001
            report["review"] = {"issues": [], "verdict": f"review didn't run: {e}"}
    report["score"] = score(report)
    report.update(sheets=(w_title, s_title), source_items=len(items))
    return report


def build_workbook(template_bytes: bytes, template_w: str, template_s: str,
                   new_id: str, parsed: dict, use_llm: bool = True,
                   source_bytes: bytes = None):
    """One property in, one workbook out — holding just its two tabs. The clone
    happens inside the template workbook, because that is what carries the styles
    and the logo across; the other properties are dropped before saving. Nothing
    on the W sheet points outside its own pair, so nothing breaks."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    # the template may already hold a pair with this name (a rebuild) — work
    # under a scratch name, then take the real one once the others are gone
    scratch = new_id
    while f"W - Cost Budget - {scratch}" in wb.sheetnames or \
            f"S - Cost Budget - {scratch}" in wb.sheetnames:
        scratch += "~"
    report = add_property(wb, template_w, template_s, scratch, parsed,
                          use_llm=use_llm, source_bytes=source_bytes)
    w_tmp, s_tmp = report["sheets"]
    for name in list(wb.sheetnames):
        if name not in (w_tmp, s_tmp):
            del wb[name]
    w_title, s_title = f"W - Cost Budget - {new_id}", f"S - Cost Budget - {new_id}"
    if scratch != new_id:
        for ws_ in wb.worksheets:                  # rename, and repoint the links
            for row in ws_.iter_rows():
                for c in row:
                    if _is_formula(c.value) and s_tmp in c.value:
                        c.value = c.value.replace(s_tmp, s_title)
        wb[w_tmp].title, wb[s_tmp].title = w_title, s_title
        report["sheets"] = (w_title, s_title)
    wb._sheets = [wb[w_title], wb[s_title]]     # worksheet first, source behind it
    wb.active = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), report


# ── UI ───────────────────────────────────────────────────────────────────────
def render():
    import openpyxl
    import pandas as pd
    import streamlit as st
    from tools import tabular

    st.header("🧱   Cost Budget")
    st.caption("One property at a time. Upload a cost-basis workbook to take the "
               "format from, plus the sponsor's cost budget for the new property — you "
               "get back a **two-tab workbook**: a **W** worksheet whose sponsor blocks "
               "are the sponsor's own line items, each wired to the cell it came from, "
               "and an **S** tab holding their document exactly as they sent it. The "
               "QCP side keeps only its headings and subtotal formulas; notes start "
               "empty.")

    if not hist_llm.available():
        st.warning("Reading sponsor documents needs **ANTHROPIC_API_KEY** in the app "
                   "secrets — the budgets arrive as scans and one-off spreadsheets, so "
                   "there's no fixed parser for them.")

    book = st.file_uploader("Cost-basis workbook (.xlsx)", type=["xlsx", "xlsm"],
                            key="cb_book")
    if not book:
        st.info("Upload the workbook to begin.")
        return
    data = book.getvalue()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data))
    except Exception as e:  # noqa: BLE001
        st.error(f"Could not open that workbook: {e}")
        return
    found = pairs(wb)
    if not found:
        st.error("No `W - Cost Budget - …` / `S - Cost Budget - …` tab pairs in this "
                 "workbook — this tool clones an existing property's pair.")
        return

    st.subheader("1 · Template")
    pick = st.selectbox("Clone this property's tabs", [p[0] for p in found],
                        format_func=lambda i: f"{i}  ({dict((p[0], p[1]) for p in found)[i]})",
                        key="cb_tmpl")
    tmpl_w, tmpl_s = next((w, s) for i, w, s in found if i == pick)
    st.caption(f"Cloning `{tmpl_w}` + `{tmpl_s}` — "
               f"{len(blocks(wb[tmpl_w]))} sections, "
               f"{sum(1 for b in blocks(wb[tmpl_w]) if b[3] == 'SPONSOR')} sponsor-side.")

    st.subheader("2 · The sponsor's cost budget")
    src = st.file_uploader("Cost basis / cost budget (PDF, Excel or CSV)",
                           type=tabular.UPLOAD_TYPES, key="cb_src")
    if not src:
        st.info("Upload the sponsor's document for the new property.")
        return

    cache = st.session_state.setdefault("cb_parsed", {})
    key = f"{src.name}:{src.size}"
    if key not in cache:
        with st.spinner(f"Reading {src.name}…"):
            try:
                cache[key] = extract_source(src.getvalue(),
                                            sponsor_categories(wb, tmpl_w))
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not read that document: {e}")
                return
    parsed = cache[key]

    items = [r for r in parsed.get("rows") or [] if r.get("kind") == "item"]
    st.success(f"Read {len(parsed.get('rows') or [])} lines — {len(items)} cost items"
               + (f" · {parsed['property_name']}" if parsed.get("property_name") else ""))
    with st.expander("Check the transcription against the document", expanded=False):
        st.dataframe(pd.DataFrame([{"": r.get("kind"), "Line": r.get("label"),
                                    "Amount": r.get("amount")}
                                   for r in parsed.get("rows") or []]),
                     use_container_width=True, hide_index=True, height=380)

    default_id = re.sub(r"[^0-9A-Za-z ]+", "", str(parsed.get("property_name") or "")).split()
    new_id = st.text_input("New property tab name", value=(default_id[0] if default_id else ""),
                           key="cb_id", help="Tabs are named 'W - Cost Budget - <this>' "
                                             "and 'S - Cost Budget - <this>'.")
    if not new_id.strip():
        st.info("Give the property a short name for its tabs.")
        return

    st.subheader("3 · Build")
    if st.button("🧱   Build this property", type="primary"):
        try:
            out, rep = build_workbook(data, tmpl_w, tmpl_s, new_id.strip(), parsed,
                                      use_llm=hist_llm.available(),
                                      source_bytes=src.getvalue())
        except Exception as e:  # noqa: BLE001
            st.error(f"Build failed: {e}")
            return
        st.session_state["cb_out"] = (out, rep, book.name)

    if "cb_out" in st.session_state:
        out, rep, orig = st.session_state["cb_out"]
        w_title, s_title = rep["sheets"]
        st.success(f"Built a two-tab workbook — `{w_title}` and `{s_title}`. "
                   f"{len(rep['linked'])} sponsor row(s) wired to the source; "
                   f"everything outside the sponsor's side cleared "
                   f"({rep['cleared'] + len(rep['zeroed'])} cells).")
        sc = rep.get("score") or {}
        if sc:
            a, b, c, d = st.columns(4)
            a.metric("Accuracy", f"{sc['score']}%")
            b.metric("Money carried across", f"{sc['coverage']}%",
                     help="Of what we read from the document, how much reached the "
                          "worksheet. Short of 100% is often correct — the template "
                          "has no row for some of the sponsor's lines.")
            c.metric("Sections that add up", f"{sc['fidelity']}%",
                     help="Sections whose lines sum to the total printed beside them "
                          "on the document. Short of 100% points at a misread number.")
            d.metric("Issues found", sc["issues"],
                     delta=None if not sc["high"] else f"{sc['high']} serious",
                     delta_color="inverse")

        rev = rep.get("review") or {}
        if rev.get("issues"):
            st.warning(f"Claude read the finished sheet back and flagged "
                       f"{len(rev['issues'])} thing(s):")
            st.dataframe(pd.DataFrame([{"Severity": i["severity"], "Where": i["where"],
                                        "Problem": i["problem"], "Suggested fix": i["fix"]}
                                       for i in rev["issues"]]),
                         use_container_width=True, hide_index=True)
        elif rev.get("verdict"):
            st.success("Claude read the finished sheet back: " + rev["verdict"])

        chk = rep.get("check") or {}
        if chk.get("sections"):
            st.markdown("**Totals check** — the sponsor's sections against what reached "
                        "the worksheet.")
            st.dataframe(pd.DataFrame([{
                "Section": r["section"],
                "Printed on the doc": r["printed"],
                "Lines we read": r["read"],
                "On the W sheet": r["on_w_sheet"],
                "Reads": "—" if r["read_gap"] is None else
                         ("ok" if abs(r["read_gap"]) < 1 else f"off by {r['read_gap']:,.0f}"),
                "Wired": "ok" if abs(r["wired_gap"]) < 1 else f"{r['wired_gap']:,.0f}",
            } for r in chk["sections"]]), use_container_width=True, hide_index=True)
            gap = chk.get("difference", 0)
            msg = (f"Read {chk['read_total']:,.0f} from the document; "
                   f"{chk['wired_total']:,.0f} reached the worksheet.")
            (st.success if abs(gap) < 1 else st.warning)(
                msg + ("" if abs(gap) < 1 else
                       f" {abs(gap):,.0f} is sitting on the source tab with no row to go to."))
            if chk.get("misread"):
                st.caption("Sections whose lines don't add to the total printed beside them: "
                           + ", ".join(chk["misread"]) +
                           " — usually the sponsor rolling other sections up inside them, "
                           "but worth opening the source tab to confirm.")

        if rep.get("wiring"):
            with st.expander(f"How the {len(rep['wiring'])} row(s) were wired — worth a look",
                             expanded=True):
                st.dataframe(pd.DataFrame(
                    [{"Template row": w,
                      "Reads from": " + ".join(lbl for lbl, _a in srcs),
                      "Amount": sum(a for _l, a in srcs if isinstance(a, (int, float)))}
                     for w, srcs in rep["wiring"]],
                ), use_container_width=True, hide_index=True)

        c1, c2 = st.columns(2)
        with c1:
            if rep["missing"]:
                st.warning(f"{len(rep['missing'])} template row(s) the sponsor's document "
                           "doesn't cover — left at zero:")
                st.dataframe(pd.DataFrame(rep["missing"], columns=["Template row", "Row"]),
                             use_container_width=True, hide_index=True)
            else:
                st.info("Every template row found a match in the source.")
        with c2:
            if rep["unused"]:
                st.warning(f"{len(rep['unused'])} line(s) in the sponsor's document with no "
                           "row in the template — **not** on the W sheet; they're on the "
                           "S tab if you want to place them by hand:")
                st.dataframe(pd.DataFrame(rep["unused"], columns=["Source line", "Amount"]),
                             use_container_width=True, hide_index=True)
            else:
                st.info("Every source line landed on a template row.")
        if rep["zeroed"]:
            with st.expander(f"{len(rep['zeroed'])} inherited number(s) zeroed"):
                st.code("\n".join(rep["zeroed"][:80]))
        st.download_button("⬇   Download workbook (.xlsx)", data=out,
                           file_name=f"Cost Budget - {new_id.strip()}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
